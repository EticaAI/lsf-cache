# ==============================================================================
#
#          FILE:  L999999999_0.py
#
#         USAGE:  It's a python library
#                    from L999999999_0 import (
#                        hxltm_carricato,
#                        qhxl_hxlhashtag_2_bcp47,
#                        # (...)
#                    )
#
#   DESCRIPTION:  Common library for 999999999_*.py cli scripts
#
#       OPTIONS:  ---
#
#  REQUIREMENTS:  - python3
#                   - pip install openpyxl pyyaml
#          BUGS:  ---
#         NOTES:  ---
#       AUTHORS:  Emerson Rocha <rocha[at]ieee.org>
# COLLABORATORS:  ---
#       COMPANY:  EticaAI
#       LICENSE:  Public Domain dedication or Zero-Clause BSD
#                 SPDX-License-Identifier: Unlicense OR 0BSD
#       VERSION:  v1.0.0
#       CREATED:  2022-05-18 19:52 UTC based on 999999999_10263485.py
#      REVISION:  ---
# ==============================================================================
"""Common library for 999999999_*.py cli scripts
"""

# pytest
#    python3 -m doctest ./999999999/0/L999999999_0.py


import csv
from functools import reduce
from genericpath import exists
import json
# from multiprocessing.sharedctypes import Value
# import importlib
import os
from pathlib import Path
# from pathlib import Path
import re
import sys
from datetime import date, datetime
from typing import (
    Any,
    Iterator,
    List,
    Tuple,
    Type,
    Union
)
import uuid

# Allow fail if user does not have, but just using part of the tools
# import yaml
try:
    import yaml
except ModuleNotFoundError:
    # Error handling
    pass

# Allow fail if user does not have, but just using part of the tools
# from openpyxl import (
#     load_workbook
# )
try:
    from openpyxl import (
        load_workbook
    )
except ModuleNotFoundError:
    # Error handling
    pass


# pylint --disable=W0511,C0103,C0302,C0116 ./999999999/0/L999999999_0.py

# SYSTEMA_SARCINAE = str(Path(__file__).parent.resolve())
# PROGRAMMA_SARCINAE = str(Path().resolve())
# ARCHIVUM_CONFIGURATIONI_L999999999_0 = [
#     PROGRAMMA_SARCINAE + '/L999999999_0.meta.yml',
#     SYSTEMA_SARCINAE + '/L999999999_0.meta.yml',
# ]
# ARCHIVUM_CONFIGURATIONI_DEFALLO = [
#     # SYSTEMA_SARCINAE + '/' + NOMEN + '.meta.yml',
#     # PROGRAMMA_SARCINAE + '/' + NOMEN + '.meta.yml',
#     PROGRAMMA_SARCINAE + '/{{PROGRAMMA}}.meta.yml',
#     SYSTEMA_SARCINAE + '/{{PROGRAMMA}}.meta.yml',
# ]
NUMERORDINATIO_BASIM = os.getenv('NUMERORDINATIO_BASIM', os.getcwd())
NUMERORDINATIO_DEFALLO = int(os.getenv('NUMERORDINATIO_DEFALLO', '60'))  # �
NUMERORDINATIO_MISSING = "�"
VENANDUM_INSECTUM = bool(os.getenv('VENANDUM_INSECTUM', ''))

EXIT_OK = 0
EXIT_ERROR = 1
EXIT_SYNTAX = 2

# def _dictionaria_linguarum():
#     # importlib.import_module('', package=None)
#     L1603_1_51 = importlib.import_module('1603_1_51', package='1603_1_1')

#     return L1603_1_51.DictionariaLinguarum()

BCP47_LANGTAG_CALLBACKS = {
    'hxl_attrs': lambda lmeta, strictum: bcp47_langtag_callback_hxl(
        lmeta, strictum=strictum),
    'hxl_minimal': lambda lmeta, strictum: bcp47_langtag_callback_hxl_minimal(
        lmeta, strictum=strictum)
}

BCP47_LANGTAG_EXTENSIONS = {
    'r': lambda r, strictum: bcp47_rdf_extension(r, strictum=strictum)
}

# BCP47_EX_HXL = {
#     'qcc-Zxxx-r-aMDCIII-alatcodicem-anop':
#     '#item+conceptum+codicem',
#     'qcc-Zxxx-r-aMDCIII-alatnumerordinatio-anop-sU2200-s1603-snop':
#     '#item+conceptum+numerordinatio',
# }
BCP47_EX_HXL = {
    '#item+conceptum+codicem': {
        'bcp47': 'qcc-Zxxx-r-aMDCIII-alatcodicem-anop',
        'hxl': '#item+rem+i_qcc+is_zxxx+rdf_a_mdciii_latcodicem',
        'hxltm': '#item+conceptum+codicem'
    },
    '#item+conceptum+numerordinatio': {
        'bcp47': 'qcc-Zxxx-r-aMDCIII-alatnumerordinatio-anop-sU2200-s1603-snop',
        'hxl': '#item+rem+i_qcc+is_zxxx'
        '+rdf_a_mdciii_latnumerordinatio+rdf_s_u2200_s1603',
        'hxltm': '#item+conceptum+numerordinatio'
    },
    # # Make sure secondary roundtrip works
    # '#item+rem+i_qcc+is_zxxx+rdf_a_mdciii_latnumerordinatio+rdf_s_u2200_s1603': {
    #     'bcp47': 'qcc-Zxxx-r-aMDCIII-alatnumerordinatio-anop-sU2200-s1603-snop',
    #     'hxl': '#item+rem+i_qcc+is_zxxx'
    #     '+rdf_a_mdciii_latnumerordinatio+rdf_s_u2200_s1603',
    #     'hxltm': '#item+conceptum+numerordinatio'
    # }
}

BCP47_AD_HXL = {
    'qcc-Zxxx-r-aMDCIII-alatcodicem-anop':
    BCP47_EX_HXL['#item+conceptum+codicem'],
    'qcc-Zxxx-r-aMDCIII-alatnumerordinatio-anop-sU2200-s1603-snop':
    BCP47_EX_HXL['#item+conceptum+numerordinatio']
}

# @TODO allow non hardcoded CSV_SEPARATORS
# Hacky way to have inline cell separators.
CSVW_SEPARATORS = {
    # HXL: +rdf_ycsvwseparator_u007c, BCP47, (...)-r-yCSVWseparator-u007c
    'u007c': '|',
    'u007cu007c': '||',
    'u0020': ' ',
    'u003b': ';',
    'u0009': '	'  # tab
}

# @see https://www.asciitable.com/

EXTRA_OPERATORS = {
    # Used for @prefix
    'STX': {
        'eng-Latn': '(start of text)',
        # 'hxl': 'u02',
        'hxl': 'u0002',
        'wikidata': 'Q10366650',  # https://www.wikidata.org/wiki/Q10366650
        # 'unicode': 'U+0002',
        'unicode': u"\x02"
    },
    # Used to explain what separator the cell value may use
    'GS': {
        'eng-Latn': '(group separator)',  # Also: information separator three
        'hxl': 'u001d',
        # 'hxl': 'u001d',
        'wikidata': 'Q10366650',  # https://www.wikidata.org/wiki/Q110028713
        # 'unicode': 'U+001D',
        'unicode': u"\x1D",
    }
}

# @see https://en.wikipedia.org/wiki/List_of_logic_symbols
FIRST_ORDER_LOGIC = {
    '∀': {
        'python': u'\u2200',
        'eng-Latn': 'universal quantification',
        'hxl': 'u2200',
        'wdata': 'Q126695',  # https://www.wikidata.org/wiki/Q126695
        'unicode': 'U+2200'
    },
    '∃': {
        'python': u'\u2203',
        'eng-Latn': 'existential quantification',
        'hxl': 'u2203',
        'wdata': 'Q773483',  # https://www.wikidata.org/wiki/Q773483
        'unicode': 'U+2203'
    },
    # print(u'{0}'.format("\U0001D53B"))
    # print(u'{0}'.format(u"\U0001D53B"))
    '𝔻': {
        'python': u"\U0001D53B",
        'eng-Latn': 'domain of discourse',
        'hxl': 'u0001D53B',
        'wdata': 'Q1228944',  # https://www.wikidata.org/wiki/Q1228944
        'unicode': 'U+1D53B'
    },
}

# @see https://hxlstandard.org/standard/1-1final/dictionary/#classification
# @see https://www.wikidata.org/wiki/Wikidata:List_of_properties
# @see https://docs.google.com/spreadsheets
#      /d/1NjSI2LaS3SqbgYc0HdD8oIb7lofGtiHgoKKATCpwVdY/edit#gid=1088874596
# @see https://www.wikidata.org/wiki/Wikidata:List_of_properties/name
# prefix 'wdata' = ttp://www.wikidata.org/wiki/Special:EntityData/
HXL_ATTRIBUTES_AD_RDF = {
    'geo': {
        # Aliases
        '__alia__': {
            '+name+alt1': '+name+alt',
            '+name+alt2': '+name+alt',
            '+v_iso2': '+code+v_iso3166p1a2',
            '+v_iso3': '+code+v_iso3166p1a3',
        },
        '+code+v_numerodinatio': {
            'numerodinatio_est': True
        },
        '+name+v_unterm': {
            # P1448: official name of the subject in its official language(s)
            'wdata': 'P1448',  # official name
            'rdftypisegolinguis': [
                # 'skos:BFO_0000029',
                'skos:prefLabel'
            ]
        },
        '+name+alt': {
            'rdf:type': [
                'skos:altLabel'
            ]
        },
        # '+name+alt1': {
        #     '___': '+name+alt'
        # },
        # '+name+alt2': {
        #     '___': '+name+alt'
        # },
        # '+name+alt1': HXL_ATTRIBUTES_AD_RDF['geo']['+name+alt'],
        '+name': {
            # P1448: short name of a place, organisation, person, journal,
            #        Wikidata property, etc.
            'wdata': 'P1813',  # short name
            'rdftypisegolinguis': [
                # 'skos:BFO_0000029',
                'skos:prefLabel'
            ]
        },
        # [+code / +v_pcode] varies by context
        # '+code': {}
        '+code+v_m49': {
            'wdata': 'P2082'  # United Nations M.49 code for the subject item
        },
        '+code+v_iso3166p1a2': {
            'wdata': 'P297'  # ISO 3166-1 alpha-2 code
        },
        '+code+v_iso3166p1a3': {
            'wdata': 'P298'  # ISO 3166-1 alpha-3 code
        },
        # Consider using UN m49 code instead of the ISO one
        '+code+v_iso3166p1n': {
            'wdata': 'P299'  # ISO 3166-1 numeric code
        },
        '+code+v_iso3166p2': {
            'wdata': 'P300'  # subdivision code ISO 3166-2
        },
    },
    # Note: avoid use generic
    'zzzgeneric': {
        '+name': {
            # name; name the subject is known by.
            # If a more specific property is available, use that
            'wdata': 'P2561'
        }
    },
}

# HXL_ATTRIBUTES_AD_RDF['geo']['+name+alt1'] = \
#     HXL_ATTRIBUTES_AD_RDF['geo']['+name+alt']
# HXL_ATTRIBUTES_AD_RDF['geo']['+name+alt2'] = \
#     HXL_ATTRIBUTES_AD_RDF['geo']['+name+alt']

# wdtaxonomy Q6256 -P P131
HXL_HASHTAGS_AD_RDF = {
    '#country': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},  # Requires processing ifer __hxlattrs better value
        'wdata': 'Q6256',  # country
        'rdftrivio': '5000',
        'rdftypis': {
            '#adm1': [
                'obo:BFO_0000170'  # BFO_0000170: location of at all times
            ]
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    # # Not a valid HXL hashtag, but using anyway as alias to country
    # '#adm0': {
    #     'hxlattrs': HXL_ATTRIBUTES_AD_RDF['geo'],
    #     'wdata': 'Q6256',  # country
    #     'rdftrivio': '5000',
    #     'rdf:type': [
    #         'obo:BFO_0000029'
    #     ]
    # },
    '#adm1': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q10864048',  # first-level administrative country subdivision
        'rdftrivio': '5001',
        'rdftypis': {
            '#adm2': [
                'obo:BFO_0000082'  # BFO_0000082: located in at all times
            ],
            '#country': [
                'obo:BFO_0000170'  # BFO_0000170: location of at all times
            ],
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    '#adm2': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q13220204',  # second-level administrative country subdivision
        'rdftrivio': '5002',
        'rdftypis': {
            '#adm1': [
                'obo:BFO_0000082'
            ],
            '#adm3': [
                'obo:BFO_0000170'
            ],
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    '#adm3': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q13221722',  # third-level administrative country subdivision
        'rdftrivio': '5003',
        'rdftypis': {
            '#adm2': [
                'obo:BFO_0000082'
            ],
            '#adm4': [
                'obo:BFO_0000170'
            ],
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    '#adm4': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q14757767',  # fourth-level administrative country subdivision
        'rdftrivio': '5004',
        'rdftypis': {
            '#adm3': [
                'obo:BFO_0000082'
            ],
            '#adm5': [
                'obo:BFO_0000170'
            ],
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    '#adm5': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q15640612',  # fifth-level administrative country subdivision
        'rdftrivio': '5005',
        'rdftypis': {
            '#adm4': [
                'obo:BFO_0000082'
            ],
            '#adm6': [
                'obo:BFO_0000170'
            ],
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
    '#adm6': {
        '__hxlattrs': 'geo',
        'hxlattrs': {},
        'wdata': 'Q22927291',  # sixth-level administrative country subdivision
        'rdftrivio': '5006',
        'rdftypis': {
            '#adm5': [
                'obo:BFO_0000082'
            ]
        },
        'rdftypisego': [
            'obo:BFO_0000029'
        ]
    },
}

# @TODO this part is somewhat hardcoded
HXL_HASH_ET_ATTRIBUTA_AD_RDF = {
    # @TODO
    #      - #country+code+v_iso3166p1a2
    '#country+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5000-snop-pOBO-pbfo124-ps5001',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5001+rdf_s_u2200_s5000'
    },
    # P297 ISO 3166-1 alpha-2 code, https://www.wikidata.org/wiki/Property:P297
    '#country+code+v_iso3166p1a2': {
        '__no1bpc47__': 'qcc-Zxxx-r-pP-pp297-ps5000-x-p297',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_p297+rdf_p_p_p297_s5000'
    },
    '#adm1+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5001-snop-pOBO-pbfo124-ps5002-pOBO-pbfo171-ps5000',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5002+rdf_p_obo_bfo171_s5000+rdf_s_u2200_s5001'
    },
    # @TODO make P-Codes using "instance of (P31)" like > p:P31 Q7200235
    #       (change for this TODO) drafted +v_pcode's
    # @see https://en.wikipedia.org/wiki/Place_code
    # @see https://www.wikidata.org/wiki/Q7200235 Q7200235
    '#adm1+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5001-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5001'
    },
    '#adm2+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5002-snop-pOBO-pbfo124-ps5002-pOBO-pbfo171-ps5001',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5002+rdf_p_obo_bfo171_s5001+rdf_s_u2200_s5002'
    },
    '#adm2+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5002-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5002'
    },
    '#adm3+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5003-snop-pOBO-pbfo124-ps5003-pOBO-pbfo171-ps5002',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5003+rdf_p_obo_bfo171_s5002+rdf_s_u2200_s5003'
    },
    '#adm3+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5003-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5003'
    },
    '#adm4+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5004-snop-pOBO-pbfo124-ps5004-pOBO-pbfo171-ps5003',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5004+rdf_p_obo_bfo171_s5003+rdf_s_u2200_s5004'
    },
    '#adm4+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5004-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5004'
    },
    '#adm5+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5005-snop-pOBO-pbfo124-ps5005-pOBO-pbfo171-ps5004',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5005+rdf_p_obo_bfo171_s5004+rdf_s_u2200_s5005'
    },
    '#adm5+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5005-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5005'
    },
    '#adm6+code+v_numerodinatio': {
        '__no1bpc47__': 'qcc-Zxxx-r-aOBO-abfo29-anop-sU2200-s5006-snop-pOBO-pbfo124-ps5005-pOBO-pbfo171-ps5004',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+rdf_a_obo_bfo29+rdf_p_obo_bfo124_s5005+rdf_p_obo_bfo171_s5004+rdf_s_u2200_s5005'
    },
    '#adm6+code+v_pcode': {
        '__no1bpc47__': 'qcc-Zxxx-r-pWDATA-pq7200235-ps5006-x-wdataq7200235',
        '__no1hxl__': '#item+rem+i_qcc+is_zxxx+ix_wdataq7200235+rdf_p_wdata_q7200235_s5006'
    },
    # @see https://www.wikidata.org/wiki/EntitySchema:E49
    # publication date (P577)
    # date or point in time when a work was first published or released
    # https://www.wikidata.org/wiki/Property:P577
    '#date+start': {
        '__no1bpc47__': 'qcc-Zxxx-r-pP-pp577-ps1603-tXSD-tdatetime-tnop',
        '__no1hxl__':
        '#item+rem+i_qcc+is_zxxx+rdf_p_p_p577_s1603+rdf_t_xsd_datetime'
    },
    # discontinued date (P2669)
    # date that the availability of a product was discontinued;
    # see also "dissolved, abolished or demolished" (P576)
    # https://www.wikidata.org/wiki/Property:P2669
    '#date+end': {
        '__no1bpc47__': 'qcc-Zxxx-r-pP-pp2699-ps1603-tXSD-tdatetime-tnop',
        '__no1hxl__':
        '#item+rem+i_qcc+is_zxxx+rdf_p_p_p2699_s1603+rdf_t_xsd_datetime'
    },
    ## retrieved (P813)
    # - https://www.wikidata.org/wiki/Property:P813
    # - https://wiki.openstreetmap.org/wiki/Key:check_date
    '#date+updated': {
        '__no1bpc47__': 'qcc-Zxxx-r-pP-pp813-ps1603-tXSD-tdatetime-tnop',
        '__no1hxl__':
        '#item+rem+i_qcc+is_zxxx+rdf_p_p_p813_s1603+rdf_t_xsd_datetime'
    },
}

HXL_WDATA = [
    # hxl_v based on
    # - https://data.humdata.org/dataset/hxl-master-vocabulary-list
    # - https://data.humdata.org/dataset/countries-and-territories
    #
    # fūnctiō/relātiō tags:
    # - abl-0: administrative boundary level 0
    # - abl-1: administrative boundary level 1
    # - abl-0123456: administrative boundary level any
    # - psl-t-0: Population Statistics by administrative boundary level 0
    # - psl-t-1: Population Statistics by administrative boundary level 1
    # - psl-t-0123456: Population Statistics by administrative boundary l. any
    {
        # Ontological meaning for fūnctiō: reason to exist; focus
        '_function': ['abl-0'],
        # Ontology meaning for relātiō: broader relation, can be used for this
        '_relatio': [],
        'hxl_ix': 'ix_iso3166p1a2',
        'hxl_v': 'v_iso2',
        'rdf_datatype': None,
        'wdata_p': 'P297',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0'],
        '_relatio': [],
        'hxl_ix': 'ix_iso3166p1a3',
        'hxl_v': 'v_iso3',
        'rdf_datatype': None,
        'wdata_p': 'P298',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0'],
        '_relatio': [],
        'hxl_ix': 'ix_iso3166p1n',
        'hxl_v': None,
        'rdf_datatype': None,
        'wdata_p': 'P299',
        'wdata_q': None,
    },
    {
        # The function (reason to exist) of  ISO 3166-2 is subdivision, adm1.
        # However, in some circustances, places can have codes as if they're
        # adm0; this means ISO 3166-2 have a relatio with adm0
        '_function': ['abl-1'],
        '_relatio': ['abl-0'],
        'hxl_ix': 'ix_iso3166p2',
        'hxl_v': 'v_iso3',
        'rdf_datatype': None,
        'wdata_p': 'P300',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': ['abl-0'],  # weak relation: "countries can have currency"
        'hxl_ix': 'ix_iso4217',
        'hxl_v': 'v_currency',
        'iri': None,
        'rdf_datatype': None,
        'wdata_p': 'P498',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_iso15924a4',
        'hxl_v': None,
        'iri': 'http://www.unicode.org/iso15924/iso15924-codes.html',
        'rdf_datatype': None,
        'wdata_p': 'P506',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_iso15924n3',
        'hxl_v': None,
        'iri': 'http://www.unicode.org/iso15924/iso15924-codes.html',
        'rdf_datatype': None,
        'wdata_p': 'P2620',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        'hxl_ix': 'ix_jpgeolod',
        'hxl_v': None,
        'iri': 'https://geolod.ex.nii.ac.jp',
        'rdf_datatype': None,
        'wdata_p': 'P5400',
        'wdata_q': None,
    },
    {
        '_function': [],
        # broader relation: "AGROVOC can have countries"
        '_relatio': ['abl-0'],
        'hxl_ix': 'ix_unagrovoc',
        'hxl_v': None,
        'iri': 'https://agrovoc.fao.org',
        'rdf_datatype': None,
        'wdata_p': 'P8061',
        'wdata_q': None,
    },
    {
        '_function': [],
        # broader relation: "AGROVOC can have countries"
        '_relatio': ['abl-0'],
        'hxl_ix': 'ix_unescothes',
        'hxl_v': None,
        'iri': 'https://vocabularies.unesco.org',
        'rdf_datatype': None,
        'wdata_p': 'P3916',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0'],
        '_relatio': [],
        'hxl_ix': 'ix_unm49',
        'hxl_v': 'v_m49',
        'iri': None,
        'rdf_datatype': None,
        'wdata_p': 'P2082',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        'hxl_ix': 'ix_unp',  # using 'unp', ommiting 'code' / 'id' from all
        'hxl_v': 'v_pcode',
        'iri': None,
        'rdf_datatype': None,
        'wdata_p': None,
        'wdata_q': None,
    },
    {
        '_function': ['abl-0'],
        '_relatio': [],
        'hxl_ix': 'ix_usfactbook',
        'hxl_v': None,
        'iri': 'https://www.cia.gov/the-world-factbook/countries',
        'rdf_datatype': None,
        'wdata_p': 'P9948',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': ['abl-0123456'],  # WordNet can have several Adm levels
                                      # But is not main function of Wordnet be
                                      # only about administrativel levels.
        'hxl_ix': 'ix_usworldnet',
        'hxl_v': None,
        'iri': 'http://wordnet-rdf.princeton.edu/',
        'rdf_datatype': None,
        'wdata_p': 'P8814',
        'wdata_q': None,
    },
    {
        '_function': ['abl-1'],
        '_relatio': [],
        'hxl_ix': 'ix_xzhasc',
        'hxl_v': None,
        # http://www.statoids.com/wab.html
        'iri': 'http://www.statoids.com/ihasc.html',
        'rdf_datatype': None,
        'wdata_p': 'P8119',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        'hxl_ix': 'ix_xzgeonames',
        'hxl_v': None,
        'iri': 'https://www.geonames.org',
        'rdf_datatype': None,
        'wdata_p': 'P1566',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_xzglide',
        'hxl_v': 'v_glide',
        'iri': 'https://glidenumber.net',
        'rdf_datatype': None,
        'wdata_p': None,
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': ['abl-0'],  # GitHub topics _can_ have adm0 (but not focus)
        'hxl_ix': 'ix_xzgithubt',
        'hxl_v': None,
        'iri': 'https://github.com/topics',
        'rdf_datatype': None,
        'wdata_p': 'P9100',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_xzglotto',
        'hxl_v': 'v_glotto',  # Proposed as 2022-07-13
        'iri': 'https://glottolog.org',
        'rdf_datatype': None,
        'wdata_p': 'P1394',
        'wdata_q': None,
    },
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        'hxl_ix': 'ix_xzosmrel',  # ommiting 'code' / 'id' from all
        'hxl_v': None,
        'iri': 'https://wiki.openstreetmap.org/wiki/Relation',
        'rdf_datatype': None,
        'wdata_p': 'P402',
        'wdata_q': None,
    },
    # administrative territorial entity of a specific level (Q1799794)
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        'hxl_ix': 'ix_zzcodablevel',
        'hxl_v': None,
        'iri': 'https://www.wikidata.org/wiki/Q1799794',
        'rdf_datatype': None,
        'wdata_p': None,
        'wdata_q': 'Q1799794',
    },
    {
        '_function': ['abl-0123456'],
        '_relatio': [],
        # ZZ prefix (differenaited from XY) inspired by UNICODE CLDR would means
        # "Unknown or Invalid Territory"
        'hxl_ix': 'ix_zzgeojson',
        'hxl_v': None,
        'iri': 'https://www.wikidata.org/wiki/Property:P3896',
        'rdf_datatype': None,
        'wdata_p': 'P3896',
        'wdata_q': None,
    },
    # @TODO think about some "default" type in textual format
    #       https://www.wikidata.org/wiki/Special:EntityData/Q155.ttl
    #       @prefix geo: <http://www.opengis.net/ont/geosparql#> .
    {
        '_function': [],
        '_relatio': ['abl-0123456'],  # Any adm level could have a point
                                      # Example for Brazil: Point(-53.0 -14.0)
        'hxl_ix': 'ix_zzwgs84point',
        'hxl_v': None,
        'iri': 'https://www.wikidata.org/wiki/Property:P625',
        # Example: wdt:P625 "Point(-53 -14)"^^geo:wktLiteral ;
        'rdf_datatype': 'geo:wktLiteral',
        'wdata_p': 'P625',
        'wdata_q': None,
    },
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_zzwikiq',
        'hxl_v': None,
        'iri': 'https://www.wikidata.org/wiki/Q43649390',
        'rdf_datatype': None,
        'wdata_p': None,
        'wdata_q': 'Q43649390',
    },
    # +ix_wikiq is same as +ix_zzwikiq, maybe deal with this inconsistency
    # of prefixes of ix_wikiq toghetuer with ix_iso*
    {
        '_function': [],
        '_relatio': [],
        'hxl_ix': 'ix_wikiq',
        'hxl_v': None,
        'iri': 'https://www.wikidata.org/wiki/Q43649390',
        'rdf_datatype': None,
        'wdata_p': None,
        'wdata_q': 'Q43649390',
    },
]

# Hotfixes for used ix_, but not as fully documented
HXL_WDATA__META_IX = [
    # 'ix_zzcodablevel',
    # 'ix_wikiq'  # @TODO this could be improved to wdata:{value}
    # @TODO implement ix_wikiq as https://www.wikidata.org/wiki/Q43649390
    #       Example: for ab 0: +rdf_p_wdata_q43649390_s5000
]

# @TODO https://www.wikidata.org/wiki/Property:P3743
#       ITU/ISO/IEC object identifier

# @TODO maybe https://www.wikidata.org/wiki/Property:P901

# @TODO maybe https://www.wikidata.org/wiki/Wikidata:Property_proposal/PM20_geo_code

# @TODO maybe https://www.wikidata.org/wiki/Property:P998

# @TODO maybe https://www.wikidata.org/wiki/Property:P5437
# @TODO maybe https://www.wikidata.org/wiki/Property:P8374
# @TODO maybe https://www.wikidata.org/wiki/Property:P3422
# @TODO maybe https://www.wikidata.org/wiki/Property:P8217
# @TODO maybe https://www.wikidata.org/wiki/Property:P4427
# @TODO maybe https://www.wikidata.org/wiki/Property:P2163
# @TODO https://www.wikidata.org/wiki/Property:P6366
# @TODO not sure https://www.wikidata.org/wiki/Property:P2983
# @TODO https://www.wikidata.org/wiki/Property:P486
# @TODO https://www.wikidata.org/wiki/Property:P2581
# @TODO https://www.wikidata.org/wiki/Property:P8714
# @TODO do it https://www.wikidata.org/wiki/Property:P8717
# @TODO do it https://www.wikidata.org/wiki/Property:P8370
# @TODO do it https://www.wikidata.org/wiki/Property:P8672
# @TODO do it https://www.wikidata.org/wiki/Property:P5337
# @TODO do it https://www.wikidata.org/wiki/Property:P3553
# @TODO do it https://www.wikidata.org/wiki/Property:P213
# @TODO check later https://www.wikidata.org/wiki/Property:P5063
# @TODO maybe https://www.wikidata.org/wiki/Property:P3097
# @TODO https://www.wikidata.org/wiki/Property:P487
# @TODO maybe https://www.wikidata.org/wiki/Property:P2979

# DONE! https://www.wikidata.org/wiki/Property:P300
#       this is for adm1, but some cases an adm0 like
#       https://www.wikidata.org/wiki/Q26273 with code NL-SX would be adm1 of NL
#       https://www.wikidata.org/wiki/Q16644 with code US-MP would be adm1 of US

# @TODO check https://www.naturalearthdata.com/downloads/110m-cultural-vectors/

# NOTE weid case for UN M.49 528:
#      - https://www.wikidata.org/wiki/Q55
#      - https://www.wikidata.org/wiki/Q29999

# Note: KML is not as popular on Wikidata as P3896 geoshape
# P3096
# https://www.wikidata.org/wiki/Property:P3096
# ix_zzkml

# P3896
# https://www.wikidata.org/wiki/Property:P3896
# ix_zzgeojson

# @TODO reorganize this


def _expand_hxl_ad_rdf():
    global HXL_ATTRIBUTES_AD_RDF
    global HXL_HASHTAGS_AD_RDF
    if '__alia__' in HXL_ATTRIBUTES_AD_RDF['geo']:
        # print('starting _prepare_HXL_HASHTAGS_AD_RDF')
        for aliud, referens in \
                HXL_ATTRIBUTES_AD_RDF['geo']['__alia__'].items():
            HXL_ATTRIBUTES_AD_RDF['geo'][aliud] = \
                HXL_ATTRIBUTES_AD_RDF['geo'][referens]
        del HXL_ATTRIBUTES_AD_RDF['geo']['__alia__']
        # print(HXL_ATTRIBUTES_AD_RDF)
    # else:
    #     print('already ready _prepare_HXL_HASHTAGS_AD_RDF')


# @TODO maybe rework this part global object part. It's used with
#       999999999_54872.py --numerordinatio-cum-antecessoribus
NUMERODINATIO_ANTECESSORIBUS__OKAY = []
NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS = [

]


# @TODO RDF_TYPUS_AD_TRIVIUM_INCOGNITA: test the implications of mixing
#       SKOS (use case: the translations) and
#       OWL (use case: data used for inferences) and how bad 'skos:Concept'
#       would break. (rocha, 2022-06-08 10:09 UTC)
# @TODO RDF_TYPUS_AD_TRIVIA_INCOGNITA: If things go bad, we can just ask
#       user to not import SKOS (or change mapping from owl:Class to
#       owl:Thing) to not break the inferences and instanceOf's
#       (rocha, 2022-06-08 10:17 UTC)


# @see https://www.w3.org/TR/skos-reference/#concepts
RDF_TYPUS_AD_TRIVIUM_INCOGNITA = 'skos:Concept'
RDF_TYPUS_AD_TRIVIUM_SEMPER = ['skos:Concept', 'owl:Thing']

# spatiīs
# - spatium, s, n, nominativus, https://en.wiktionary.org/wiki/spatium#Latin
# - nōminālī, s, n, dativus, https://en.wiktionary.org/wiki/nominalis#Latin
# - ... /spatium nominali/ (s, n)
#   - /spatia nōminālibus/ (pl, n)
# RDF_NAMESPACES = {
RDF_SPATIA_NOMINALIBUS = {
    'rdf': 'https://www.w3.org/1999/02/22-rdf-syntax-ns#',
    'rdfs': 'http://www.w3.org/2000/01/rdf-schema#',
    'xsd': 'http://www.w3.org/2001/XMLSchema#',
    'owl': 'http://www.w3.org/2002/07/owl#',
    'obo': 'http://purl.obolibrary.org/obo/',
    'skos': 'http://www.w3.org/2004/02/skos/core#',
    # 'mdciii': 'urn:mdciii:',
    # https://www.w3.org/ns/csvw.ttl
    # 'csvw': '<http://www.w3.org/ns/csvw#>',
    # 'p': 'http://www.wikidata.org/prop/',
    # 'dct': 'http://purl.org/dc/terms/',
    # 'dc': 'http://purl.org/dc/elements/1.1/',
    # @TODO see also https://www.w3.org/ns/prov.ttl boostrapper imported by
    #       https://www.w3.org/ns/csvw.ttl
    # 'unescothes': 'http://vocabularies.unesco.org/thesaurus/',
}

# This can be pre-populated by tools before being used
# @see rdf_namespaces_extras()
# @see https://www.wikidata.org/wiki/EntitySchema:E49
RDF_SPATIA_NOMINALIBUS_EXTRAS = {
    'devnull': 'http://example.org/dev/null/',
    'geo': 'http://www.opengis.net/ont/geosparql#',
    'wdata': 'http://www.wikidata.org/wiki/Special:EntityData/',
    # 'p': 'http://www.wikidata.org/prop/',
    'wdt': 'http://www.wikidata.org/prop/direct/',
    'wdv': 'http://www.wikidata.org/value/',
    'p': 'http://www.wikidata.org/prop/',  # NOTE: preffer use wdata for datasets
    'xsd': 'http://www.w3.org/2001/XMLSchema#',
}
# For "Base OWL" of Wikidata, download link: http://wikiba.se/ontology

# Note: prefixes that already are lower case do not be here
RDF_SPATIA_NOMINALIBUS_PREFIX = {

    # From https://www.w3.org/1999/02/22-rdf-syntax-ns# ______________________
    # 'rdf:type': 'rdf:type',
    # 'rdf:subject': 'rdf:subject',
    # 'rdf:predicate': 'rdf:predicate',
    # 'rdf:object': 'rdf:object',
    # 'rdf:rest': 'rdf:rest',
    # 'rdf:language': 'rdf:language',
    # 'rdf:direction': 'rdf:direction',
    # 'rdf:nil': 'rdf:nil',
    'rdf:alt': 'rdf:Alt',
    'rdf:bag': 'rdf:Bag',
    'rdf:compoundliteral': 'rdf:CompoundLiteral',
    'rdf:html': 'rdf:HTML',  # www.w3.org/TR/rdf11-concepts/#h3_section-html
    'rdf:json': 'rdf:JSON',
    'rdf:langstring': 'rdf:langString',
    'rdf:list': 'rdf:List',
    'rdf:plainliteral': 'rdf:PlainLiteral',
    'rdf:property': 'rdf:Property',
    'rdf:seq': 'rdf:Seq',
    'rdf:statement': 'rdf:Statement',
    'rdf:xmlliteral': 'rdf:XMLLiteral',

    # From http://www.w3.org/2000/01/rdf-schema# _____________________________
    'rdfs:class': 'rdfs:Class',
    'rdfs:container': 'rdfs:Container',
    # 'rdfs:member': 'rdfs:member',
    # 'rdfs:label': 'rdfs:label',
    # 'rdfs:domain': 'rdfs:domain',
    # 'rdfs:range': 'rdfs:range',
    'rdfs:containermembershipproperty': 'rdfs:ContainerMembershipProperty',
    'rdfs:datatype': 'rdfs:Datatype',
    'rdfs:isdefinedby': 'rdfs:isDefinedBy',
    'rdfs:literal': 'rdfs:Literal',
    'rdfs:resource': 'rdfs:Resource',
    'rdfs:seealso': 'rdfs:seeAlso',
    'rdfs:subclassof': 'rdfs:subClassOf',
    'rdfs:subpropertyof': 'rdfs:subPropertyOf',

    # From https://www.w3.org/TR/rdf11-concepts/#h3_xsd-datatypes ____________
    'xsd:datetime': 'xsd:dateTime',
    'xsd:datetimestamp': 'xsd:dateTimeStamp',
    'xsd:gyear': 'xsd:gYear',
    'xsd:gmonth': 'xsd:gMonth',
    'xsd:gday': 'xsd:gDay',
    'xsd:gyearmonth': 'xsd:gYearMonth',
    'xsd:gmonthday': 'xsd:gMonthDay',
    'xsd:yearmonthduration': 'xsd:yearMonthDuration',
    'xsd:daytimeduration': 'xsd:dayTimeDuration',
    'xsd:unsignedbyte': 'xsd:unsignedByte',
    'xsd:unsignedshort': 'xsd:unsignedShort',
    'xsd:unsignedint': 'xsd:unsignedInt',
    'xsd:unsignedlong': 'xsd:unsignedLong',
    'xsd:positiveinteger': 'xsd:positiveInteger',
    'xsd:nonnegativeinteger': 'xsd:nonNegativeInteger',
    'xsd:negativeinteger': 'xsd:negativeInteger',
    'xsd:nonpositiveinteger': 'xsd:nonPositiveInteger',
    'xsd:hexbinary': 'xsd:hexBinary',
    'xsd:base64binary': 'xsd:base64Binary',
    'xsd:anyuri': 'xsd:anyURI',
    'xsd:normalizedstring': 'xsd:normalizedString',
    'xsd:nmtoken': 'xsd:NMTOKEN',
    'xsd:name': 'xsd:Name',
    'xsd:ncname': 'xsd:NCName',
    # xsd:ID and xsd:IDREF are for cross references within an XML document.
    'xsd:id': 'xsd:ID',  # https://www.w3.org/TR/xmlschema11-2/#ID
    'xsd:idref': 'xsd:IDREF',  # https://www.w3.org/TR/xmlschema11-2/#IDREF
    'xsd:idrefs': 'xsd:IDREFS',  # https://www.w3.org/TR/xmlschema11-2/#IDREFS

    # From http://www.w3.org/2002/07/owl# ____________________________________
    # 'owl:imports': 'owl:imports',
    # 'owl:cardinality': 'owl:cardinality',
    # 'owl:deprecated': 'owl:deprecated',
    # 'owl:members': 'owl:members',
    'owl:alldifferent': 'owl:AllDifferent',
    'owl:alldisjointclasses': 'owl:AllDisjointClasses',
    'owl:alldisjointproperties': 'owl:AllDisjointProperties',
    'owl:annotation': 'owl:Annotation',
    'owl:annotationproperty': 'owl:AnnotationProperty',
    'owl:asymmetricproperty': 'owl:AsymmetricProperty',
    'owl:axiom': 'owl:Axiom',
    'owl:class': 'owl:Class',
    'owl:datarange': 'owl:DataRange',
    'owl:datatypeproperty': 'owl:DatatypeProperty',
    'owl:deprecatedclass': 'owl:DeprecatedClass',
    'owl:deprecatedproperty': 'owl:DeprecatedProperty',
    'owl:functionalproperty': 'owl:FunctionalProperty',
    'owl:inversefunctionalproperty': 'owl:InverseFunctionalProperty',
    'owl:irreflexiveproperty': 'owl:IrreflexiveProperty',
    'owl:namedindividual': 'owl:NamedIndividual',
    'owl:negativepropertyassertion': 'owl:NegativePropertyAssertion',
    'owl:nothing': 'owl:Nothing',
    'owl:objectproperty': 'owl:ObjectProperty',
    'owl:ontology': 'owl:Ontology',
    'owl:ontologyproperty': 'owl:OntologyProperty',
    'owl:reflexiveproperty': 'owl:ReflexiveProperty',
    'owl:restriction': 'owl:Restriction',
    'owl:symmetricproperty': 'owl:SymmetricProperty',
    'owl:transitiveproperty': 'owl:TransitiveProperty',
    'owl:thing': 'owl:Thing',
    'owl:allvaluesfrom': 'owl:allValuesFrom',
    'owl:annotatedproperty': 'owl:annotatedProperty',
    'owl:annotatedsource': 'owl:annotatedSource',
    'owl:annotatedtarget': 'owl:annotatedTarget',
    'owl:assertionproperty': 'owl:assertionProperty',
    'owl:backwardcompatiblewith': 'owl:backwardCompatibleWith',
    'owl:bottomdataproperty': 'owl:bottomDataProperty',
    'owl:bottomobjectproperty': 'owl:bottomObjectProperty',
    'owl:complementof': 'owl:complementOf',
    'owl:datatypecomplementof': 'owl:datatypeComplementOf',
    'owl:differentfrom': 'owl:differentFrom',
    'owl:disjointunionof': 'owl:disjointUnionOf',
    'owl:disjointwith': 'owl:disjointwith',
    'owl:distinctmembers': 'owl:distinctMembers',
    'owl:equivalentclass': 'owl:equivalentClass',
    'owl:equivalentproperty': 'owl:equivalentProperty',
    'owl:haskey': 'owl:hasKey',
    'owl:hasself': 'owl:hasSelf',
    'owl:hasvalue': 'owl:hasValue',
    'owl:incompatiblewith': 'owl:incompatibleWith',
    'owl:intersectionof': 'owl:intersectionOf',
    'owl:inverseof': 'owl:inverseOf',
    'owl:maxcardinality': 'owl:maxCardinality',
    'owl:maxqualifiedcardinality': 'owl:maxQualifiedCardinality',
    'owl:mincardinality': 'owl:minCardinality',
    'owl:minqualifiedcardinality': 'owl:minQualifiedCardinality',
    'owl:onclass': 'owl:onClass',
    'owl:ondatarange': 'owl:onDataRange',
    'owl:ondatatype': 'owl:onDatatype',
    'owl:oneof': 'owl:oneOf',
    'owl:onproperties': 'owl:onProperties',
    'owl:onproperty': 'owl:onProperty',
    'owl:priorversion': 'owl:priorVersion',
    'owl:propertychainaxiom': 'owl:propertyChainAxiom',
    'owl:propertydisjointwith': 'owl:propertyDisjointWith',
    'owl:qualifiedcardinality': 'owl:qualifiedCardinality',
    'owl:sameas': 'owl:sameAs',
    'owl:somevaluesfrom': 'owl:someValuesFrom',
    'owl:sourceindividual': 'owl:sourceIndividual',
    'owl:targetindividual': 'owl:targetIndividual',
    'owl:targetvalue': 'owl:targetValue',
    'owl:topdataproperty': 'owl:topDataProperty',
    'owl:topobjectproperty': 'owl:topObjectProperty',
    'owl:unionof': 'owl:unionOf',
    'owl:versioninfo': 'owl:versionInfo',
    'owl:versioniri': 'owl:versionIRI',
    'owl:withrestrictions': 'owl:withRestrictions',

    # @SEE http://www.w3.org/2004/02/skos/core# ______________________________

    'skos:collection': 'skos:Collection',
    'skos:concept': 'skos:Concept',
    'skos:conceptscheme': 'skos:ConceptScheme',
    'skos:orderedcollection': 'skos:OrderedCollection',
    'skos:altlabel': 'skos:altLabel',
    'skos:broadmatch': 'skos:broadMatch',
    # 'skos:broader': 'skos:broader',
    'skos:broadertransitive': 'skos:broaderTransitive',
    'skos:changenote': 'skos:changeNote',
    'skos:closematch': 'skos:closeMatch',
    # 'skos:definition': 'skos:definition',
    'skos:editorialnote': 'skos:editorialNote',
    'skos:exactmatch': 'skos:exactMatch',
    'skos:exactmatch': 'skos:exactMatch',
    # 'skos:example': 'skos:example',
    'skos:hastopconcept': 'skos:hasTopConcept',
    'skos:hiddenlabel': 'skos:hiddenLabel',
    'skos:historynote': 'skos:historyNote',
    'skos:inscheme': 'skos:inScheme',
    'skos:mappingrelation': 'skos:mappingRelation',
    # 'skos:member': 'skos:member',
    'skos:memberlist': 'skos:memberList',
    'skos:narrowmatch': 'skos:narrowMatch',
    # 'skos:narrower': 'skos:narrower',
    # 'skos:notation': 'skos:notation',
    # 'skos:note': 'skos:note',
    'skos:preflabel': 'skos:prefLabel',
    # 'skos:related': 'skos:related',
    'skos:relatedmatch': 'skos:relatedMatch',
    'skos:scopenote': 'skos:scopeNote',
    'skos:semanticrelation': 'skos:semanticRelation',
    'skos:topconceptof': 'skos:topConceptOf',

    # @SEE https://oborel.github.io/ _________________________________________
    # https://www.ebi.ac.uk/ols/ontologies/ro
    # ------------------------------- Entities -------------------------------
    # continuant
    'obo:bfo2': 'obo:BFO_0000002',  # purl.obolibrary.org/obo/BFO_0000002
    # occurrent
    'obo:bfo3': 'obo:BFO_0000003',  # purl.obolibrary.org/obo/BFO_0000003

    # BFO_0000029: site (used also for places)
    'obo:bfo29': 'obo:BFO_0000029',

    # @TODO: add at least the core of BFO here (or do a funcion to generate
    #        at least the numeral form of them, without the aliases)

    # ------------------------------- Relations -------------------------------
    # @TODO https://www.ebi.ac.uk/ols/ontologies/ro
    'obo:bfopartof': 'obo:BFO_0000050',  # purl.obolibrary.org/obo/BFO_0000050
    'obo:bfo50': 'obo:BFO_0000050',
    # bfo_part_of <- ro_located_in
    'obo:ro1025': 'obo:RO_0001025',  # purl.obolibrary.org/obo/RO_0001025
    'obo:rolocatedin': 'obo:BFO_0000002',

    # -------------- TOPIC: Administrative regions, COD-AB, etc --------------
    # RO_0001025 ~ BFO_0000082 ?
    'obo:locatedinatalltimes': 'obo:BFO_0000082',
    'obo:bfo82': 'obo:BFO_0000082',
    'obo:bfolocatedinatsometime': 'obo:BFO_0000171',
    'obo:bfo171': 'obo:BFO_0000171',
    'obo:locationofatsometime': 'obo:BFO_0000124',
    'obo:bfo124': 'obo:BFO_0000124',

    # @SEE http://www.opengis.net/ont/geosparql# _______________________________
    # @TODO maybe add others (the ones necessary)
    'geo:wktliteral': 'geo:wktLiteral',
}
# Note: prefixes that already are lower case should not be here
RDF_SPATIA_NOMINALIBUS_PREFIX_EXTRAS = {
}

# ------- Needs refactoring later, START -------

# def _rdf_spatia_nominalibus_prefix(rem: str, methodos: str = 'asa') -> dict:


def _rdf_spatia_nominalibus_prefix_normali(rem: str) -> dict:
    # exemplum: obo:bfo29 -> obo:BFO_0000029
    # exemplum: wdata:p2082 -> wdata:P2082
    rem_l = rem.lower()
    if rem_l in RDF_SPATIA_NOMINALIBUS_PREFIX:
        return RDF_SPATIA_NOMINALIBUS_PREFIX[rem_l]
    elif rem in RDF_SPATIA_NOMINALIBUS_PREFIX.values():
        # already normalized; redundant function call
        return rem

    if rem_l.startswith('obo:'):
        rem_ls = rem_l.replace('obo:', ''). replace('_', '')
        rem_digits = ''.join(filter(str.isdigit, rem_ls))
        rem_item_alpha = rem_ls.replace(rem_digits, '').upper()
        rem_digits_full = rem_digits.zfill(7)
        return 'obo:{0}_{1}'.format(rem_item_alpha, rem_digits_full)
    if rem_l.startswith('wdata:'):
        rem_ls = rem_l.replace('wdata:', '')
        rem_item_alpha = rem_ls.upper()
        return 'wdata:{0}'.format(rem_item_alpha)

    # Worst case: assume input already is normalized
    return rem


def _rdf_spatia_nominalibus_prefix_simplici(rem: str) -> dict:
    # exemplum: obo:BFO_0000029 -> obo:bfo29
    # exemplum: obo:BFO_0098765 -> obo:bfo98765
    # rem = 'obo:BFO_0000129'
    rem_l = rem.lower()
    # raise ValueError('oi')
    if rem_l in RDF_SPATIA_NOMINALIBUS_PREFIX:
        return rem_l
    elif rem in RDF_SPATIA_NOMINALIBUS_PREFIX.values():
        # already normalized; redundant function call
        # raise ValueError('oi2')
        clavem = list(RDF_SPATIA_NOMINALIBUS_PREFIX.keys())
        index = list(RDF_SPATIA_NOMINALIBUS_PREFIX.values()).index(rem)
        return clavem[index]

    if rem_l.startswith('obo:'):
        rem_ls = rem_l.replace('obo:', ''). replace('_', '')
        rem_digits = ''.join(filter(str.isdigit, rem_ls))
        rem_item_alpha = rem_ls.replace(rem_digits, '').lower()
        # rem_digits_full = rem_digits.zfill(7)
        return 'obo:{0}{1}'.format(rem_item_alpha, rem_digits)

    return rem

# ------- Needs refactoring later, END -------


def bcp47_langtag(
        rem: str,
        clavem: Type[Union[str, list]] = None,
        strictum: bool = True
) -> dict:
    """Public domain python function to process BCP47 langtag
    The BCP47Langtag is an public domain python function to
    aid parsing of the IETF BCP 47 language tag. It implements the syntactic
    analysis of RFC 5646 and does not require lookup tables which makes
    it friendly for quick analysis.
    Limitations:
       - subtags such as "-t-" and "-u- '' are not parsed on this current
         version. They are treated as generic BCP47 extensions.
       - The Language-Tag_normalized result would still need external data
         lookups to stricter normalization. Yet BCP47Langtag can be used as
         the very first step to analyze a language tag which is viable to port
         across programming languages.
       - The default behavior is to throw exceptions with at least some types
         of syntactic errors, a feature which can be disabled yet reasoning is
         exposed with _error. However, already very malformated can have
        further bugs which different programming implementations do not need
        to be consistent with each other.
    Versions in other programming languages:
       The initial author encourages versions to other programming languages or
       libraries or tools which make use of this. No need to ask permission
       upfront and there is no problem with using other licenses
       than public domain.
    Changelog:
       - 2022-05-28: Uses global variable BCP47_LANGTAG_EXTENSIONS to search
                     for functions to process extensions.
       - 2021-11-22: Partial implementation of BCP47 (RFC 5646)
       - 2021-01-02: Fixes on Language-Tag_normalized (discoversed when ported
                     JavaScript version was created)

    Args:
        rem (str):                       The BCP47 langtag
        clavem (Type[Union[str, list]]): Key (string) for specific value or keys
                                         (list) to return a dict (optional)
        strictum (bool):                 Throw exceptions. False replace values
                                        with False (optional)
    Returns:
        dict: Python dictionary. None means not found. False means the feature
                                 is not implemented
    Author:
        Emerson Rocha <rocha(at)ieee.org>
    License:
        SPDX-License-Identifier: Unlicense OR 0BSD

    -------------
    The syntax of the language tag in ABNF [RFC5234] is:
    Language-Tag  = langtag             ; normal language tags
                / privateuse          ; private use tag
                / grandfathered       ; grandfathered tags
    langtag       = language
                    ["-" script]
                    ["-" region]
                    *("-" variant)
                    *("-" extension)
                    ["-" privateuse]
    language      = 2*3ALPHA            ; shortest ISO 639 code
                    ["-" extlang]       ; sometimes followed by
                                        ; extended language subtags
                / 4ALPHA              ; or reserved for future use
                / 5*8ALPHA            ; or registered language subtag
    extlang       = 3ALPHA              ; selected ISO 639 codes
                    *2("-" 3ALPHA)      ; permanently reserved
    script        = 4ALPHA              ; ISO 15924 code
    region        = 2ALPHA              ; ISO 3166-1 code
                / 3DIGIT              ; UN M.49 code
    variant       = 5*8alphanum         ; registered variants
                / (DIGIT 3alphanum)
    extension     = singleton 1*("-" (2*8alphanum))
                                        ; Single alphanumerics
                                        ; "x" reserved for private use
    singleton     = DIGIT               ; 0 - 9
                / %x41-57             ; A - W
                / %x59-5A             ; Y - Z
                / %x61-77             ; a - w
                / %x79-7A             ; y - z
    privateuse    = "x" 1*("-" (1*8alphanum))
    grandfathered = irregular           ; non-redundant tags registered
                / regular             ; during the RFC 3066 era
    irregular     = "en-GB-oed"         ; irregular tags do not match
                / "i-ami"             ; the 'langtag' production and
                / "i-bnn"             ; would not otherwise be
                / "i-default"         ; considered 'well-formed'
                / "i-enochian"        ; These tags are all valid,
                / "i-hak"             ; but most are deprecated
                / "i-klingon"         ; in favor of more modern
                / "i-lux"             ; subtags or subtag
                / "i-mingo"           ; combination
                / "i-navajo"
                / "i-pwn"
                / "i-tao"
                / "i-tay"
                / "i-tsu"
                / "sgn-BE-FR"
                / "sgn-BE-NL"
                / "sgn-CH-DE"
    regular       = "art-lojban"        ; these tags match the 'langtag'
                / "cel-gaulish"       ; production, but their subtags
                / "no-bok"            ; are not extended language
                / "no-nyn"            ; or variant subtags: their meaning
                / "zh-guoyu"          ; is defined by their registration
                / "zh-hakka"          ; and all of these are deprecated
                / "zh-min"            ; in favor of a more modern
                / "zh-min-nan"        ; subtag or sequence of subtags
                / "zh-xiang"
    alphanum      = (ALPHA / DIGIT)     ; letters and numbers
    -------------

    Most tests use examples from https://tools.ietf.org/search/bcp47 and
    https://github.com/unicode-org/cldr/blob/main/tools/cldr-code
    /src/main/resources/org/unicode/cldr/util/data/langtagTest.txt

    Exemplōrum gratiā (et Python doctest, id est, automata testīs):
        (python3 -m doctest myscript.py)

    >>> bcp47_langtag('pt-Latn-BR', 'language')
    'pt'
    >>> bcp47_langtag('pt-Latn-BR', 'script')
    'Latn'
    >>> bcp47_langtag('pt-Latn-BR', 'region')
    'BR'
    >>> bcp47_langtag('de-CH-1996', 'variant')
    ['1996']
    >>> bcp47_langtag('x-fr-CH', ['language', 'region', 'privateuse'])
    {'language': None, 'region': None, 'privateuse': ['fr', 'CH']}
    >>> bcp47_langtag('i-klingon', 'grandfathered')
    'i-klingon'
    >>> bcp47_langtag('zh-min-nan', 'language')
    'zh'
    >>> bcp47_langtag('zh-min-nan', 'variant')
    ['min-nan']
    >>> bcp47_langtag('es-419', 'region')
    '419'
    >>> bcp47_langtag('en-oxendict', 'variant') # Oxford English Dictionary
    ['oxendict']
    >>> bcp47_langtag('zh-pinyin', 'variant') # Pinyin romanization
    ['pinyin']
    >>> bcp47_langtag('zh-pinyin', 'script') # Limitation: cannot infer Latn
    >>> bcp47_langtag('en-a-bbb-x-a-ccc', 'privateuse')
    ['a', 'ccc']
    >>> bcp47_langtag('en-a-bbb-x-a-ccc', 'extension')
    {'a': 'bbb'}
    >>> bcp47_langtag('tlh-a-b-foo', '_error')
    Traceback (most recent call last):
    ...
    ValueError: Errors for [tlh-a-b-foo]: extension [a] empty
    >>> bcp47_langtag('tlh-a-b-foo', '_error', False)
    ['extension [a] empty']
    >>> bcp47_langtag(
    ... 'zh-Latn-CN-variant1-a-extend1-x-wadegile-private1',
    ... ['variant', 'extension', 'privateuse'])
    {'variant': ['variant1'], 'extension': {'a': 'extend1'}, \
'privateuse': ['wadegile', 'private1']}
    >>> bcp47_langtag(
    ... 'en-Latn-US-lojban-gaulish-a-12345678-ABCD-b-ABCDEFGH-x-a-b-c-12345678',
    ... strictum=False)
    {'Language-Tag': \
'en-Latn-US-lojban-gaulish-a-12345678-ABCD-b-ABCDEFGH-x-a-b-c-12345678', \
'Language-Tag_normalized': \
'en-Latn-US-lojban-gaulish-a-12345678-ABCD-b-ABCDEFGH-x-a-b-c-12345678', \
'language': 'en', 'script': 'Latn', 'region': 'US', \
'variant': ['lojban', 'gaulish'], \
'extension': {'a': '12345678-ABCD', 'b': 'ABCDEFGH'}, \
'privateuse': ['a', 'b', 'c', '12345678'], 'grandfathered': None, \
'_callbacks': {'hxl_attrs': '+i_en+is_latn+ix_12345678', 'hxl_minimal': None}, \
'_unknown': [], '_error': ['private tag len = 1 [a]', \
'private tag len = 1 [b]', 'private tag len = 1 [c]', \
'private tag len = 1 [a]', 'private tag len = 1 [b]', \
'private tag len = 1 [c]']}

    # BCP47: "Example: The language tag "en-a-aaa-b-ccc-bbb-x-xyz" is in
    # canonical form, while "en-b-ccc-bbb-a-aaa-X-xyz" is well-formed (...)
    >>> bcp47_langtag(
    ... 'en-b-ccc-bbb-a-aaa-X-xyz')
    {'Language-Tag': 'en-b-ccc-bbb-a-aaa-X-xyz', \
'Language-Tag_normalized': 'en-a-aaa-b-ccc-bbb-x-xyz', \
'language': 'en', 'script': None, 'region': None, 'variant': [], \
'extension': {'a': 'aaa', 'b': 'ccc-bbb'}, 'privateuse': ['xyz'], \
'grandfathered': None, '_callbacks': {'hxl_attrs': '+i_en+ix_xyz', \
'hxl_minimal': None}, '_unknown': [], '_error': []}
    """
    # For sake of copy-and-paste portability, we ignore a few pylints:
    # pylint: disable=too-many-branches,too-many-statements,too-many-locals
    result = {
        # The input Language-Tag, _as it is_
        'Language-Tag': rem,
        # The Language-Tag normalized syntax, if no errors
        'Language-Tag_normalized': None,
        'language': None,
        'script': None,
        'region': None,
        'variant': [],
        'extension': {},   # Example {'a': ['bbb', 'ccc'], 'd': True}
        'privateuse': [],  # Example: ['wadegile', 'private1']
        'grandfathered': None,
        '_callbacks': {},
        '_unknown': [],
        '_error': [],
    }

    skip = 0

    if not isinstance(rem, str) or len(rem) == 0:
        result['_error'].append('Empty/wrong type')
        skip = 1
    else:
        rem = rem.replace('_', '-').strip()

    # The weird tags first: grandfathered/irregular
    if rem in [
        'en-GB-oed', 'i-ami', 'i-bnn', 'i-default', 'i-enochian',
        'i-hak', 'i-klingon', 'i-lux', 'i-ming', 'i-navajo', 'i-pwn',
            'i-tao', 'i-tay', 'i-tsu', 'sgn-BE-FR', 'sgn-BE-NL', 'sgn-CH-DE']:
        # result['langtag'] = None
        result['language'] = rem.lower()
        result['grandfathered'] = rem
        skip = 1
    # The weird tags first: grandfathered/regular
    if rem in [
            'art-lojban', 'cel-gaulish', 'no-bok', 'no-nyn', 'zh-guoyu',
            'zh-hakka', 'zh-min', 'zh-min-nan', 'zh-xiang']:

        parts_r = rem.split('-')
        # result['langtag'] = None
        result['language'] = parts_r.pop(0).lower()
        result['variant'].append('-'.join(parts_r).lower())
        result['grandfathered'] = rem
        skip = 1

    parts = rem.split('-')
    leftover = []

    deep = 0
    while len(parts) > 0 and skip == 0 and deep < 100:
        deep = deep + 1

        # BCP47 can start with private tag, without language at all
        if parts[0].lower() == 'x':
            parts.pop(0)
            while len(parts) > 0:
                result['privateuse'].append(parts.pop(0))
            break

        # BCP47 extensions start with one US-ASCII letter.
        if len(parts[0]) == 1 and parts[0].isalpha():
            if parts[0].isalpha() == 'i':
                result['_error'].append('Only grandfathered can use i-')

            extension_key = parts.pop(0).lower()
            if len(parts) == 0 or len(parts[0]) == 1:
                # BCP47 2.2.6. : "Each singleton MUST be followed by at least
                # one extension subtag (...)
                # result['extension'][extension_key] = [None]
                result['extension'][extension_key] = {}
                result['_error'].append(
                    'extension [' + extension_key + '] empty')
                continue

            result['extension'][extension_key] = ''
            while len(parts) > 0 and len(parts[0]) != 1:
                # Extensions may have more strict rules than -x-
                # @see https://datatracker.ietf.org/doc/html/rfc6497 (-t-)
                # @see https://datatracker.ietf.org/doc/html/rfc6067 (-u-)

                # Let's avoid atempt to lowercase extensions, since this is not
                # not explicity on BCP47 for unknow extensions
                # result['extension'][extension_key] = \
                #     result['extension'][extension_key] + \
                #     '-' + parts.pop(0).lower()
                result['extension'][extension_key] = \
                    result['extension'][extension_key] + \
                    '-' + parts.pop(0)

                result['extension'][extension_key] = \
                    result['extension'][extension_key].strip('-')

            continue

        # for part in parts:
        if result['language'] is None:
            if parts[0].isalnum() and len(parts[0]) == 2 or len(parts[0]) == 3:
                result['language'] = parts[0].lower()
            else:
                result['language'] = False
                result['_error'].append('language?')
            parts.pop(0)
            continue

        # Edge case to test for numeric in 4 (not 3): 'de-CH-1996'
        if len(parts[0]) == 4 and parts[0].isalpha() \
                and result['script'] is None:
            # if parts[0].isalpha() and result['script'] is None:
            if parts[0].isalpha():
                if result['region'] is None and len(result['privateuse']) == 0:
                    result['script'] = parts[0].capitalize()
                else:
                    result['script'] = False
                    result['_error'].append('script after region/privateuse')
            else:
                result['script'] = False
                result['_error'].append('script?')
            parts.pop(0)
            continue

        # Regions, such as ISO 3661-1, like BR
        if len(parts[0]) == 2 and result['region'] is None:
            if parts[0].isalpha():
                result['region'] = parts[0].upper()
            else:
                result['region'] = False
                result['_error'].append('region?')
            parts.pop(0)
            continue

        # Regions, such as ISO 3661-1, like 076
        if len(parts[0]) == 3 and result['region'] is None:
            if parts[0].isnumeric():
                result['region'] = parts.pop(0)
            else:
                result['region'] = False
                result['_error'].append('region?')
                parts.pop(0)
            continue

        if len(result['extension']) == 0 and len(result['privateuse']) == 0:
            # 2.2.5. Variant Subtags
            #   4.1 "Variant subtags that begin with a (US-ASCII)* letter
            #       (a-z, A-Z) MUST be at least five characters long."
            #   4.2 "Variant subtags that begin with a digit (0-9) MUST be at
            #       least four characters long."
            if parts[0][0].isalpha() and len(parts[0]) >= 5:
                result['variant'].append(parts.pop(0))
                continue
            if parts[0][0].isnumeric() and len(parts[0]) >= 4:
                result['variant'].append(parts.pop(0))
                continue

        leftover.append(parts.pop(0))

    result['_unknown'] = leftover

    # @TODO: maybe re-implement only for know extensions, like -t-, -u-, -h-
    # if len(result['extension']) > 0:
    #     extension_norm = {}
    #     # keys
    #     keys_sorted = sorted(result['extension'])
    #     # values
    #     for key in keys_sorted:
    #         extension_norm[key] = sorted(result['extension'][key])

    #     result['extension'] = extension_norm

    # Language-Tag_normalized
    if len(result['_error']) == 0:

        if result['grandfathered']:
            result['Language-Tag_normalized'] = result['grandfathered']
        else:
            norm = []
            if result['language']:
                norm.append(result['language'])
            if result['script']:
                norm.append(result['script'])
            if result['region']:
                norm.append(result['region'])
            if len(result['variant']) > 0:
                norm.append('-'.join(result['variant']))

            if len(result['extension']) > 0:
                #  TODO: maybe re-implement only for know extensions,
                #        like -t-, -u-, -h-. For now we're not trying to
                #        normalize ordering of unknow future extensions, BUT
                #        we sort key from different extensions
                sorted_extension = {}
                for key in sorted(result['extension']):
                    sorted_extension[key] = result['extension'][key]
                result['extension'] = sorted_extension

                for key in result['extension']:
                    if result['extension'][key][0] is None:
                        norm.append(key)
                    else:
                        norm.append(key)
                        # norm.extend(result['extension'][key])
                        norm.append(result['extension'][key])

            if len(result['privateuse']) > 0:
                norm.append('x-' + '-'.join(result['privateuse']))

            result['Language-Tag_normalized'] = '-'.join(norm)

    if len(result['extension'].keys()) > 0 and \
            'BCP47_LANGTAG_EXTENSIONS' in globals():
        for extension_key, extension_raw in result['extension'].items():
            if extension_key in BCP47_LANGTAG_EXTENSIONS:
                result['extension'][extension_key] = \
                    BCP47_LANGTAG_EXTENSIONS[extension_key](
                        extension_raw,
                        strictum=strictum
                )

    if 'BCP47_LANGTAG_CALLBACKS' in globals():
        if clavem is None or \
                ('_callbacks' == clavem or '_callbacks' in clavem):
            # pass
            for cb_key, cb_fn in BCP47_LANGTAG_CALLBACKS.items():
                result['_callbacks'][cb_key] = cb_fn(result, strictum=strictum)
        # raise NotImplementedError(BCP47_LANGTAG_CALLBACKS)

    if strictum and len(result['_error']) > 0:
        raise ValueError(
            'Errors for [' + rem + ']: ' + ', '.join(result['_error']))

    if clavem is not None:
        if isinstance(clavem, str):
            return result[clavem]
        if isinstance(clavem, list):
            result_partial = {}
            for item in clavem:
                result_partial[item] = result[item]
            return result_partial
        raise TypeError(
            'clavem [' + str(type(clavem)) + '] != [str, list]')

    return result


def bcp47_langtag_callback_hxl(
        langtag_meta: dict,
        strictum: bool = True
) -> str:
    """bcp47_langtag_callback_hxl convert a bcp47_langtag meta to hxl attributes

    Args:
        langtag_meta (dict): a bcp47_langtag compatible metadata
        strictum (bool, optional): (not implemented yet). Defaults to True.

    Returns:
        str: return HXL attributes (without HXL hashtag)
    """

    # raise ValueError(langtag_meta)

    resultatum = []
    # resultatum.append('+todo')
    resultatum.append('+i_{0}'.format(langtag_meta['language'].lower()))

    if langtag_meta['script']:
        resultatum.append('+is_{0}'.format(langtag_meta['script'].lower()))

    if langtag_meta['privateuse'] and len(langtag_meta['privateuse']) > 0:
        for item in langtag_meta['privateuse']:
            # print('    444', item)
            if len(item) > 1:
                resultatum.append('+ix_{0}'.format(item.lower()))
            else:
                langtag_meta['_error'].append(
                    'private tag len = 1 [{0}]'.format(item))

    if langtag_meta['extension'] and 'r' in langtag_meta['extension']:
        _r = langtag_meta['extension']['r']

        if 'rdf:type' in _r and len(_r['rdf:type']) > 0:
            for item in _r['rdf:type']:

                _temp1, temp2 = item.split('||')
                subject_key = _temp1
                subject_namespace = temp2
                # _predicate = _rdf_spatia_nominalibus_prefix_normali(_temp1)
                _predicate = _rdf_spatia_nominalibus_prefix_simplici(_temp1)
                # raise ValueError(_predicate)
                __predicate_prefix, _predicate_item = _predicate.split(':')
                subject_namespace = subject_namespace.replace(':nop', '')
                resultatum.append(
                    '+rdf_a_{0}_{1}'.format(
                        __predicate_prefix, _predicate_item))

        if _r['rdf:predicate'] and len(_r['rdf:predicate']) > 0:
            for item in _r['rdf:predicate']:
                # prefix, term, subject_domain, _nop2 = item.lower().split(':')
                # raise ValueError(item)
                # print(item)
                normalized_predicate, _subject_part = item.lower().split('||')
                subject_domain = _subject_part

                # discarting not yet implemented additional subject meta :NOP
                _subject_part = _subject_part.replace(':nop', '')

                # SPECIAL CASE: OBO prefix we remove leading zeroes and _
                if normalized_predicate.startswith('obo:'):
                    normalized_predicate = normalized_predicate.lower()
                    _predicate_ns = 'obo'
                    _predicate_item_raw = \
                        normalized_predicate.replace('obo:', '')
                    _predicate_item_raw_digits = ''.join(
                        filter(str.isdigit, _predicate_item_raw))
                    _predicate_item_raw_alpha = \
                        _predicate_item_raw.replace(
                            _predicate_item_raw_digits, '').replace(
                                '_', '')
                    _predicate_item = '{0}{1}'.format(
                        _predicate_item_raw_alpha,
                        _predicate_item_raw_digits.lstrip('0')
                    )
                    # prefix, term = _predicate_item.split(':')
                    subject_domain = _subject_part.lstrip('0')
                    prefix = _predicate_ns
                    # term = _predicate_item_raw_digits
                    term = _predicate_item
                else:
                    subject_domain, _nop = subject_domain.split(':')
                    prefix, term = normalized_predicate.split(':')
                # raise ValueError(_predicate_part)
                resultatum.append('+rdf_p_{0}_{1}_s{2}'.format(
                    prefix, term, subject_domain))

        if 'rdf:subject' in _r and len(_r['rdf:subject']) > 0:
            for item in _r['rdf:subject']:
                # subject_key, subject_namespace, _nop = item.lower().split(':')
                _temp1, temp2 = item.lower().split('||')
                subject_key = _temp1
                subject_namespace = temp2
                subject_namespace = subject_namespace.replace(':nop', '')
                resultatum.append(
                    '+rdf_s_{0}_s{1}'.format(subject_key, subject_namespace))

        if _r['rdfs:Datatype'] and len(_r['rdfs:Datatype']) > 0:
            # prefix, term, _nop = _r['rdfs:Datatype'].lower().split(':')
            _temp1, _temp2 = _r['rdfs:Datatype'].lower().split('||')
            prefix, term = _temp1.split(':')
            resultatum.append('+rdf_t_{0}_{1}'.format(prefix, term))

        if _r["xsl:transform"] and len(_r["xsl:transform"]) > 0:
            value_prefixes = None
            value_separator = None
            for titem in _r["xsl:transform"]:
                # tverb, tval_1, _nop_tval_2 = titem.split(':')

                _temp1, temp2 = titem.split('||')
                tverb = _temp1
                tval_1, _nop_tval_2 = temp2.split(':')

                if tverb.lower() == EXTRA_OPERATORS['STX']['hxl']:
                    if value_prefixes is None:
                        value_prefixes = []
                    value_prefixes.append(tval_1)

                elif tverb.lower() == EXTRA_OPERATORS['GS']['hxl']:
                    # @TODO: check this at compile time
                    if tval_1 in CSVW_SEPARATORS:
                        value_separator = tval_1
                    else:
                        raise NotImplementedError(
                            'Separator [{0}] not implemented '
                            'Context: [{1}] ; Options <{2}>'.format(
                                tval_1, langtag_meta, CSVW_SEPARATORS
                            ))

                    # value_separator = CSVW_SEPARATORS[tval_1]

            if value_prefixes is not None:
                for titem in value_prefixes:
                    resultatum.append('+rdf_y_{0}_{1}'.format(
                        EXTRA_OPERATORS['STX']['hxl'], titem))

            if value_separator is not None:
                resultatum.append('+rdf_y_{0}_{1}'.format(
                    EXTRA_OPERATORS['GS']['hxl'], value_separator))

    resultatum = sorted(resultatum)

    # raise ValueError(resultatum, langtag_meta)

    return ''.join(resultatum)


def bcp47_langtag_callback_hxl_minimal(
        langtag_meta: dict,
        strictum: bool = True
) -> str:
    """bcp47_langtag_callback_hxl convert a bcp47_langtag meta to hxl attributes

    Args:
        langtag_meta (dict): a bcp47_langtag compatible metadata
        strictum (bool, optional): (not implemented yet). Defaults to True.

    Returns:
        str: return HXL attributes (without HXL hashtag)
    """
    # print('langtag_meta', langtag_meta)
    res = bcp47_langtag_callback_hxl(langtag_meta, strictum)
    minimal_parts = []
    extra_parts = []
    parts = res.replace('+i_qcc+is_zxxx', '').split('+')

    parts = list(filter(None, parts))

    # raise ValueError(res)

    # We only try to compact interlingual concepts, not linguistic
    if not res.startswith('+i_qcc+is_zxxx'):
        return None

    # Do no run twice
    if res.find('+ix_hxlattrs') > -1:
        return None

    # If does not have an RDF tab about being a pivot key, we return False
    # This may either signal a error OR a HXL tag that subject is implicit
    # u2203
    if res.find('+rdf_s_u2200_s') == -1 and res.find('+rdf_s_u2203_s') == -1:
        # The folowing block is disabled for now as theres some cases
        # valid +ix_ prefixes exist, but large tables they would conflict.

        # # ... however, if it does have a keys -x-LLL (+ix_LLL), let's attempt
        # # to assume they are unique engouth and sort it.
        # # ... except if is ix_error
        # if res.find('+ix_') > -1 and res.find('+ix_error') == -1:
        #     for item in parts:
        #         if item.startswith('ix_'):
        #             minimal_parts.append(item)
        #         else:
        #             extra_parts.append(item)
        #     if len(extra_parts) == 0:
        #         return [res, None]

        #     minimal = '+i_qcc+is_zxxx+' + '+'.join(minimal_parts)
        #     extra = '+'.join(extra_parts)

        #     return [minimal, extra]

        # ... and do this also for data types such as +rdf_t_xsd_datetime
        # even if does not have ix_
        if res.find('+rdf_t_') > -1:
            for item in parts:
                if not item.startswith('rdf_t_'):
                    minimal_parts.append(item)
                else:
                    extra_parts.append(item)
            if len(extra_parts) == 0:
                return [res, None]

            minimal = '+i_qcc+is_zxxx+' + '+'.join(minimal_parts)
            extra = '+'.join(extra_parts)

            return [minimal, extra]

        return False

    for item in parts:
        if item.startswith('rdf_s_u2200_s') or item.startswith('rdf_s_u2203_s'):
            minimal_parts.append(item)
        else:
            extra_parts.append(item)
    if len(extra_parts) == 0:
        return [res, None]

    minimal = '+i_qcc+is_zxxx+' + '+'.join(minimal_parts)
    extra = '+'.join(extra_parts)

    return [minimal, extra]


def bcp47_rdf_extension(
        rem: str,
        clavem: Type[Union[str, list]] = None,
        strictum: bool = True
) -> dict:
    """""Public domain python function to process RDF "G" extension for BCP47

    The bcp47_rdf_extension is an public domain python function to
    aid parsing of the IETF BCP 47 language tag. It implements the syntactic
    analysis of RFC 5646 and does not require lookup tables which makes
    it friendly for quick analysis.

    The general description of extensions is described at
    https://www.rfc-editor.org/rfc/rfc5646.html#section-2.2.6

    Args:
        bcp47_extension_r (str):         The r extention part of a BCP47
                                         language tag
        clavem (Type[Union[str, list]]): Key (string) for specific value or keys
                                         (list) to return a dict (optional)
        strictum (bool):                 Throw exceptions. False replace values
                                        with False (optional)

    Returns:
        dict: Python dictionary. None means not found. False means the feature
                                 is not implemented

    Changelog:
        - 2022-05-28: Created.

    Author:
        Emerson Rocha <rocha(at)ieee.org>

    License:
        SPDX-License-Identifier: Unlicense OR 0BSD

    -----
    See also
        - en.wikipedia.org/wiki/Resource_Description_Framework#Vocabulary
        - https://en.wikipedia.org/wiki/Reification_(knowledge_representation)
        - https://en.wikipedia.org/wiki/First-order_logic
        - https://en.wikipedia.org/wiki/Universal_quantification
        - https://en.wikipedia.org/wiki/List_of_logic_symbols
          - "∀" U+2200
          - “🔗” (U+1F517)


    RDF Vocabulary / Properties
        - rdf:subject – the subject of the RDF statement
        - rdf:predicate – the predicate of the RDF statement

    @TODO: - https://en.wikipedia.org/wiki/First-order_logic#Logical_symbols
             - The logical connectives: ∧ for conjunction,
               ∨ for disjunction, → for implication, ↔ for biconditional,
               ¬ for negation. Occasionally other logical connective symbol
               are included. Some authors[7] use Cpq, instead of →,
               and Epq, instead of ↔, especially in contexts where → is used
               for other purposes. Moreover, the horseshoe ⊃ may replace →;
               the triple-bar ≡ may replace ↔;
               a tilde (~), Np, or Fp, may replace ¬;
               a double bar {\displaystyle \|}\|, {\displaystyle +}+ or
               Apq may replace ∨; and ampersand &, Kpq, or the middle dot, ⋅,
               may replace ∧, especially if these symbols are not available
               for technical reasons. (The aforementioned symbols
               Cpq, Epq, Np, Apq, and Kpq are used in Polish notation.)
           - en.wikipedia.org/wiki/Polish_notation#Polish_notation_for_logic
           - https://en.wikipedia.org/wiki/Hungarian_notation
           - https://en.wikipedia.org/wiki/Leszynski_naming_convention

    -----

    Exemplōrum gratiā (et Python doctest, id est, automata testīs):
        (python3 -m doctest myscript.py)

    >>> bcp47_rdf_extension('pskos-pprefLabel-ps1', 'rdf:predicate')
    ['skos:prefLabel||1:NOP']

    >>> bcp47_rdf_extension(
    ... 'pDC-pcontributor-ps2-pDC-pcreator-ps3-pDC-ppublisher-ps4',
    ... 'rdf:predicate')
    ['dc:contributor||2:NOP', 'dc:creator||3:NOP', 'dc:publisher||4:NOP']

    >>> bcp47_rdf_extension('pDCT-pmodified-ps1-tXSD-tdateTime-tnop',
    ... ['rdf:predicate', 'rdfs:Datatype'])
    {'rdf:predicate': ['dct:modified||1:NOP'], \
'rdfs:Datatype': 'xsd:dateTime||NOP'}

    """
    # For sake of copy-and-paste portability, we ignore a few pylints:
    # pylint: disable=too-many-branches,too-many-statements,too-many-locals
    result = {
        'rdf:Statement_raw': rem,
        # 'bcp47_extension_r_normalized': None,
        'rdf:subject': [],
        'rdf:predicate': [],
        'rdf:object': [],
        'rdf:type': [],
        'rdfs:Datatype': None,
        'xsl:transform': [],
        '_unknown': [],
        '_error': [],
        # 'csvw:separator': '', # Added only if necessary
        # 'prefix': [],  # Added only if necessary
    }
    # _predicates = []
    _subjects = []
    _objects = []

    # result['bcp47_extension_r_normalized'] = \
    #     result['bcp47_extension_r'].lower()

    if rem.find('-') > 0:
        r_parts = rem.split('-')
        r_parts_tot = len(r_parts)
        r_rest = r_parts_tot
        is_disbalanced = False

        if len(r_parts) % 3 != 0:
            is_disbalanced = True
            result['_error'].append('G extension not groups of 3: {0}'.format(
                len(r_parts)
            ))

        while (r_rest > 0) and (is_disbalanced is False):
            r_verb = r_parts[r_parts_tot - r_rest]
            r_op_1 = r_parts[r_parts_tot - r_rest + 1]
            r_op_2 = r_parts[r_parts_tot - r_rest + 2]

            r_op = r_verb[0].lower()

            if r_op_1[0].lower() != r_op or r_op_2[0].lower() != r_op:
                result['_error'].append(
                    'Prefix not equal [{0}]: [{1}-{2}-{3}]'.format(
                        r_op,
                        r_verb,
                        r_op_1,
                        r_op_2,
                    ))
                r_rest = r_rest - 3
                continue

            r_verb = r_verb[1:]
            r_op_1 = r_op_1[1:]
            r_op_2 = r_op_2[1:]
            if r_op_2 == 'nop':
                r_op_2 = 'NOP'

            if r_op == 'a':

                verb = r_verb.lower() + ':' + r_op_1

                if verb in RDF_SPATIA_NOMINALIBUS_PREFIX:
                    verb = RDF_SPATIA_NOMINALIBUS_PREFIX[verb]

                result['rdf:type'].append('{0}||0:{1}'.format(
                    verb, r_op_2
                ))
                # pass

            elif r_op == 'p':
                if r_op_2[0].lower() != 's':
                    result['_error'].append(
                        '[{3}] only implemented for reference subject. '
                        'This means require prefix [s]: '
                        'Used: [{0}{1}-{0}{2}{0}-{3}]'.format(
                            r_op,
                            r_verb,
                            r_op_1,
                            r_op_2,

                        ))
                else:
                    r_op_2 = r_op_2[1:]

                    result['rdf:predicate'].append('{0}:{1}||{2}:{3}'.format(
                        r_verb.lower(), r_op_1, r_op_2, 'NOP'
                    ))
            elif r_op == 's':
                _subjects.append('{0}||{1}:{2}'.format(
                    r_verb.upper(), r_op_1.lower(), r_op_2
                ))
            elif r_op == 'o':
                result['_error'].append(
                    'rdf:object not implemented yet.'
                    'Used: [{0}{1}-{0}{2}{0}-{3}]'.format(
                        r_op,
                        r_verb,
                        r_op_1,
                        r_op_2,
                    ))
                # continue
            elif r_op == 't':
                if result['rdfs:Datatype'] is None:
                    datatype = _rdf_spatia_nominalibus_prefix_normali(
                        '{0}:{1}'.format(r_verb.lower(), r_op_1)
                    )
                    # raise ValueError(r_verb, datatype)
                    # r_verb = r_verb.lower()
                    # result['rdfs:Datatype'] = '{0}:{1}||{2}'.format(
                    #     r_verb, r_op_1, 'NOP'
                    # )
                    result['rdfs:Datatype'] = '{0}||{1}'.format(
                        datatype, 'NOP'
                    )
                else:
                    result['_error'].append(
                        'rdfs:Datatype duplicated [{0}]-[{1}]'.format(
                            r_verb.lower(),
                            r_op_1
                        ))
            elif r_op == 'y':
                result['xsl:transform'].append('{0}||{1}:{2}'.format(
                    r_verb.upper(), r_op_1.lower(), r_op_2,
                ))
            else:
                result['_error'].append(
                    'Unknown prefix [{0}]: [{0}{1}-{0}{2}{0}-{3}]'.format(
                        r_op,
                        r_verb,
                        r_op_1,
                        r_op_2,
                    ))

            r_verb = r_verb.upper()
            r_op_2 = r_op_2.lower()

            r_rest = r_rest - 3

        # if len(_predicates) > 0:
        #     _predicates.sort()
        #     result['rdf:predicate'] = _predicates
        if len(result['rdf:predicate']) > 0:
            result['rdf:predicate'].sort()
            # raise ValueError(result['rdf:predicate'])
            # print('all', result['rdf:predicate'])
            for index in range(len(result['rdf:predicate'])):
                _item_parts = result['rdf:predicate'][index]

                # print(_item_parts)

                # _predicate_ns, _predicate_item, _subject, _subject_nop = \
                #     _item_parts.split(':')
                _temp1, _temp2 = _item_parts.split('||')
                _predicate_ns, _predicate_item = _temp1.split(':')
                _subject, _subject_nop = _temp2.split(':')

                raw_predicate = f'{_predicate_ns}:{_predicate_item}'
                normalized_predicate = None

                if raw_predicate in RDF_SPATIA_NOMINALIBUS_PREFIX:
                    normalized_predicate = RDF_SPATIA_NOMINALIBUS_PREFIX[raw_predicate]
                elif raw_predicate in RDF_SPATIA_NOMINALIBUS_PREFIX_EXTRAS:
                    normalized_predicate = \
                        RDF_SPATIA_NOMINALIBUS_PREFIX_EXTRAS[raw_predicate]

                if normalized_predicate is not None:
                    if normalized_predicate.startswith('obo:'):
                        normalized_predicate = normalized_predicate.lower()
                        _predicate_ns = 'obo'
                        _predicate_item_raw = \
                            normalized_predicate.replace('obo:', '')
                        _predicate_item_raw_digits = ''.join(
                            filter(str.isdigit, _predicate_item_raw))
                        _predicate_item_raw_alpha = \
                            _predicate_item_raw.replace(
                                _predicate_item_raw_digits, '').replace(
                                    '_', '')
                        _predicate_item = '{0}{1}'.format(
                            _predicate_item_raw_alpha,
                            _predicate_item_raw_digits.lstrip('0')
                        )
                    else:
                        _predicate_ns, _predicate_item = \
                            normalized_predicate.split(':')
                        # pass

                    result['rdf:predicate'][index] = \
                        '{0}||{1}:{2}'.format(
                        RDF_SPATIA_NOMINALIBUS_PREFIX[raw_predicate],
                        _subject, _subject_nop)
                else:
                    # _p1, _p2, _s1, _s2 = \
                    #     result['rdf:predicate'][index].split(':')
                    # result['rdf:predicate'][index] = '{0}:{1}||{2}:{3}'.format(
                    #     _p1, _p2, _s1, _s2
                    # )
                    result['rdf:predicate'][index] = '{0}:{1}||{2}:{3}'.format(
                        _predicate_ns, _predicate_item, _subject, _subject_nop
                    )
                    # pass

                # if raw_predicate in RDF_SPATIA_NOMINALIBUS_PREFIX:
                #     result['rdf:predicate'][index] = '{0}:{1}'.format(
                #         RDF_SPATIA_NOMINALIBUS_PREFIX[raw_predicate], subject)
                # pass

        if len(_objects) > 0:
            _objects.sort()
            result['rdf:object'] = _objects

        if len(_subjects) > 0:
            _subjects.sort()
            result['rdf:subject'] = _subjects

        if len(result['xsl:transform']) > 0:
            result['xsl:transform'].sort()

            # result['rdf:subject'] = _subjects

    else:
        result['_error'].append('G extension do not have -')

    # if len(r_parts) % 3 == 0:
    #     pass
    # else:
    #     result['_error'].append('G extension not groups of 3')

    # if len(r_parts) % 2 == 0:
    #     pass
    # else:
    #     result['_error'].append('G extension not even number')

    if strictum and len(result['_error']) > 0:
        raise SyntaxError('[{0}]: <{1}>'.format(
            rem, result['_error']))

    if clavem is not None:
        if isinstance(clavem, str):
            return result[clavem]
        if isinstance(clavem, list):
            result_partial = {}
            for item in clavem:
                result_partial[item] = result[item]
            return result_partial
        raise TypeError(
            'clavem [' + str(type(clavem)) + '] != [str, list]')

    return result


def bcp47_rdf_extension_caput_ad_columnae_i(
        caput_originali: List[str],
        caput_originali_asa: List[str],
        namespaces: List[dict] = None,
        strictum: bool = True
) -> dict:
    resultatum = []
    extras = []
    for index in range(len(caput_originali)):
        _hxl_minimal = caput_originali_asa[index]['_callbacks']['hxl_minimal']
        if not _hxl_minimal:
            resultatum.append(caput_originali[index])
        else:
            # resultatum.append(caput_originali[index])
            _meta = hxl_hashtag_to_bcp47('#item' + _hxl_minimal[0])
            resultatum.append(_meta['Language-Tag_normalized'])

            # @TODO this strategy to check if already -x- is too simplistic.
            #       eventually could be improved. Just doing a quick hacky here
            #       (Rocha, 2022-07-26 03:36 UTC)
            if _meta['Language-Tag_normalized'].find('-x-') > -1:
                extras.append([
                    _meta['Language-Tag_normalized'] + '-hxlattrs',
                    _hxl_minimal[1]
                ])
            else:
                extras.append([
                    _meta['Language-Tag_normalized'] + '-x-hxlattrs',
                    _hxl_minimal[1]
                ])
            # resultatum.append(caput_originali[index])
        # resultatum.append(caput_originali_asa[index])
        # print(caput_originali_asa[index])
        # break
    # return caput_originali
    if len(extras):
        resultatum.extend(extras)

    return resultatum


def bcp47_rdf_extension_relationship(
        header: List[str],
        namespaces: List[dict] = None,
        strictum: bool = True
) -> dict:
    """""Metadata processing of the bcp47_rdf_extension version

    Note about caput_ad_columnae_i:
    - We can generate an ASA about how would be if long column names
      would be renamed and added as text value, (e.g. do it from a
      wide dataset), but not designed to do it from narrow dataset

    Args:
        caput (List[str]): _description_
        strictum (bool, optional): _description_. Defaults to True.

    Returns:
        dict: _description_
    """
    # About caput_**, see
    # - https://en.wikipedia.org/wiki/Wide_and_narrow_data
    # - https://tidyr.tidyverse.org/
    # - pandas.pydata.org/pandas-docs/stable/reference/api/pandas.melt.html
    # - pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.pivot.html

    # @TODO: this function is obviously doing too much at once. Eventually
    #        can be refactored. (rocha, 2022-06-09 10:50 UTC)
    result = {
        'caput_originali': header,
        'caput_ad_columnae_i': [],
        'caput_originali_asa': [],
        # 'rdf:subject': None,
        # 'rdf:predicate': [],
        # 'rdf:object': None,
        # 'rdfs:Datatype': None,
        # '_unknown': [],
        'rdfs:Container': {},
        'rdf_spatia_nominalibus': RDF_SPATIA_NOMINALIBUS,
        # 'rdf_spatia_nominalibus': {
        #     'rdf': RDF_SPATIA_NOMINALIBUS['rdf'],
        #     'rdfs': RDF_SPATIA_NOMINALIBUS['rdfs'],
        #     'xsd': RDF_SPATIA_NOMINALIBUS['xsd'],
        #     'owl': RDF_SPATIA_NOMINALIBUS['owl'],
        #     'obo': RDF_SPATIA_NOMINALIBUS['obo'],
        #     'skos': RDF_SPATIA_NOMINALIBUS['skos'],
        #     'wdata': RDF_SPATIA_NOMINALIBUS['wdata'],
        # },
        'trivium_aliis_per_indicem': {},
        '_error': [],
    }

    result['rdf_spatia_nominalibus'] = RDF_SPATIA_NOMINALIBUS
    # RDF_TYPUS_AD_TRIVIA_INCOGNITA

    if namespaces is not None and len(namespaces) > 0:
        for item in namespaces:
            result['rdf_spatia_nominalibus'][item['prefix']] = item['iri']

    # print('header', header)

    def _aux_init_container(result: dict, subject: str) -> dict:
        """_aux_init_container

        Each time a column asks for a subjec group, we create one with default
        values. Note that the container MAY not actually have a primary key
        (which could lead to break) but as long as the user don't ask
        for such group, it will not break
        """
        result['rdfs:Container'][subject] = {
            'trivium': {
                # 'alia': (str(subject), str(subject)),
                # 'alia': {str(subject)},
                'alia': [str(subject)],
                # 'alia': [str(subject)],
                'index': -1,
                # 'iri': inline_namespace_iri,
                # 'rdf_praefixum': 'urn',
                'rdf_praefixum': 'urnmdciii',
                # We will fallback the pivots as generic classes
                # We should enable later override this behavior
                # via language tag on the pivot
                'rdf:predicate': [],
                # @TODO: implement the semantics of is_a
                'rdf:type': [],
                # # aliīs, pl, m/f/n, dativus, en.wiktionary.org/wiki/alius#Latin
                # 'trivium_aliis': []
            },
            'indices_columnis': [],
            # aliīs, pl, n, ablativus, en.wiktionary.org/wiki/alius#Latin
            'indices_cum_aliis': []
        }
        return result

    def _aux_recalc_containers(result: dict) -> dict:
        """_aux_recalc_containers

        The major reason for this recalculation is at the end resolve the
        aliases for same group. This allow copy-and-pasting strategies
        while the user could ADD (not replace) new names for keys
        and use these names to create more focused relations
        """
        trivium_aliis = {}

        # First pass
        for item_caput_asa in result['caput_originali_asa']:
            if 'extension' not in item_caput_asa or \
                    'r' not in item_caput_asa['extension']:
                continue
            _r = item_caput_asa['extension']['r']
            # print('oi foi', item_caput_asa)
            item_subject_list = _r['rdf:subject']
            if len(item_subject_list) == 0:
                continue
            index_ex_tabula = item_caput_asa['_index_ex_tabula']
            for item_subject in item_subject_list:
                # print(item_subject)
                # Exemplum: "U2200||500:NOP"
                _temp1, _temp2 = item_subject.split('||')
                aliud = _temp2.split(':').pop(0)
                if index_ex_tabula not in trivium_aliis:
                    trivium_aliis[index_ex_tabula] = set()
                trivium_aliis[int(index_ex_tabula)].add(aliud)

        # print(trivium_aliis, result['rdfs:Container'])

        # Second pass
        _trivium_aliis = []
        for trivium_alii, _item in result['rdfs:Container'].items():
            _trivium_aliis.append(trivium_alii)
            _trivium_indici = _item['trivium']['index']
            _cum_aliis = []

            if _trivium_indici == -1:
                # Item is referenced by others, but does not explicitly exist
                continue

            for _item in trivium_aliis[_trivium_indici]:
                _cum_aliis.extend(
                    result['rdfs:Container'][_item]['indices_columnis'])

            # trivium_aliis[_trivium_indici] = \
            #     set(list(trivium_aliis[_trivium_indici]).sort())
            # trivium_aliis[_trivium_indici].sort()
            trivium_aliis[_trivium_indici] = \
                list(trivium_aliis[_trivium_indici])
            trivium_aliis[_trivium_indici].sort(key=int)

            result['rdfs:Container'][trivium_alii]['trivium']['alia'] = \
                trivium_aliis[_trivium_indici]

            _cum_aliis = list(set(_cum_aliis))
            _cum_aliis.sort(key=int)

            result['rdfs:Container'][trivium_alii]['indices_cum_aliis'] = \
                _cum_aliis
            # print(trivium_alii, _item)
            # if 'extension' not in item_caput_asa or \
            #         'r' not in item_caput_asa['extension']:
            #     continue

        # Third pass; mostly to re-order rdfs:Container to generate same
        # abstract syntax tree even if order of columns change
        _trivium_aliis.sort(key=int)
        rdfs_container_novo = {}
        for trivium_alii in _trivium_aliis:
            rdfs_container_novo[trivium_alii] = \
                result['rdfs:Container'][trivium_alii]
            # pass

        result['rdfs:Container'] = rdfs_container_novo

        result['trivium_aliis_per_indicem'] = trivium_aliis
        return result

    # ========= Fist iteration over each column, START =========
    for index, item in enumerate(header):
        item_meta = bcp47_langtag(
            item, ['language', 'script', 'extension',
                   'privateuse', '_callbacks'],
            strictum=False)
        # @TODO; get erros and export them to upper level
        # item_meta['_column'] = index
        item_meta['_index_ex_tabula'] = index
        inline_namespace = None
        inline_namespace_iri = None
        is_inline_namespace = False

        # raise ValueError(item_meta)

        is_pivot_key = False

        object_prefixes = None
        if 'r' in item_meta['extension'] and \
                'prefix' in item_meta['extension']['r'] and \
                len(item_meta['extension']['r']['prefix']) > 0:
            object_prefixes = item_meta['extension']['r']['prefix']

        # RDFStatement: Subject -> Predicate -> [[ Object ]]
        if 'r' in item_meta['extension'] and \
                len(item_meta['extension']['r']['rdf:object']) > 0:
            for object in item_meta['extension']['r']['rdf:object']:
                # if object.startswith('🔗'):
                #     # is_inline_namespace = True
                #     inline_namespace = object.replace('🔗', '')
                #     strictum = True
                if object.startswith('_'):
                    # is_inline_namespace = True
                    inline_namespace = object.replace('_', '')
                    strictum = True
                    if inline_namespace in RDF_SPATIA_NOMINALIBUS:
                        inline_namespace_iri = \
                            RDF_SPATIA_NOMINALIBUS[inline_namespace]
                        result['rdf_spatia_nominalibus'][inline_namespace] = \
                            inline_namespace_iri
                    elif inline_namespace in RDF_SPATIA_NOMINALIBUS_EXTRAS:
                        inline_namespace_iri = \
                            RDF_SPATIA_NOMINALIBUS_EXTRAS[inline_namespace]
                        result['rdf_spatia_nominalibus'][inline_namespace] = \
                            inline_namespace_iri
                    else:
                        if strictum:
                            raise SyntaxError(
                                'inline_namespace ({0}) ? <{1}> <{2}>'.format(
                                    inline_namespace, header, item_meta
                                ))
                        else:
                            inline_namespace_iri = '_' + inline_namespace

        # RDFStatement: [[ Subject ]] -> Predicate -> Object
        if 'r' in item_meta['extension'] and \
                len(item_meta['extension']['r']['rdf:subject']) > 0:
            for subject in item_meta['extension']['r']['rdf:subject']:
                # is_pivot_key = False
                subject_key, subject_value = subject.split(':')
                _temp1, _temp2 = subject.split('||')
                subject_key = _temp1
                subject_value = _temp2.replace(':NOP', '')
                # raise ValueError(subject)

                if subject.startswith('∀') or subject.startswith('∃') or \
                        subject.lower().startswith('u2200') or \
                        subject.lower().startswith('u2203'):
                    is_pivot_key = True
                    # raise ValueError('deu', is_pivot_key)

                if subject_value not in result['rdfs:Container']:
                    result = _aux_init_container(result, subject_value)

                if inline_namespace is not None:
                    result['rdfs:Container'][subject_value][
                        'trivium']['rdf_praefixum'] = inline_namespace

                # Add itself.
                # @TODO test better corner cases
                result['rdfs:Container'][subject_value][
                    'indices_columnis'].append(
                    index)

                if is_pivot_key:
                    if result['rdfs:Container'][subject_value][
                            'trivium']['index'] > -1:
                        SyntaxError('{0} <{1}>'.format(header, item_meta))
                    if object_prefixes is not None and len(object_prefixes) > 1:
                        SyntaxError('{0} <{1}>:: > 1 prefix [{2}]'.format(
                            header, item_meta, object_prefixes))
                    if object_prefixes is not None:
                        result['rdfs:Container'][subject_value][
                            'trivium']['rdf_praefixum'] = object_prefixes[0]
                    # raise ValueError('deu', index)
                    result['rdfs:Container'][subject_value][
                        'trivium']['index'] = index

        # RDFStatement: Subject -> [[ Predicate ]] -> Object
        if 'r' in item_meta['extension'] and \
                len(item_meta['extension']['r']['rdf:predicate']) > 0:
            for item_p in item_meta['extension']['r']['rdf:predicate']:
                # raise ValueError(predicate)
                # prefix, suffix = predicate.split(':')
                predicate, subject = item_p.split('||')
                predicate_namespce, _ignore = predicate.split(':')
                subject_value = subject.replace(':NOP', '')  # Not used... yet
                if predicate_namespce not in result['rdf_spatia_nominalibus']:
                    # if prefix not in RDF_SPATIA_NOMINALIBUS:
                    if predicate_namespce not in RDF_SPATIA_NOMINALIBUS_EXTRAS:
                        raise SyntaxError(
                            'prefix [{0}]? <{1}> <{2}>'.format(
                                predicate_namespce, header,
                                RDF_SPATIA_NOMINALIBUS_EXTRAS
                            ))
                    result['rdf_spatia_nominalibus'][predicate_namespce] = \
                        RDF_SPATIA_NOMINALIBUS_EXTRAS[predicate_namespce]

                # print(item, item_meta)
                # print(item_meta)
                if subject_value not in result['rdfs:Container']:
                    result = _aux_init_container(result, subject_value)

                result['rdfs:Container'][subject_value][
                    'indices_columnis'].append(
                    item_meta['_index_ex_tabula'])

        result['caput_originali_asa'].append(item_meta)

    # ========= Fist iteration over each column, END =========
    # Note: after here still necessary do some additional checks

    # Clean up, general changes, etc on rdfs:Container
    for item in result['rdfs:Container']:
        if RDF_TYPUS_AD_TRIVIUM_INCOGNITA and \
                len(result['rdfs:Container'][item]['trivium']['rdf:type']) == 0:
            result['rdfs:Container'][item]['trivium']['rdf:type'] = \
                [RDF_TYPUS_AD_TRIVIUM_INCOGNITA + '||0:NOP']
        if len(RDF_TYPUS_AD_TRIVIUM_SEMPER) > 0:
            for semper_typus in RDF_TYPUS_AD_TRIVIUM_SEMPER:
                if semper_typus + '||0:NOP' not in \
                        result['rdfs:Container'][item]['trivium']['rdf:type']:
                    result['rdfs:Container'][item][
                        'trivium']['rdf:type'].append(
                        semper_typus + '||0:NOP'
                    )
                # pass

        # remove duplicates
        _rdf_container_indices = result['rdfs:Container'].keys()
        # print(_rdf_container_indices)
        for _, item2 in enumerate(_rdf_container_indices):
            # print(item, result['rdfs:Container'])
            indices_columnis_unicus = \
                result['rdfs:Container'][item2]['indices_columnis']
            indices_columnis_unicus = list(set(indices_columnis_unicus))
            result['rdfs:Container'][item2]['indices_columnis'] = \
                indices_columnis_unicus

        # print(result['rdfs:Container'][item]['trivium']['index'])
        _trivium_indici = int(
            result['rdfs:Container'][item]['trivium']['index'])

        _trivum_xsl = result['caput_originali_asa'][
            _trivium_indici]['extension']['r']['xsl:transform']

        if _trivum_xsl and len(_trivum_xsl) > 0:
            for _itemxls in _trivum_xsl:
                # Exemplum: U0002||unescothes:NOP
                _temp1, _temp2 = _itemxls.split('||')
                tverb = _temp1
                tval_1, _nop_tval_2 = _temp2.split(':')
                if tverb.lower() == EXTRA_OPERATORS['STX']['hxl']:
                    # print('tval_1', tval_1)
                    result['rdfs:Container'][item]['trivium'][
                        'rdf_praefixum'] = tval_1

        _trivum_rdftypes = result['caput_originali_asa'][
            _trivium_indici]['extension']['r']['rdf:type']
        if _trivum_rdftypes and len(_trivum_rdftypes) > 0:
            # Exemplum: ['obo:BFO_0000029||0:NOP']
            for _itemtype in _trivum_rdftypes:
                # _temp1, _temp2 = _itemtype.split('||')
                # tverb = _temp1
                # tval_1, _nop_tval_2 = _temp2.split(':')
                if _itemtype not in \
                        result['rdfs:Container'][item]['trivium']['rdf:type']:
                    result['rdfs:Container'][item]['trivium'][
                        'rdf:type'].append(_itemtype)

    result = _aux_recalc_containers(result)

    result['caput_ad_columnae_i'] = bcp47_rdf_extension_caput_ad_columnae_i(
        result['caput_originali'],
        result['caput_originali_asa'])

    return result


# @TODO implement via command line specify which objective_bag.
# #     Defaults to 1 but we actually can get the others

def bcp47_rdf_extension_poc(
        header: List[str],
        data: List[List],
        objective_bag: str = '1',
        rdf_trivio_hxla: List[str] = None,
        _auxiliary_bags: List[str] = None,
        namespaces: List[dict] = None,
        rdf_sine_spatia_nominalibus: List = None,
        cum_antecessoribus: bool = False,
        rdf_ontologia_ordinibus: list = None,
        est_meta: bool = False,
        strictum: bool = True
) -> dict:
    """bcp47_rdf_extension_poc _summary_

    _extended_summary_

    Args:
        header (List[str]): _description_
        data (List[List]): _description_
        strictum (bool, optional): _description_. Defaults to True.

    Returns:
        dict: _description_

    Exemplōrum gratiā (et Python doctest, id est, automata testīs):
        (python3 -m doctest myscript.py)


    >>> namespaces = [
    ...    {'prefix': 'dc', 'iri': 'http://purl.org/dc/elements/1.1/'}
    ... ]

    >>> header_1 = ['qcc-Zxxx-r-sRDF-subject',
    ...             'eng-Latn-r-pDC-contributor-pDC-creator-pDC-publisher']
    >>> header_2 = ['qcc-Zxxx-r-sU2200-s1-snop',
    ...  'eng-Latn-r-pDC-pcontributor-ps1-pDC-pcreator-ps1-pDC-publisher-ps1']
    >>> data_1 = [['<http://vocabularies.unesco.org/thesaurus>',
    ...             'UNESCO']]
    >>> poc1 = bcp47_rdf_extension_poc(header_2, data_1, namespaces=namespaces)

    # >>> poc1['header_result']
    #'pskos-prefLabel'

    """
    if not data:
        data = []

    # raise NotImplementedError(header)
    result = {
        # 'caput': header,
        'caput_asa': {},
        # 'rdf:subject': None,
        # 'rdf:predicate': [],
        # 'rdf:object': None,
        # 'rdfs:Datatype': None,
        # '_unknown': [],
        # We always start with default prefixes
        # 'rdf_spatia_nominalibus': RDF_SPATIA_NOMINALIBUS,
        'rdf_spatia_nominalibus': {},
        'data': data,
        'rdf_triplis': [],
        # antecessōribus, pl, dativus, en.wiktionary.org/wiki/antecessor#Latin
        'antecessoribus_rdf_triplis': [],
        '_error': [],
    }

    # Enforce normalization for Wikidata predicates, we apply
    # _rdf_spatia_nominalibus_prefix_normali()
    _to_upper = ['wdata']

    if not rdf_sine_spatia_nominalibus or len(rdf_sine_spatia_nominalibus) == 0:
        rdf_sine_spatia_nominalibus = None

    # return {}

    # print('header', header)

    # header.pop()

    if est_meta:
        # Try harder to discover issues
        strictum = False

    meta = bcp47_rdf_extension_relationship(
        header, namespaces=namespaces, strictum=strictum)
    # raise ValueError(meta)
    result['caput_asa'] = meta
    # meta['data'] = data

    if objective_bag not in meta['rdfs:Container']:
        possible_bags = meta['rdfs:Container'].keys()
        if est_meta:
            result['_error'].append(
                'objective_bag({0})? possible <{1}>'.format(
                    objective_bag, possible_bags
                ))
            return result
        else:
            if len(data) > 2:
                # meta['data'] = meta['data'][0:1]
                data = data[0]

            raise SyntaxError(
                'objective_bag({0})? possible <{1}>: '
                'header <{2}> Meta <{3}>, data <{4}>'.format(
                    objective_bag, possible_bags, header, meta, data))

    # print(meta)
    # print('')
    # print(meta['rdfs:Container'][objective_bag])

    bag_meta = result['caput_asa']['rdfs:Container'][objective_bag]
    is_urn = bag_meta['trivium']['rdf_praefixum'].startswith('urn')
    is_urn_mdciii = bag_meta[
        'trivium']['rdf_praefixum'].startswith('urnmdciii')
    prefix_pivot = None

    # return bag_meta
    if not is_urn:
        # @todo solve this error later
        prefix_pivot = bag_meta['trivium']['rdf_praefixum']
        if prefix_pivot not in result['rdf_spatia_nominalibus']:
            if prefix_pivot not in RDF_SPATIA_NOMINALIBUS_EXTRAS:
                raise ValueError('prefix [{0}] not in [{1}]'.format(
                    prefix_pivot, RDF_SPATIA_NOMINALIBUS_EXTRAS
                ))
            result['rdf_spatia_nominalibus'][prefix_pivot] = \
                RDF_SPATIA_NOMINALIBUS_EXTRAS[prefix_pivot]

    xsl_transform_prefix = EXTRA_OPERATORS['STX']['hxl'].upper()
    xsl_transform_prefix_missing = set()
    for caput_originali_asa in result['caput_asa']['caput_originali_asa']:
        # print(caput_originali_asa)
        # print(caput_originali_asa['extension']['r']['xsl:transform'])

        if 'r' not in caput_originali_asa['extension']:
            continue

        _dt = caput_originali_asa['extension']['r']['rdfs:Datatype']
        if _dt is not None:
            _predicate, _ = _dt.split('||')
            prefix_datatype, _ = _predicate.split(':')
            if prefix_datatype not in result['rdf_spatia_nominalibus']:
                if prefix_datatype not in RDF_SPATIA_NOMINALIBUS_EXTRAS:
                    raise ValueError('prefix_datatype [{0}] not in [{1}]'.format(
                        prefix_datatype, RDF_SPATIA_NOMINALIBUS_EXTRAS
                    ))
                result['caput_asa']['rdf_spatia_nominalibus'][prefix_datatype] = \
                    RDF_SPATIA_NOMINALIBUS_EXTRAS[prefix_datatype]

            # raise ValueError(caput_originali_asa, result['rdf_spatia_nominalibus'])
        # result['rdf_spatia_nominalibus'][prefix_pivot]

        xsl_items = caput_originali_asa['extension']['r']['xsl:transform']
        if not xsl_items or len(xsl_items) == 0:
            continue
        # Example: ['U0002||unescothes:NOP', 'U001D||u007c:NOP']
        for xsx_item_meta in xsl_items:
            if not xsx_item_meta.startswith(xsl_transform_prefix):
                continue
            _temp1, _temp2 = xsx_item_meta.split('||')
            xsl_transform_prefix_missing.add(_temp2.split(':').pop(0))

        # pass
        # print(xsl_transform_prefix_missing)

    # raise ValueError(result['rdf_spatia_nominalibus'])
    if len(xsl_transform_prefix_missing) > 0:
        for xsl_item in xsl_transform_prefix_missing:
            # print('todo', xsl_item,
            #       result['caput_asa']['rdf_spatia_nominalibus'])
            if xsl_item in result['caput_asa']['rdf_spatia_nominalibus']:
                continue
            if xsl_item in RDF_SPATIA_NOMINALIBUS:
                result['caput_asa']['rdf_spatia_nominalibus'][xsl_item] = \
                    RDF_SPATIA_NOMINALIBUS[xsl_item]
            elif xsl_item in RDF_SPATIA_NOMINALIBUS_EXTRAS:
                result['caput_asa']['rdf_spatia_nominalibus'][xsl_item] = \
                    RDF_SPATIA_NOMINALIBUS_EXTRAS[xsl_item]
            else:
                raise ValueError('prefix [{0}] not in <{1}> or <{2}>'.format(
                    prefix_pivot, RDF_SPATIA_NOMINALIBUS_EXTRAS,
                    RDF_SPATIA_NOMINALIBUS_EXTRAS
                ))

    index_id = bag_meta['trivium']['index']
    triples_delayed = []

    # raise ValueError(meta['caput_originali_asa'])
    def _helper_aux_triple(
        bag_meta, bcp47_lang=None, subject=None,
        object_literal=None, object_tabula_indici: int = None,
        bcp47_privateuse: list = None, rdf_trivio_hxla: list = None
    ) -> Tuple:
        triples = []

        if subject is None or len(subject) == 0:
            return []

        # raise ValueError(bag_meta)
        # raise ValueError(rdf_trivio_hxla)
        # print(bag_meta)

        # @TODO: implement some way to discover implicit relations
        #        (up to one level). Would need scan table twice
        triples_delayed = []
        # This obviously is simplistic, because we can reference multiple
        # columns for same hashtags.

        is_literal = True

        value_separator = None
        value_prefixes = None
        if len(bag_meta["xsl:transform"]) > 0:
            for titem in bag_meta["xsl:transform"]:

                # print(titem)

                # tverb, tval_1, _nop_tval_2 = titem.split(':')
                _temp1, temp2 = titem.split('||')
                tverb = _temp1
                tval_1, _nop_tval_2 = temp2.split(':')

                tverb = tverb.lower()

                if tverb == EXTRA_OPERATORS['STX']['hxl']:
                    if value_prefixes is None:
                        value_prefixes = []
                    value_prefixes.append(tval_1)

                elif tverb == EXTRA_OPERATORS['GS']['hxl']:
                    # @TODO: check this at compile time
                    value_separator = CSVW_SEPARATORS[tval_1]
                else:
                    raise SyntaxError('{0} not in [{1}] context <{2}>'.format(
                        tverb, EXTRA_OPERATORS, [bag_meta, object_literal]
                    ))

        # if 'csvw:separator' in bag_meta and \
        #         len(bag_meta['csvw:separator']) > 0:
        #     value_separator = bag_meta['csvw:separator']

        # if 'prefix' in bag_meta and \
        #         len(bag_meta['prefix']) > 0:
        #     value_prefixes = bag_meta['prefix']

        for predicate_and_subject in bag_meta['rdf:predicate']:
            if not object_literal:
                continue

            _temp1, _temp2 = predicate_and_subject.split('||')
            predicate = _temp1
            for _item in _to_upper:
                if predicate.startswith(_item + ':'):
                    predicate = _rdf_spatia_nominalibus_prefix_normali(
                        predicate)

            if value_separator is not None and \
                object_literal.find(value_separator) > -1 and \
                    object_literal.find('\\' + value_separator) == -1:
                # @TODO: if a field have both separator and escaped separator
                #        we will fail. This would require a regex split, but
                #        but not implementing for now for performance reasons.
                object_results = object_literal.split(value_separator)
                # pass
            else:
                object_results = [object_literal]

            if value_prefixes is not None:
                is_literal = False
                for prefix in value_prefixes:
                    for index in range(len(object_results)):
                        if not object_results[index].startswith(prefix):
                            object_results[index] = '{0}:{1}'.format(
                                prefix, object_results[index]
                            )

            is_object_also_a_key = False

            if len(bag_meta['rdf:subject']) > 0:
                # TODO: deal with cases where a subject can be subject for
                #       more than a group
                for sitem in bag_meta['rdf:subject']:
                    if sitem.startswith((
                        FIRST_ORDER_LOGIC['∀']['hxl'].upper(),
                        FIRST_ORDER_LOGIC['∃']['hxl'].upper()
                    )):
                        is_object_also_a_key = True
                        _temp1, _temp2 = sitem.split('||')
                        _subject_group = _temp2.split(':').pop(0)
                        _subject_bag = result[
                            'caput_asa']['rdfs:Container'][_subject_group]
                        is_object_also_urnmcdiii = False
                        prefix_object_also_a_key = ''
                        if _subject_bag['trivium']['rdf_praefixum'] == \
                                'urnmdciii':
                            is_object_also_urnmcdiii = True
                        else:
                            prefix_object_also_a_key = \
                                _subject_bag['trivium']['rdf_praefixum']

            for item in object_results:
                # object_result = _helper_aux_object(item)
                is_namespaced_object = False

                if is_object_also_a_key and is_object_also_urnmcdiii:
                    object_result = '<urn:mcdiii:{0}>'.format(item)
                elif is_object_also_a_key and len(prefix_object_also_a_key) > 0:
                    object_result = '{0}:{1}'.format(
                        prefix_object_also_a_key, item)
                    is_namespaced_object = True
                elif 'rdfs:Datatype' in bag_meta and \
                        bag_meta['rdfs:Datatype']:
                    _temp1, _temp2 = bag_meta['rdfs:Datatype'].split('||')
                    object_result = '"{0}"^^{1}'.format(item, _temp1)
                elif not bcp47_lang.startswith('qcc'):
                    # @TODO escape " on item (if any)
                    object_result = '"{0}"@{1}'.format(
                        rdf_literal_escape(item), bcp47_lang)
                elif not is_literal:
                    # Example: prefixed result
                    object_result = item
                else:
                    # TODO: implement other data types
                    # object_result = _helper_aux_object(
                    #     item, object_tabula_indici)
                    object_result = '"{0}"'.format(item)

                # raise ValueError(item)
                if rdf_sine_spatia_nominalibus is not None \
                        and predicate.find(':') > -1:
                    _item_p_ns = predicate.split(':').pop(0)
                    if _item_p_ns in rdf_sine_spatia_nominalibus:
                        continue
                # raise ValueError(item, is_namespaced_object)
                if rdf_sine_spatia_nominalibus is not None \
                        and is_namespaced_object is True:
                    # raise ValueError(item)
                    _item_o_ns = object_result.split(':').pop(0)
                    if _item_o_ns in rdf_sine_spatia_nominalibus:
                        continue
                # raise ValueError(item)
                triples.append([subject, predicate, object_result])
                # continue

        # print('bag_meta', rdf_trivio_hxla, bcp47_privateuse)
        # raise ValueError(bag_meta)
        if rdf_trivio_hxla and len(rdf_trivio_hxla) > 0 \
                and bcp47_privateuse and len(bcp47_privateuse) > 0:
            # print('bag_meta', rdf_trivio_hxla, bcp47_privateuse)
            # print('bag_meta', bcp47_privateuse[0].startswith(rdf_trivio_hxla[0]))
            _trivio_hxla = None
            _trivio_hxla_list = []
            for _hxla in rdf_trivio_hxla:
                for _item in bcp47_privateuse:
                    if _item.startswith(_hxla):
                        # raise ValueError('foi', triples)
                        # _trivio_hxla = 'ix:{0}'.format(_item)
                        _trivio_hxla_list.append(_item)
                        # triples.append(['#<foi>', _trivio_hxla, _hxla])

            if len(_trivio_hxla_list) > 0:
                sorted(_trivio_hxla_list)
                if len(_trivio_hxla_list) == 1:
                    _trivio_hxla = 'ix:{0}'.format(_trivio_hxla_list[0])
                else:
                    _parts = []
                    # if using owl:unionOf (ix:item1 ix:item2 ix:item3)
                    # for _item in _trivio_hxla_list:
                    #     _parts.append('ix:{0}'.format(_item))
                    # _trivio_hxla = 'owl:unionOf ({0})'.format(' '.join(_parts))
                    for _item in _trivio_hxla_list:
                        _parts.append('ix:{0}'.format(_item))
                    _trivio_hxla = 'ix:{0}'.format(
                        '__'.join(_trivio_hxla_list))

            if _trivio_hxla is not None:
                triples_novo = []
                for _triple_item in triples:
                    _bn = '_:{0}'.format(uuid.uuid3(
                        uuid.NAMESPACE_DNS,
                        # _triple_item[0]
                        str('{}{}'.format(_triple_item[0], _triple_item[1]))
                    ))
                    triples_novo.append([
                        _triple_item[0],
                        _triple_item[1],
                        _bn
                    ])
                    triples_novo.append([
                        _bn,
                        _trivio_hxla,
                        _triple_item[2]
                    ])
                    # pass
                    # if len(triples):
                    #     print(_bn)
                # print('old', triples)
                triples = triples_novo
                # if len(triples):
                #     raise ValueError(triples, triples_novo)

        return triples, triples_delayed

    # raise ValueError(data, bag_meta)
    for linea in data:
        # triple = []
        # First pivot

        if not linea[index_id] or len(linea[index_id]) == 0:
            continue

        triple_rdfs_label_literal = None
        if is_urn_mdciii:

            triple_subject = '<urn:mdciii:{0}>'.format(linea[index_id])
            triple_rdfs_label_literal = linea[index_id]

            # ''.split()
            trivium_antecessori = linea[index_id].split(':')
            # This initialize
            trivium_antecessori.pop()
            if len(trivium_antecessori) > 0 and len(trivium_antecessori[0]) > 0:
                # trivium_antecessori = list(trivium_antecessori)
                numerordinatio_cum_antecessoribus(
                    trivium_antecessori,
                    rdf_ontologia_ordinibus=rdf_ontologia_ordinibus)

                _ns = ':'.join(trivium_antecessori)

                if len(_ns.strip()) > 0:
                    result['rdf_triplis'].append([
                        '<urn:mdciii:{0}()>'.format(_ns),
                        'skos:member',
                        triple_subject
                    ])
        elif is_urn:
            triple_subject = '<urn:{0}>'.format(linea[index_id])
            triple_rdfs_label_literal = linea[index_id]
        else:
            if rdf_sine_spatia_nominalibus is not None and \
                    prefix_pivot in rdf_sine_spatia_nominalibus:
                continue
            triple_subject = '{0}:{1}'.format(prefix_pivot, linea[index_id])

        if triple_rdfs_label_literal is not None:
            result['rdf_triplis'].append([
                triple_subject, 'rdfs:label',
                '"' + rdf_literal_escape(triple_rdfs_label_literal) + '"'
            ])

        for ego_typus in bag_meta['trivium']['rdf:type']:
            if not ego_typus.endswith('||0:NOP'):
                raise NotImplementedError('[{0}] <{1}>'.format(
                    ego_typus, bag_meta))
            _ego_typus = ego_typus.replace('||0:NOP', '')
            if rdf_sine_spatia_nominalibus is not None:
                _ego_typus.split(':').pop(0)
                if _ego_typus.split(':').pop(0) in rdf_sine_spatia_nominalibus:
                    continue
            triple = [triple_subject, 'a', _ego_typus]
            result['rdf_triplis'].append(triple)

        # Predicate for self is Subject here
        for predicate in bag_meta['trivium']['rdf:predicate']:
            # raise ValueError(predicate)
            triple = [triple_subject, 'a', predicate]
            result['rdf_triplis'].append(triple)

        # for referenced_by in bag_meta['indices_columnis']:
        for referenced_by in bag_meta['indices_cum_aliis']:
            if referenced_by == index_id:
                continue

            _bcp47lang = '{0}-{1}'.format(
                meta['caput_originali_asa'][referenced_by]['language'],
                meta['caput_originali_asa'][referenced_by]['script'],
            )
            object_literal = linea[referenced_by]
            object_tabula_indici = linea.index(object_literal)

            # _objects_parsed = _helper_aux_object(
            #     object_literal, bcp47lang=_bcp47lang, trivium=objective_bag)

            _privateuse = \
                meta['caput_originali_asa'][referenced_by]['privateuse']
            aux_triples, triples_delayed = _helper_aux_triple(
                meta['caput_originali_asa'][referenced_by]['extension']['r'],
                bcp47_lang=_bcp47lang,
                subject=triple_subject,
                object_literal=object_literal,
                object_tabula_indici=object_tabula_indici,
                bcp47_privateuse=_privateuse,
                rdf_trivio_hxla=rdf_trivio_hxla
            )

            if len(aux_triples) > 0:
                result['rdf_triplis'].extend(aux_triples)

    if rdf_sine_spatia_nominalibus is not None:
        _temp1 = {}
        for item_ns, item_iri in \
                result['caput_asa']['rdf_spatia_nominalibus'].items():
            if item_ns not in rdf_sine_spatia_nominalibus:
                _temp1[item_ns] = item_iri
        result['caput_asa']['rdf_spatia_nominalibus'] = _temp1
        # raise ValueError(result['rdf_spatia_nominalibus'])
        # pass
        # and is_namespaced_object is True:

    if 'rdf_spatia_nominalibus' in result:
        # @TODO remove this later.
        del result['rdf_spatia_nominalibus']

    # raise ValueError(result['caput_asa'].keys())
    if rdf_trivio_hxla and len(rdf_trivio_hxla) > 0:
        # We're likely to need upfront. So lets already add it
        result['caput_asa']['rdf_spatia_nominalibus']['ix'] = 'urn:hxl:vocab:a:ix:'

    result['antecessoribus_rdf_triplis'] = \
        NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS

    if est_meta:
        # @TODO: annex extra information
        # return bag_meta
        return result

    return result


class CodAbTabulae:
    """CodAbTabulae simple wrapper to work with raw P-Code/COD-AB tabular data

    - COD:
      - https://cod.unocha.org/
      - https://emergency.unhcr.org/entry/50306/common-operational-datasets-
        cods-and-fundamental-operational-datasets-fods
    - tabulae, f, s, dat./gen/, https://en.wiktionary.org/wiki/tabula
    - simplicī, m/f/n, s, dativus, https://en.wiktionary.org/wiki/simplex#Latin

    @TODO:
    - Check strategies to encode geo data in plain text.
      - https://giswiki.hsr.ch/GeoCSV
      -  https://github.com/hxl-team/HXL-Vocab/blob/master/Tools/static
         /hxl-geolocation-standard-draft.pdf
        - HXL circa 2012 already was using RDF to map information
      - https://www.ogc.org/standards/geosparql
      - https://github.com/hxl-team/HXL-Vocab/blob/master/Tools/static/hxl.ttl
    - SPARQL, Geo, etc
      - http://www.geosparql.org/

    """
    caput_originali: List[str] = None
    caput_hxl: List[str] = None
    caput_hxltm: List[str] = None
    caput_no1: List[str] = None
    caput_no1bcp47: List[str] = None
    data: List[list] = None
    dictionaria_linguarum: Type['DictionariaLinguarum'] = None
    ordo: int = 1
    numerordinatio_praefixo: str = None
    pcode_praefixo: str = None
    unm49: str = None
    experimentum_est: bool = False

    # identitās, f, s, nom., https://en.wiktionary.org/wiki/identitas#Latin
    # ex (+ ablative), https://en.wiktionary.org/wiki/ex#Latin
    # locālī, n, s, dativus, https://en.wiktionary.org/wiki/localis#Latin
    # identitas_locali_ex_hxl_hashtag: str = '#item+conceptum+codicem'
    identitas_locali_index: int = -1
    numerordinatio_indici: int = -2

    # https://en.wiktionary.org/wiki/originalis#Latin

    # caput_originali: List[str] = None
    # _caput_hxl: List[str] = None
    # objectīvō, n, dativus, https://en.wiktionary.org/wiki/dictionarium#Latin
    # objectīvō, n, dativus, https://en.wiktionary.org/wiki/objectivus#Latin
    _objectivo_dictionario: str = None

    # Columns marked merge from output
    _column_merge: list = []

    # Columns marked to discard from output
    _column_devnull: list = []

    def __init__(
        self,
        caput: list,
        data: list = None,
        ordo: int = 1,
        numerordinatio_praefixo: str = None,
        pcode_praefixo: str = None,
        unm49: str = None,
        experimentum_est: bool = False,
    ):
        """__init__"""
        self.caput_originali = caput
        self.data = data
        self.ordo = ordo
        self.numerordinatio_praefixo = numerordinatio_praefixo
        self.pcode_praefixo = pcode_praefixo
        self.unm49 = unm49
        self.experimentum_est = experimentum_est

    def imprimere(self) -> list:
        """imprimere /print/@eng-Latn

        Trivia:
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - fōrmātum, s, n, nominativus, https://en.wiktionary.org/wiki/formatus

        Args:
            formatum (str, optional): output format. Defaults to 'csv'.

        Returns:
            [list]: data, caput
        """
        # pylint: disable=chained-comparison
        # caput = []
        # data = []
        # if formatum:
        #     self._formatum = formatum

        # if formatum is None or formatum == 'csv':

        # print(self._objectivo_dictionario)

        # raise ValueError(self._objectivo_dictionario)

        caput = self.caput_originali
        data = self.data
        if self._objectivo_dictionario == 'hxl':
            caput = self.caput_hxl
        if self._objectivo_dictionario == 'hxltm':
            caput = self.caput_hxltm
        if self._objectivo_dictionario == 'no1':
            caput = self.caput_no1
        if self._objectivo_dictionario == 'no1bcp47':
            caput = self.caput_no1bcp47
            # caput = self.caput_no1

        # @TODO potentially re-arrange the order of columns on the result
        return caput, data

    def praeparatio(self, formatum: str):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """

        self._objectivo_dictionario = formatum

        self.caput_hxl = []
        self.dictionaria_linguarum = None

        # print('self.data', self.data)

        # Let's assume is a plain CSV (we skip if start with #)
        for index, res in enumerate(self.caput_originali):
            column = []
            if res and len(res) > 0 and not res.startswith('#') and \
                    len(self.data) > 0:
                for l_index, linea in enumerate(self.data):
                    if l_index > 500:
                        break
                    column.append(linea[index])

            self.caput_hxl.append(self.quod_hxl_de_caput_rei(res, column))

        # if formatum not in ['hxltm', 'no1']:
        if formatum not in ['hxltm', 'no1', 'no1bcp47']:
            return self
        # if formatum in ['hxltm', 'no1']:
        self.caput_hxltm = []

        self.dictionaria_linguarum = DictionariaLinguarum()

        for index, res in enumerate(self.caput_hxl):
            caput_novi = self.quod_hxltm_de_hxl_rei(res)

            if caput_novi == '#item+conceptum+codicem':
                self.identitas_locali_index = index
            # if caput_novi == '#item+conceptum+numerordinatio':
            #     self.numerordinatio_indici = index

            self.caput_hxltm.append(caput_novi)

        if self.identitas_locali_index < 0:
            ōrdinēs = range(self.ordo + 1)
            for ordo in reversed(ōrdinēs):
                self.praeparatio_identitas_numerodinatio(ordo)
            # self.praeparatio_identitas_numerodinatio(0)
            self.praeparatio_identitas_locali()

        # if formatum == 'no1':
        if formatum in ['no1', 'no1bcp47']:
            self.caput_no1 = []

            # self.dictionaria_linguarum = DictionariaLinguarum()

            for index, res in enumerate(self.caput_hxltm):
                caput_novi = self.quod_no1_de_hxltm_rei(res)

                # if caput_novi == '#item+conceptum+codicem':
                #     self.identitas_locali_index = index
                if caput_novi == '#item+conceptum+numerordinatio':
                    self.numerordinatio_indici = index

                self.caput_no1.append(caput_novi)

            if len(self.caput_no1) != len(set(self.caput_no1)):

                caput_novi__reversed = []
                # caput_no1 = list(self.caput_no1)
                caput_no1 = self.caput_no1[:]
                _len_caput = len(caput_no1)
                _index = _len_caput - 1
                _done = []
                while _index > -1:
                    # print('oi', _index, _totale)
                    _res = caput_no1[_index]
                    _totale = caput_no1.count(_res)
                    # print(self.caput_no1[_index])
                    if _totale > 1:
                        _done.append(_res)
                        _alt_now = (_totale - _done.count(_res))
                        # print('  >>repeated', _res, caput_no1)

                        # Only add altN to duplicates
                        if _alt_now == 0:
                            _res_novo = _res
                        else:
                            # @TODO make a more semantic way to have duplicated
                            #       items than ix_altN
                            _res_novo = '{0}+ix_alt{1}'.format(
                                _res, _alt_now
                            )
                        # caput_no1[caput_no1.index(_res)] = '__done' + str(index)
                        # caput_no1[caput_no1.index(_res)] = None
                        caput_novi__reversed.append(_res_novo)
                        # print('   > > repeated neo', _res, caput_no1)
                    else:
                        caput_novi__reversed.append(_res)
                    # _index += 1
                    _index = _index - 1
                # for _index in range(len(self.caput_no1)):
                # for _item in self.caput_no1:

                caput_novi__reversed.reverse()
                self.caput_no1 = caput_novi__reversed

            # self.caput_no1

        # if formatum in ['hxltm', 'no1'] and self.identitas_locali_index < 0:
        #     self.praeparatio_identitas_locali()

            if self.numerordinatio_indici < 0:
                self.praeparatio_numerordinatio()

        if formatum in ['no1bcp47']:
            self.caput_no1bcp47 = []

            # self.dictionaria_linguarum = DictionariaLinguarum()

            caput = self.caput_hxltm

            hashtag_numerordinatio = '#country+v_numerordinatio'
            if self.ordo > 0:
                hashtag_numerordinatio = '#adm{0}+v_numerordinatio'.format(
                    self.ordo)
            caput.insert(0, hashtag_numerordinatio)
            caput.insert(0, '#item+conceptum+numerordinatio')

            # for index, res in enumerate(self.caput_no1):
            for index, res in enumerate(caput):
                caput_novi = self.quod_no1bcp47_de_hxltm_rei(res, caput)

                # if caput_novi == '#item+conceptum+codicem':
                #     self.identitas_locali_index = index
                # if caput_novi == '#item+conceptum+numerordinatio':
                #     self.numerordinatio_indici = index

                self.caput_no1bcp47.append(caput_novi)

        # if formatum in ['hxltm', 'no1'] and self.identitas_locali_index < 0:
        #     self.praeparatio_identitas_locali()

            # if self.numerordinatio_indici < 0:
            #     self.praeparatio_numerordinatio()

        return self

    def praeparatio_identitas_locali(
            self, ordo: int = None, hxl_hashtag: str = '#item+conceptum+codicem'):
        """praeparatio_identitas_locali
        """
        if ordo is None:
            ordo = self.ordo
        pcode_index = None
        pcode_hashtag_de_facto = ''
        if ordo == 0:
            pcode_hashtag = [
                '#country+code+v_pcode', '#country+code+v_iso2',
                '#country+code+v_iso3166p1a2']
        else:
            pcode_hashtag = ['#adm{0}+code+v_pcode'.format(ordo)]

        for item in pcode_hashtag:
            if item in self.caput_hxltm:
                pcode_hashtag_de_facto = item
                pcode_index = self.caput_hxltm.index(item)
                break

        if pcode_index is None:
            raise SyntaxError(
                '{0} not in (hxltm)<{1}>/(hxl)<{2}>(csv){3}'.format(
                    pcode_hashtag, self.caput_hxltm,
                    self.caput_hxl, self.caput_originali
                ))
        # pcode_index = self.caput_hxltm.index(pcode_hashtag)

        # self.caput_hxltm.insert(0, '#item+conceptum+codicem')
        self.caput_hxltm.insert(0, hxl_hashtag)
        data_novis = []

        for linea in self.data:
            linea_novae = []
            pcode_completo = linea[pcode_index]
            if ordo == 0:
                # linea_novae.append(pcode_completo)  # Ex. BR
                linea_novae.append(self.unm49)  # Ex. 76 ex BR
            else:

                # Ex. 31 ad BR31
                pcode_numeri = pcode_completo.replace(self.pcode_praefixo, '')

                # Ex: Haiti admin3Pcode HT0111-01, HT0111-02, HT0111-03
                pcode_numeri = re.sub('[^0-9]', '', pcode_numeri)

                try:
                    linea_novae.append(int(pcode_numeri))
                except ValueError as err:
                    raise ValueError('<{0}:{1}> -> int({2})?? [{3}]'.format(
                        pcode_hashtag_de_facto,
                        pcode_completo,
                        pcode_numeri,
                        err
                    ))

            linea_novae.extend(linea)
            data_novis.append(linea_novae)

        self.data = data_novis

    def praeparatio_identitas_numerodinatio(
            self, ordo: int = None):
        """praeparatio_identitas_numerodinatio

        Change self.data and self.caput_hxltm

        Args:
            ordo (int, optional): _description_. Defaults to None.

        """
        # raise NotImplementedError
        hxl_hashtag = '#adm{0}+code+v_numerodinatio'.format(ordo)
        # if ordo is None:
        #     ordo = self.ordo
        pcode_index = None
        pcode_hashtag_de_facto = ''
        if ordo == 0:
            hxl_hashtag = '#country+code+v_numerodinatio'
            pcode_hashtag = [
                '#country+code+v_pcode', '#country+code+v_iso2',
                '#country+code+v_iso3166p1a2']
        else:
            pcode_hashtag = ['#adm{0}+code+v_pcode'.format(ordo)]

        for item in pcode_hashtag:
            if item in self.caput_hxltm:
                pcode_hashtag_de_facto = item
                pcode_index = self.caput_hxltm.index(item)
                break

        if pcode_index is None:
            raise SyntaxError(
                '{0} not in (hxltm)<{1}>/(hxl)<{2}>(csv){3}'.format(
                    pcode_hashtag, self.caput_hxltm,
                    self.caput_hxl, self.caput_originali
                ))
        # pcode_index = self.caput_hxltm.index(pcode_hashtag)

        # self.caput_hxltm.insert(0, '#item+conceptum+codicem')
        self.caput_hxltm.insert(0, hxl_hashtag)
        data_novis = []

        for linea in self.data:
            linea_novae = []
            pcode_completo = linea[pcode_index]
            if ordo == 0:
                # raise ValueError(ordo)
                _numerordinatio = '{0}:{1}:{2}'.format(
                    self.numerordinatio_praefixo,
                    self.unm49,
                    "0",
                )
                # linea_novae.append(pcode_completo)  # Ex. 76 for BR
                linea_novae.append(_numerordinatio)  # Ex. 76 for BR
            else:

                # Ex. 31 ad BR31
                pcode_numeri = pcode_completo.replace(self.pcode_praefixo, '')

                # Ex: Haiti admin3Pcode HT0111-01, HT0111-02, HT0111-03
                pcode_numeri = re.sub('[^0-9]', '', pcode_numeri)

                try:
                    pcode_numeri = int(pcode_numeri)
                    _numerordinatio = '{0}:{1}:{2}:{3}'.format(
                        self.numerordinatio_praefixo,
                        self.unm49,
                        ordo,
                        str(pcode_numeri),
                    )

                    linea_novae.append(_numerordinatio)
                except ValueError as err:
                    raise ValueError('<{0}:{1}> -> int({2})?? [{3}]'.format(
                        pcode_hashtag_de_facto,
                        pcode_completo,
                        pcode_numeri,
                        err
                    ))

            linea_novae.extend(linea)
            data_novis.append(linea_novae)

        self.data = data_novis

    def praeparatio_numerordinatio(self):
        """numerordinatio
        """
        # print('oi')
        identitas_locali_index = self.caput_hxltm.index(
            '#item+conceptum+codicem')
        hashtag_numerordinatio = '#country+v_numerordinatio'
        if self.ordo > 0:
            hashtag_numerordinatio = '#adm{0}+v_numerordinatio'.format(
                self.ordo)
        # self.caput_no1.insert(0, hashtag_numerordinatio)
        self.caput_no1.insert(0, '#item+conceptum+numerordinatio')
        data_novis = []

        for linea in self.data:
            _identitas_locali = linea[identitas_locali_index]
            # Special case: replace ISO2/ISO3 with UN m49 if is country
            if self.ordo == 0:
                # _identitas_locali = self.unm49
                # adm0 is an special case: the unm49 already is at the base
                # so we as only one option can exist ("0"), no redundancy need
                _numerordinatio = '{0}:{1}:{2}'.format(
                    self.numerordinatio_praefixo,
                    self.unm49,
                    str(self.ordo)
                )
            else:
                _numerordinatio = '{0}:{1}:{2}:{3}'.format(
                    self.numerordinatio_praefixo,
                    self.unm49,
                    str(self.ordo),
                    _identitas_locali,
                )
            # linea_novae = [_numerordinatio, _numerordinatio]
            linea_novae = [_numerordinatio]
            linea_novae.extend(linea)

            data_novis.append(linea_novae)

        self.data = data_novis

    def praeparatio_numerordinatio_bcp47(self):
        """numerordinatio_bcp47
        """

        self.praeparatio_numerordinatio()

        self.caput_no1bcp47 = self.caput_no1
        # @TODO ...

        # identitas_locali_index = self.caput_hxltm.index(
        #     '#item+conceptum+codicem')
        # self.caput_no1.insert(0, '#item+conceptum+numerordinatio')
        # data_novis = []

        # for linea in self.data:
        #     linea_novae = ['{0}:{1}:{2}:{3}'.format(
        #         self.numerordinatio_praefixo,
        #         self.unm49,
        #         str(self.ordo),
        #         linea[identitas_locali_index],

        #     )]
        #     linea_novae.extend(linea)
        #     data_novis.append(linea_novae)

        # self.data = data_novis

        return self.praeparatio_numerordinatio()

    @staticmethod
    def quod_columna_alphabeto_orini(column: list = None) -> str:
        """quod_columna_alphabeto_orini

       For a list of values, the non empty ones which have letters, if
       the number of letters are consistent, how many? Used to know
       if data columns could be labeled, as example, as ISO 3661-2 or
       ISO 3661-3.

        Args:
            column (list, optional):. Defaults to None.

        Returns:
            str: _description_
        """
        ordo = None
        _ordo = set()
        if column and len(column) > 0:
            for item in column:
                alpha = re.sub('[0-9]', '', item)
                if alpha and len(alpha) > 1:
                    _ordo.add(len(alpha))
            if len(_ordo) == 1:
                ordo = _ordo.pop()

        # print('  ordo', ordo, column)
        return ordo

    @staticmethod
    def quod_hxl_de_caput_rei(res: str, data_exemplis: list = None) -> str:
        """quod_hxl_de_caput_rei

        Args:
            res (str):
            data_exemplis (list):

        Returns:
            str:
        """
        res_originale = res
        if not res or len(res) == 0 or res.startswith('#'):
            # pipelining hxl again? let's not change it.
            return res

        # @see https://data.humdata.org/tools/hxl-example/
        praefīxum = ''
        suffīxum = ''
        # ōrdō  = ''
        lingua = ''
        # iso = ''
        res = res.lower()
        numerus = re.sub('[^0-9]', '', res)
        geo_ōrdinī = ''
        nomen_ōrdinī = ''
        if len(numerus) == 2:
            geo_ōrdinī = numerus[0]
            nomen_ōrdinī = numerus[1]
        if len(numerus) == 1:
            geo_ōrdinī = numerus

        if len(numerus) > 2:
            # Something weird like admin2pcode2016. Better human review
            return ''

        if res in ['date', 'validon', 'validto']:
            praefīxum = '#date'
            if res == 'date':
                suffīxum = '+start'
            if res == 'validon':  # validOn
                suffīxum = '+updated'
            if res == 'validto':  # validTo
                suffīxum = '+end'
        elif res.startswith('adm') and len(geo_ōrdinī) == 1:
            if geo_ōrdinī == '0':
                praefīxum = '#country'
            else:
                praefīxum = '#adm{0}'.format(geo_ōrdinī)

            if res.find('pcode') > -1:
                if geo_ōrdinī == '0':
                    ordo_codici = CodAbTabulae.quod_columna_alphabeto_orini(
                        data_exemplis)
                    if ordo_codici and ordo_codici in [2, 3]:
                        suffīxum = '+code+v_iso{0}'.format(ordo_codici)
                    elif ordo_codici is None:
                        # Weird case: no data at all on adm0. Lets force
                        # as ISO 3661p1a2
                        suffīxum = '+code+v_iso2'
                    else:
                        suffīxum = '+code'
                else:
                    suffīxum = '+code+v_pcode'

            elif res.find('code') > -1:
                suffīxum = '+code'

            elif res.find('name') > -1:
                if res.find('nameref') > -1 or res.find('nameref') > -1:
                    suffīxum = '+name+preferred'
                if res.find('altname') > -1:
                    suffīxum = '+name+alt{0}'.format(nomen_ōrdinī)
                else:
                    suffīxum = '+name'
                if res.find('_') > -1:
                    _temp = res.split('_')
                    if len(_temp) == 2:
                        lingua = '+i_' + _temp[1]

        if len(praefīxum) == 0:
            # Something not planned was tagged. Labeling as Meta.
            praefīxum = '#meta'

        # use case: ukr.xlsx
        #   admin2Name_en	admin2Name_ua	admin2Name_ru	admin2Pcode
        #   admin2ClassCode	admin2ClassType	admin2PoliticalType
        #   admin2pcode2016	admin1ClassType	admin1ClassCode
        if len(praefīxum) > 0 and praefīxum != '#date' and len(suffīxum) == 0:
            splitted = re.sub(
                '([A-Z][a-z]+)', r' \1',
                re.sub('([A-Z]+)', r' \1', res_originale)).split()
            praefīxum = '#meta'
            for _item in splitted:
                suffīxum = suffīxum + '+' + _item.lower()

            # Something weird happened

        return '{0}{1}{2}'.format(praefīxum, suffīxum, lingua)

    def quod_hxltm_de_hxl_rei(self, hxlhashtag: str) -> str:
        """quod_hxltm_de_hxl_rei

        Args:
            res (str):

        Returns:
            str:
        """
        # lingua = ''

        if not hxlhashtag or len(hxlhashtag) == 0 or \
                not hxlhashtag.startswith('#'):
            return ''

        bcp47_rei = qhxl_hxlhashtag_2_bcp47(hxlhashtag, True)
        # raise ValueError('oi1 {0} {1}'.format(hxlhashtag, bcp47_rei))
        # hxlhashtag = '#x_todo+' + hxlhashtag.replace('#', '')
        if bcp47_rei and len(bcp47_rei) < 9:  # < 'zzz-Zzzz'
            _bcp47_rei_copia = bcp47_rei
            if bcp47_rei == 'ua':
                # Ukrainian ISO 639-2 is 'uk', not 'ua' (code of the country)
                # https://iso639-3.sil.org/code/ukr
                bcp47_rei = 'uk'

            res1603_1_51 = self.dictionaria_linguarum.quod(
                bcp47_rei, abecedarium_incognito=True)
            if res1603_1_51 and \
                    '#item+rem+i_qcc+is_zxxx+ix_hxla' in res1603_1_51:
                ix_hxla = res1603_1_51['#item+rem+i_qcc+is_zxxx+ix_hxla']
                if ix_hxla:
                    hxlhashtag = hxlhashtag.replace(
                        '+i_{0}'.format(_bcp47_rei_copia), ix_hxla)

        if hxlhashtag.startswith('#country'):
            hxlhashtag_parts = hxlhashtag.split('+')
            if 'v_iso2' in hxlhashtag_parts:
                hxlhashtag_parts[hxlhashtag_parts.index(
                    'v_iso2')] = 'v_iso3166p1a2'
                hxlhashtag = '+'.join(hxlhashtag_parts)
            if 'v_iso3' in hxlhashtag_parts:
                hxlhashtag_parts[hxlhashtag_parts.index(
                    'v_iso2')] = 'v_iso3166p1a3'
                hxlhashtag = '+'.join(hxlhashtag_parts)

        return hxlhashtag

    def quod_no1_de_hxltm_rei(self, hxlhashtag: str) -> str:
        """quod_no1_de_hxltm_rei

        Args:
            res (str):

        Returns:
            str:
        """
        # lingua = ''

        if not hxlhashtag or len(hxlhashtag) == 0 or \
                not hxlhashtag.startswith('#'):
            return ''

        _hxl = HXLHashtagSimplici(hxlhashtag).praeparatio()

        return _hxl.quod_numerordinatio()

        # print('hxlhashtag', hxlhashtag)
        # Time-related hashtags
        # @see https://www.wikidata.org/wiki/Wikidata:List_of_properties/time
        # # https://www.wikidata.org/wiki/Property:P571
        # # inception (P571)
        # # time when an entity begins to exist
        # # equivalent property
        # #  - https://schema.org/dateCreated
        # #  - http://id.nlm.nih.gov/mesh/vocab#dateCreated
        # if hxlhashtag == '#date+start':
        #     return '#item+rem+i_qcc+is_zxxx+ix_wdatap571'

        # # P729 service entry
        # # date or point in time on which a piece or class of equipment
        # # https://www.wikidata.org/wiki/Property:P729
        # if hxlhashtag == '#date+start':
        #     return '#item+rem+i_qcc+is_zxxx+ix_wdatap729'

        # # P729 service entry
        # # date or point in time on which a piece or class of equipment
        # # https://www.wikidata.org/wiki/Property:P730
        # if hxlhashtag == '#date+end':
        #     return '#item+rem+i_qcc+is_zxxx+ix_wdatap730'

        # publication date (P577)
        # date or point in time when a work was first published or released
        # https://www.wikidata.org/wiki/Property:P577
        if hxlhashtag == '#date+start':
            return '#item+rem+i_qcc+is_zxxx+ix_wdatap577'

        # discontinued date (P2669)
        # date that the availability of a product was discontinued;
        # see also "dissolved, abolished or demolished" (P576)
        # https://www.wikidata.org/wiki/Property:P2669
        if hxlhashtag == '#date+end':
            return '#item+rem+i_qcc+is_zxxx+ix_wdatap2669'

        ## retrieved (P813)
        # - https://www.wikidata.org/wiki/Property:P813
        # - https://wiki.openstreetmap.org/wiki/Key:check_date
        if hxlhashtag == '#date+updated':
            return '#item+rem+i_qcc+is_zxxx+ix_wdatap813'

        # ISO 3166-1 alpha-2 code (P297)
        # https://www.wikidata.org/wiki/Property:P297
        if hxlhashtag in [
                '#country+code+v_iso2', '#country+code+v_iso3166p1a2']:
            # @TODO: make qualifier if this is not adm0
            return '#item+rem+i_qcc+is_zxxx+ix_wdatap297'

        # @TODO '#adm1+code+v_pcode' likely to be alpha 2 (needs check data)

        return self.quod_hxltm_de_hxl_rei(hxlhashtag)

    @staticmethod
    def quod_no1bcp47_de_hxltm_rei(hxlhashtag: str, caput: List) -> str:
        """quod_no1bcp47_de_hxltm_rei

        Args:
            res (str):

        Returns:
            str:
        """
        # @TODO this is no1 in HXL (change later for BCP47)

        if not hxlhashtag or len(hxlhashtag) == 0 or \
                not hxlhashtag.startswith('#'):
            return ''

        # if hxlhashtag in BCP47_EX_HXL:
        #     return BCP47_EX_HXL[hxlhashtag]['bcp47']

        if hxlhashtag in BCP47_EX_HXL:
            # 'hxl': '#item+rem+i_qcc+is_zxxx+rdf_a_mdciii_latcodicem',
            # 'hxltm': '#item+conceptum+codicem'
            return BCP47_EX_HXL[hxlhashtag]['hxltm']

        _hxl = HXLHashtagSimplici(hxlhashtag).praeparatio()

        # @TODO make exported formats of higher admX also export previous
        #       levels (so the RDF would work to make the relations)

        # print(hxlhashtag)
        # print("\t\t hashtag", _hxl.hashtag)
        # print("\t\t quod_ad_rdf", _hxl.quod_ad_rdf(''))
        # print("\t\t hxlattrs", _hxl.quod_ad_rdf('hxlattrs'))
        # print("\t\t quod_numerordinatio", _hxl.quod_numerordinatio())

        # raise ValueError( _hxl.hashtag)

        return _hxl.hashtag

        # return hxlhashtag + '+oi'
        return hxlhashtag


def configuratio(
        archivum_configurationi: List[str] = None,
        strictum: bool = True
) -> dict:
    """cōnfigūrātiō

    Args:
        archivum_configurationi (str, optional):

    Returns:
        (dict):
    """
    # archivae = ARCHIVUM_CONFIGURATIONI_DEFALLO
    # if archivum_configurationi is not None:
    #     if not os.path.exists(archivum_configurationi):
    #         raise FileNotFoundError(
    #             'archivum_configurationi {0}'.format(
    #                 archivum_configurationi))
    #     archivae.append(archivum_configurationi)

    for item in archivum_configurationi:
        if os.path.exists(item):
            with open(item, "r") as read_file:
                datum = yaml.safe_load(read_file)
                return datum
    if strictum:
        raise FileNotFoundError(
            'archivum_configurationi {0}'.format(
                archivum_configurationi))
    return None


def csv_imprimendo(
        caput: list = None, data: list = None, punctum_separato: str = ",",
        archivum_trivio: str = None):
    if archivum_trivio:
        raise NotImplementedError('{0}'.format(archivum_trivio))
    # imprimendō, v, s, dativus, https://en.wiktionary.org/wiki/impressus#Latin

    _writer = csv.writer(sys.stdout, delimiter=punctum_separato)
    if caput and len(caput) > 0:
        _writer.writerow(caput)
    _writer.writerows(data)

    # print(type(data), data)
    # print(type(data), caput)


class DictionariaLinguarum:
    """DictionariaLinguarum 1603_1_51
    """

    def __init__(self, fontem_archivum: str = None):
        if fontem_archivum:
            self.D1613_1_51_fontem = fontem_archivum
        else:
            self.D1613_1_51_fontem = NUMERORDINATIO_BASIM + \
                "/1603/1/51/1603_1_51.no1.tm.hxl.csv"

        # self.codex = codex
        self.dictionaria_codex = self._init_dictionaria()

        # print('oioioi', self.dictionaria_codex )

    def _init_dictionaria(self):

        datum = {}
        with open(self.D1613_1_51_fontem) as file:
            csv_file = csv.DictReader(file)
            # return list(tsv_file)
            for conceptum in csv_file:
                # print('conceptum', conceptum)
                int_clavem = int(conceptum['#item+conceptum+codicem'])
                datum[int_clavem] = {}
                for clavem, rem in conceptum.items():
                    datum[int_clavem][clavem] = rem
                    # if not clavem.startswith('#item+conceptum+codicem'):
                    #     datum[int_clavem][clavem] = rem
        return datum

    def imprimere(
            self, linguam: list = None,
            codex: Type['Codex'] = None) -> list:  # type: ignore
        """imprimere /print/@eng-Latn

        Trivia:
        - cōdex, m, s, (Nominative), https://en.wiktionary.org/wiki/codex#Latin
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - pāginae, f, pl, (Nominative), https://en.wiktionary.org/wiki/pagina

        Returns:
            [list]:
        """
        # pylint: disable=too-many-locals,too-many-statements

        resultatum = []
        resultatum_corpus = []
        resultatum_corpus_totale = 0

        # resultatum.append('')
        # resultatum.append(str(codex.usus_linguae_concepta))
        # resultatum.append('')

        linguam_clavem = []
        if linguam:
            for item in linguam:
                linguam_clavem.append(
                    item.replace('#item+rem', '')
                )
        # resultatum_corpus.append(linguam_clavem)
        # resultatum_corpus.append(len(linguam_clavem))
        for _clavem, lineam in self.dictionaria_codex.items():

            if len(linguam_clavem) > 0:
                if lineam['#item+rem+i_qcc+is_zxxx+ix_hxla'] not in \
                        linguam_clavem:
                    continue

            clavem_i18n = lineam['#item+rem+i_qcc+is_zxxx+ix_uid']
            item_text_i18n = lineam['#item+rem+i_lat+is_latn']
            ix_glottocode = lineam['#item+rem+i_qcc+is_zxxx+ix_glottocode']
            ix_iso639p3a3 = lineam['#item+rem+i_qcc+is_zxxx+ix_iso639p3a3']
            ix_wikiq = lineam['#item+rem+i_qcc+is_zxxx+ix_wikiq+ix_linguam']
            usus = 0
            if clavem_i18n in codex.usus_linguae_concepta:
                usus = codex.usus_linguae_concepta[clavem_i18n]

            if len(ix_glottocode) > 0:
                ix_glottocode = \
                    "https://glottolog.org/resource" + \
                    "/languoid/id/{0}[{0}]".format(
                        ix_glottocode)

            if len(ix_iso639p3a3) > 0:
                ix_iso639p3a3 = \
                    "https://iso639-3.sil.org/code/{0}[{0}]".format(
                        ix_iso639p3a3)
            if len(ix_wikiq) > 0:
                ix_wikiq = \
                    "https://www.wikidata.org/wiki/{0}[{0}]".format(
                        ix_wikiq)
            # resultatum_corpus.append(str(lineam))
            # resultatum_corpus.append(linguam)
            # resultatum_corpus.append(
            # resultatum_corpus.append("| {0}".format(clavem_i18n))
            resultatum_corpus.append("|")
            resultatum_corpus.append("{0}".format(clavem_i18n))
            resultatum_corpus.append("|")
            resultatum_corpus.append(
                "{0}\n+++<br>+++\n{1}\n+++<br>+++ {2}".format(
                    ix_glottocode, ix_iso639p3a3, ix_wikiq))
            # resultatum_corpus.append("| {0}".format(ix_glottocode))
            # resultatum_corpus.append("| {0}".format(ix_iso639p3a3))
            # resultatum_corpus.append("| {0}".format(ix_wikiq))
            # resultatum_corpus.append("| {0}".format(item_text_i18n))
            resultatum_corpus.append("|")
            resultatum_corpus.append("{0}".format(item_text_i18n))
            resultatum_corpus.append("|")
            resultatum_corpus.append("{0}".format(usus))
            resultatum_corpus.append('')
            resultatum_corpus_totale += 1

        if resultatum_corpus:
            resultatum.append("")

            # resultatum.append("=== Linguae in cōdex: {0}".format(
            #     resultatum_corpus_totale))
            resultatum.append("==== Rēs linguālibus: {0}".format(
                resultatum_corpus_totale))

            # cōdex, m, s, (nominative)
            # tōtālis, m/f, s, (Nominative)
            # linguae, f, s, (Dative)
            # resultatum.append(
            #     "Tōtālis linguae in cōdex: {0}".format(
            #         resultatum_corpus_totale))
            resultatum.append("")

            resultatum.append('[%header,cols="15h,25a,~,17"]')
            resultatum.append('|===')
            # https://en.wiktionary.org/wiki/latinus#Latin
            # nōmina, n, pl, (Nominative)
            #     shttps://en.wiktionary.org/wiki/nomen#Latin
            # "nōmen Latīnum"
            # https://en.wiktionary.org/wiki/Latinus#Latin
            # resultatum.append(
            #     "| <span lang='la'>Cōdex<br>linguae</span> | "
            #     "<span lang='la'>Glotto<br>cōdicī</span> | "
            #     "<span lang='la'>ISO<br>639-3</span> | "
            #     "<span lang='la'>Wiki QID<br>cōdicī</span> | "
            #     "<span lang='la'>Nōmen Latīnum</span> |")
            # resultatum.append("| --- | --- | --- | --- | --- |")
            # resultatum.append("| Cōdex linguae")
            resultatum.append("|")
            resultatum.append("Cōdex linguae")
            # resultatum.append("a| Glotto cōdicī ++<br>++ ISO 639-3 ++<br>++ Wiki QID cōdicī")
            resultatum.append("|")
            resultatum.append(
                "Glotto cōdicī +++<br>+++ ISO 639-3 +++<br>+++ Wiki QID cōdicī")
            # resultatum.append("| ISO 639-3")
            # resultatum.append("| Wiki QID cōdicī")
            resultatum.append("|")
            resultatum.append("Nōmen Latīnum")
            resultatum.append("|")
            resultatum.append("Concepta")
            resultatum.append('')
            resultatum.extend(resultatum_corpus)
            resultatum.append('|===')
            resultatum.append("")

            # concepta, f, pl, Nominative,
            #    https://en.wiktionary.org/wiki/conceptus

        return resultatum

    def quod(
        self,
        terminum: str,
        #  factum: str = '#item+rem+i_lat+is_latn',
        # clavem: str = None
        # clavem: str = None,
        abecedarium_incognito=False,
        _data_exemplis: list = None
    ) -> dict:
        """quod_

        Search 1603_1_51 full from a raw term, like:
          - '+i_rus+is_cyrl'
          - '__i_rus__is_cyrl'
          - 'rus-Cyrl'

        Trivia:
        - abecedārium, n, s, nom., en.wiktionary.org/wiki/incognitus#Latin
        - incognitō, n, s, dat. https://en.wiktionary.org/wiki/incognitus#Latin

        Args:
            terminum (str): _description_
            clavem (str, optional): _description_. Defaults to None.

        Returns:
            dict: _description_
        """
        _clavem = [
            '#item+rem+i_qcc+is_zxxx+ix_hxla',
            # '#item+rem+i_qcc+is_zxxx+ix_hxla',
            '#item+rem+i_qcc+is_zxxx+ix_csvsffxm',
            '#item+rem+i_qcc+is_zxxx+ix_uid'
        ]
        if abecedarium_incognito:
            _clavem.extend([
                # Glottocode Langoid
                '#item+rem+i_qcc+is_zxxx+ix_glottocode',
                # Wikidata QID; language
                '#item+rem+i_qcc+is_zxxx+ix_wikiq+ix_linguam',
                # ISO 639-3 (lang alpha 3)
                '#item+rem+i_qcc+is_zxxx+ix_iso639p3a3',
                # ISO 639-2 or ISO 639-3 (used on Wikidata/Wikipedia)
                '#item+rem+i_qcc+is_zxxx+ix_wikilngm',
            ])
        # _clavem = clavem_defallo if clavem is None else [clavem]
        # _clavem = clavem_defallo

        if _data_exemplis and len(_data_exemplis) > 0:
            raise NotImplementedError(
                '_data_exemplis [{0}] [[@TODO check Unicode range '
                'to infer script]]'.format(
                    _data_exemplis))

        # print(self.dictionaria_codex.items())
        # raise ValueError('b')

        for item in _clavem:
            # print('item', item)
            for _k, linguam in self.dictionaria_codex.items():
                if linguam['#item+conceptum+codicem'].startswith('0_'):
                    continue
                # print('linguam', terminum, item, linguam[item], linguam)
                # print('')
                # if terminum.find(linguam[item]) > -1:
                if terminum == linguam[item]:
                    # return linguam[factum]
                    # raise ValueError([terminum, linguam])
                    return linguam

        return None


def de_dotted(dotted_key: str,  # pylint: disable=invalid-name
              default: Any = None, fontem: dict = None) -> Any:
    """
    Trivia: dē, https://en.wiktionary.org/wiki/de#Latin
    Examples:
        >>> exemplum = {'a': {'a2': 123}, 'b': 456}
        >>> de_dotted('a.a2', fontem=exemplum)
        123

    Args:
        dotted_key (str): Dotted key notation
        default ([Any], optional): Value if not found. Defaults to None.
        fontem (dict): An nested object to search
    Returns:
        [Any]: Return the result. Defaults to default
    """

    keys = dotted_key.split('.')
    return reduce(
        lambda d, key: d.get(
            key) if d else default, keys, fontem
    )


class DictionariaInterlinguarum:
    """DictionariaInterlinguarum 1603_1_7
    """

    def __init__(self, fontem_archivum: str = None):
        if fontem_archivum:
            self.D1613_1_7_fontem = fontem_archivum
        else:
            self.D1613_1_7_fontem = NUMERORDINATIO_BASIM + \
                "/1603/1/7/1603_1_7.no1.tm.hxl.csv"

        self.dictionaria = self._init_dictionaria()

    def _init_dictionaria(self):

        datum = {}
        with open(self.D1613_1_7_fontem) as file:
            csv_file = csv.DictReader(file)
            # return list(tsv_file)
            for conceptum in csv_file:
                # print('conceptum', conceptum)
                int_clavem = int(conceptum['#item+conceptum+codicem'])
                datum[int_clavem] = {}
                for clavem, rem in conceptum.items():
                    datum[int_clavem][clavem] = rem
                    # if not clavem.startswith('#item+conceptum+codicem'):
                    #     datum[int_clavem][clavem] = rem
        return datum

    def formatum_nomen(
            self, clavem: str,
            _objectivum_linguam: str = None,
            _auxilium_linguam: list = None) -> str:
        # fōrmātum, f, s, (Nominative) https://en.wiktionary.org/wiki/formatus

        meta_langs = [
            '#item+rem+i_mul+is_zyyy',
            '#item+rem+i_lat+is_latn'
        ]

        ix_clavem = clavem.replace('#item+rem+i_qcc+is_zxxx+', '')
        terminum = None
        dictionaria_res = self.quod(ix_clavem)
        if dictionaria_res:
            terminum = qhxl(dictionaria_res, meta_langs)
            # terminum = terminum + ' lalala'

        return terminum if terminum else ix_clavem

    # def formatum_res_facto(
    #         self, res: dict, clavem: str,
    #         _objectivum_linguam: str = None,
    #         _auxilium_linguam: list = None
    # ) -> str:
    @staticmethod
    def formatum_res_facto(
            res: dict, clavem: str,
            _objectivum_linguam: str = None,
            _auxilium_linguam: list = None
    ) -> str:
        # fōrmātum, f, s, (Nominative) https://en.wiktionary.org/wiki/formatus
        # TODO: this still need improvement
        # return res[clavem]
        return res_interlingualibus_formata(res, clavem)
        # return res[clavem] + '[' + clavem + ']'

    def imprimere(self, linguam: list = None) -> list:
        """imprimere /print/@eng-Latn

        @DEPRECATED using methodo direct from codex

        Trivia:
        - cōdex, m, s, (Nominative), https://en.wiktionary.org/wiki/codex#Latin
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - pāginae, f, pl, (Nominative), https://en.wiktionary.org/wiki/pagina

        Returns:
            [list]:
        """
        resultatum = []
        # resultatum_corpus = []
        resultatum_corpus_totale = 0
        resultatum_corpus_obj = []
        linguam_clavem = []
        if linguam:
            for item in linguam:
                linguam_clavem.append(
                    item.replace('#item+rem+i_qcc+is_zxxx+', '')
                )
        # resultatum_corpus.append(linguam_clavem)
        # resultatum_corpus.append(len(linguam_clavem))
        for clavem, lineam in self.dictionaria.items():

            if len(linguam_clavem) > 0:
                if lineam['#item+rem+i_qcc+is_zxxx+ix_hxlix'] not in \
                        linguam_clavem:
                    continue

            neo_lineam = {}
            for k, v in lineam.items():
                # neo_lineam[k] = v
                if v:
                    neo_lineam[k] = v

            resultatum_corpus_obj.append(neo_lineam)

            resultatum_corpus_totale += 1

        if resultatum_corpus_obj:

            resultatum.append("==== Rēs interlinguālibus: {0}".format(
                resultatum_corpus_totale))

            # @TODO: 1603_1_99_10_11; {0} = code complete for this book
            resultatum.extend(descriptio_tabulae_de_lingua(
                'Lingua Anglica (Abecedarium Latinum)',
                'The result of this section is a preview. '
                'We\'re aware it is not well formatted for a book format. '
                'Sorry for the temporary inconvenience.'
            ))

            # About description lists

            # Two columns here (requires some changes on CSS)
            #  https://github.com/asciidoctor/asciidoctor-pdf/issues/327
            #  https://github.com/Mogztter/asciidoctor-web-pdf/tree/main
            #  /examples/cheat-sheet

            # import pprint
            resultatum.append("")
            for res in resultatum_corpus_obj:
                resultatum.append("")
                # resultatum.append('===== {0} '.format(
                #     res['#item+conceptum+numerordinatio'])
                # )
                # resultatum.append('**{0}**'.format(
                #     res['#item+conceptum+numerordinatio'])
                # )
                resultatum.append("")
                resultatum.append("{0}::".format(
                    res['#item+rem+i_lat+is_latn']))

                for clavem, rem_textum in res.items():
                    if clavem in [
                        '#item+conceptum+numerordinatio',
                        '#item+conceptum+codicem',
                        '#status+conceptum+definitionem',
                        '#status+conceptum+codicem',
                        '#item+rem+i_lat+is_latn',
                    ]:
                        continue

                    if len(rem_textum) < 1:
                        continue

                    resultatum.append("{0}::: {1}".format(clavem, rem_textum))
                    # datum[int_clavem][clavem] = rem

                # resultatum.append("")
                # resultatum.append("[source,json]")
                # resultatum.append("----")
                # resultatum.append(json.dumps(
                #     res, indent=4, sort_keys=True, ensure_ascii=False))
                # resultatum.append("----")
                # resultatum.append(pprint.pprint(res))

        return resultatum

    def imprimereTabula(self, linguam: list = None) -> list:
        """imprimere /print/@eng-Latn

        Trivia:
        - cōdex, m, s, (Nominative), https://en.wiktionary.org/wiki/codex#Latin
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - pāginae, f, pl, (Nominative), https://en.wiktionary.org/wiki/pagina

        Returns:
            [list]:
        """
        resultatum = []
        resultatum_corpus = []
        resultatum_corpus_totale = 0
        linguam_clavem = []
        if linguam:
            for item in linguam:
                linguam_clavem.append(
                    item.replace('#item+rem+i_qcc+is_zxxx+', '')
                )
        # resultatum_corpus.append(linguam_clavem)
        # resultatum_corpus.append(len(linguam_clavem))
        for _clavem, lineam in self.dictionaria.items():

            if len(linguam_clavem) > 0:
                if lineam['#item+rem+i_qcc+is_zxxx+ix_hxlix'] not in \
                        linguam_clavem:
                    continue

            # clavem_i18n = lineam['#item+rem+i_qcc+is_zxxx+ix_uid']
            clavem_i18n = lineam['#item+rem+i_qcc+is_zxxx+ix_hxlix']
            definitionem = lineam['#item+rem+definitionem+i_eng+is_latn']
            item_text_i18n = lineam['#item+rem+i_lat+is_latn']
            ix_wdatap = lineam['#item+rem+i_qcc+is_zxxx+ix_wdatap']
            # ix_glottocode = ''
            # ix_iso639p3a3 = lineam['#item+rem+i_qcc+is_zxxx+ix_iso639p3a3']
            ix_iso639p3a3 = ''
            # ix_wikiq = lineam['#item+rem+i_qcc+is_zxxx+ix_wikiq+ix_linguam']
            # ix_wikiq = ''
            ix_wdatap = ''
            if len(ix_wdatap) > 0:
                ix_wdatap = \
                    "https://www.wikidata.org/wiki/Property:{0}[{0}]".format(
                        ix_wdatap)

            if len(ix_iso639p3a3) > 0:
                ix_iso639p3a3 = \
                    "https://iso639-3.sil.org/code/{0}[{0}]".format(
                        ix_iso639p3a3)
            # if len(ix_wikiq):
            #     ix_wikiq = \
            #         "https://www.wikidata.org/wiki/{0}[{0}]".format(
            #             ix_wikiq)
            # resultatum_corpus.append(str(lineam))
            # resultatum_corpus.append(linguam)
            # resultatum_corpus.append(
            resultatum_corpus.append("| {0}".format(clavem_i18n))
            resultatum_corpus.append("| {0}".format(ix_wdatap))
            resultatum_corpus.append("| {0}".format(ix_iso639p3a3))
            # resultatum_corpus.append("| {0}".format(ix_wikiq))
            resultatum_corpus.append("| {0}".format(item_text_i18n))
            resultatum_corpus.append("| {0}".format(definitionem))
            resultatum_corpus.append('')
            resultatum_corpus_totale += 1

        if resultatum_corpus:
            resultatum.append("")

            resultatum.append("=== Interlinguae in cōdex: {0}".format(
                resultatum_corpus_totale))

            # cōdex, m, s, (nominative)
            # tōtālis, m/f, s, (Nominative)
            # linguae, f, s, (Dative)
            resultatum.append(
                "Tōtālis linguae in cōdex: {0}".format(
                    resultatum_corpus_totale))
            resultatum.append("")

            resultatum.append('[%header,cols="~,~,~,~,~"]')
            resultatum.append('|===')
            # https://en.wiktionary.org/wiki/latinus#Latin
            # nōmina, n, pl, (Nominative)
            #     shttps://en.wiktionary.org/wiki/nomen#Latin
            # "nōmen Latīnum"
            # https://en.wiktionary.org/wiki/Latinus#Latin
            # resultatum.append(
            #     "| <span lang='la'>Cōdex<br>linguae</span> | "
            #     "<span lang='la'>Glotto<br>cōdicī</span> | "
            #     "<span lang='la'>ISO<br>639-3</span> | "
            #     "<span lang='la'>Wiki QID<br>cōdicī</span> | "
            #     "<span lang='la'>Nōmen Latīnum</span> |")
            # resultatum.append("| --- | --- | --- | --- | --- |")
            resultatum.append("| Interlinguae")
            resultatum.append("| /Wiki P/")
            resultatum.append("| ISO 639-3")
            # resultatum.append("| Wiki QID cōdicī")
            resultatum.append("| Nōmen Latīnum")
            resultatum.append("| Definitionem")
            resultatum.append('')
            resultatum.extend(resultatum_corpus)
            resultatum.append('|===')
            resultatum.append("")

        return resultatum

    def quod(self, terminum: str,
             #  factum: str = '#item+rem+i_lat+is_latn',
             clavem: str = None) -> str:
        # clavem_defallo = [
        #     '#item+rem+i_qcc+is_zxxx+ix_hxla',
        #     '#item+rem+i_qcc+is_zxxx+ix_csvsffxm'
        # ]
        # raise ValueError(terminum, self.dictionaria_codex.keys())
        if not terminum:
            return None

        clavem_defallo = [
            '#item+rem+i_qcc+is_zxxx+ix_hxlix',
            '#item+rem+i_qcc+is_zxxx+ix_hxlvoc'
        ]
        _clavem = clavem_defallo if clavem is None else [clavem]
        # _clavem = clavem_defallo

        for item in _clavem:
            # print('item', item)
            for _k, linguam in self.dictionaria.items():
                # print('linguam', linguam)
                if terminum == linguam[item]:
                    # if terminum.find(linguam[item]) > -1 and linguam[item]:
                    # return linguam[factum]
                    return linguam

        return None


def descriptio_tabulae_de_lingua(
        lingua_textum: Union[str, list], rem_textum: Union[str, list]) -> list:

    if isinstance(lingua_textum, str):
        lingua_textum = [lingua_textum]

    if isinstance(rem_textum, str):
        rem_textum = [rem_textum]

    paginae = []
    paginae.append('[%header,cols="25h,~a"]')
    paginae.append('|===')
    paginae.append("|")
    paginae.append("Lingua de verba")
    paginae.append("|")
    paginae.append("Verba de conceptiō")
    paginae.append('')

    for item_lingua in lingua_textum:
        item_textum = rem_textum.pop(0)
        paginae.append('|')
        paginae.append(item_lingua)
        paginae.append('|')
        paginae.append(item_textum.strip())
        # paginae.append(item_textum.strip().replace("\n\n",
        #                " +++<br><br>+++").replace("\n", " "))
        paginae.append('')

    paginae.append('|===')

    return paginae


class HXLHashtagSimplici:
    """HXLHashtagSimplici very primitive parse of single HXL hashtag.

    """
    # orīgināle, s, n, nominativus, en.wiktionary.org/wiki/originalis#Latin
    originale: str
    hashtag: str = None
    tag: str = None
    attributes: List[str] = []
    ad_rdf: dict = None

    def __init__(
        self,
        hashtag: str = None
    ):
        """__init__"""

        self.originale = hashtag
        _expand_hxl_ad_rdf()

    def habeo_attributa(
            self,
            quaestio: Union[list, str],
            ignorationes: list = None,
            minimae: int = None) -> bool:
        """habeo_attributa _summary_

        Have these attributes?

        Args:
            quaestio (Union[list, str]): _description_
            minimae (int, optional): How many must have. Defaults to all items.
                                     Use 1 to something equivalent to 'OR'
            ignōrātiōnēs (list, optional): Ignore self attributes if start with
                                           these prefixes

        Returns:
            bool: _description_
        """

        # https://en.wiktionary.org/wiki/habeo#Latin
        # attribūta, pl, nominativus, en.wiktionary.org/wiki/attributus#Latin
        if not quaestio or len(quaestio) == 0:
            return False

        if not isinstance(quaestio, list):
            quaestio = quaestio.lower()
            quaestio = quaestio.split('+')
        quaestio = list(filter(len, quaestio))

        if len(quaestio) == 0:
            return False

        if minimae is None:
            minimae = len(quaestio)
            if ignorationes is not None:
                for _item_ignoration in ignorationes:
                    for _item in self.attributes:
                        if _item.startswith(_item_ignoration):
                            minimae -= 1

        _habeo_attributa_total = 0
        # print(quaestio)
        for _item in quaestio:
            if _item in self.attributes:
                _habeo_attributa_total += 1
            # if _item not in self.attributes:
            #     # print('n foi', quaestio)
            #     return False
        # print('foi')
        # return True
        return (_habeo_attributa_total >= minimae)

    def habeo_attributa__hxl_wdata(
            self,
            quaestio: str) -> dict:
        """habeo_attributa__hxl_wdata Pre-filter HXL_WDATA value (if exists)


        Args:
            quaestio (str): _description_

        Returns:
            dict: _description_
        """
        # @TODO: this function is very simplistic. Does not handle two ix_ attrs
        #        at this moment.
        res = None
        if quaestio == 'hxl_ix':
            for _item in self.attributes:
                if _item.startswith('ix_'):
                    for _wditem in HXL_WDATA:
                        if 'hxl_ix' in _wditem and _wditem['hxl_ix'] == _item:
                            res = _wditem
                            return res
        return None

    def quod_ad_rdf(self, quaestio: str, default: Any = None):
        if self.ad_rdf is None:
            return None

        if quaestio == '' and self.ad_rdf is not None:
            # empty strying is 'debug', but lets ommit verbose __hxlattrs
            ad_rdf = self.ad_rdf
            if '__hxlattrs' in ad_rdf:
                del ad_rdf['__hxlattrs']
            # return self.ad_rdf
            return ad_rdf

        return de_dotted(quaestio, default=default, fontem=self.ad_rdf)

    def quod_attributa(self, praefixa: str) -> list:
        resultatum = []
        for _item in self.attributes:
            if _item.startswith(praefixa):
                resultatum.append(_item)

        return resultatum

    def quod_bcp47(self, caput_contextui: List[str] = None,
                   strictum=True) -> str:

        if self.hashtag in BCP47_EX_HXL:
            # Already '#item+conceptum+codicem'/'#item+conceptum+numerordinatio'
            return BCP47_EX_HXL[self.hashtag]['bcp47']

        if self.hashtag in HXL_HASH_ET_ATTRIBUTA_AD_RDF:
            return HXL_HASH_ET_ATTRIBUTA_AD_RDF[self.hashtag]['__no1bpc47__']

        hxl_base = '#item+rem'
        # numerordinatio = self.quod_numerordinatio(caput_contextui)
        # item_meta = hxl_hashtag_to_bcp47(numerordinatio)
        item_meta = hxl_hashtag_to_bcp47(self.hashtag)

        # print(item_meta)

        if len(item_meta['_error']) == 0 and \
                item_meta['Language-Tag_normalized']:
            return item_meta['Language-Tag_normalized']
            # print('item_meta    ', item_meta)
            # bcp47 = '{0}{1}'.format(
            #     hxl_base,
            #     item_meta['_callbacks']['hxl_attrs']
            # )
            return bcp47
        else:
            # print('item_meta    ', item_meta)
            if strictum:
                raise SyntaxError('{0} <{1}> <{2}>'.format(
                    self.hashtag, caput_contextui, item_meta)
                )
            return 'qcc-Zxxx-r-aDEVNULL-abnop-anop-x-error'

    def quod_numerordinatio(self, caput_contextui: List[str] = None):
        if self.hashtag in BCP47_EX_HXL:
            # Already '#item+conceptum+codicem'/'#item+conceptum+numerordinatio'
            return self.hashtag

        if self.hashtag in HXL_HASH_ET_ATTRIBUTA_AD_RDF:
            return HXL_HASH_ET_ATTRIBUTA_AD_RDF[self.hashtag]['__no1hxl__']

        resultatum = []
        rdf_parts = []
        # if self.ad_rdf is None:
        #     return None
        _ignorationes = ['i_', 'is_']
        _linguae = self.quod_attributa('i_')
        _script = self.quod_attributa('is_')
        est_linguae = len(_linguae) > 0 and 'i_qcc' not in _linguae

        # # Initialize assuming if already is language content, can't be key
        # est_trivium = not est_linguae
        rdftrivio = self.quod_ad_rdf('rdftrivio')
        rdftypisego = self.quod_ad_rdf('rdftypisego')
        rdftypisegolinguis = self.quod_ad_rdf('hxlattrs.rdftypisegolinguis')
        # print('rdftrivio', rdftrivio)
        # print('rdftypisegolinguis', rdftypisegolinguis)

        est_trivium = False
        if not est_linguae:
            if self.habeo_attributa('v_numerordinatio', _ignorationes):
                est_trivium = True
                # TODO: implement check if input already was encoded with
                #       full normalized form.
        else:
            if rdftrivio is not None and rdftypisegolinguis is not None:
                for _item in rdftypisegolinguis:
                    _item2 = _item.lower().replace(':', '_')
                    rdf_parts.append('rdf_p_{0}_s{1}'.format(
                        _item2, rdftrivio
                    ))

        contextus = []
        if caput_contextui is not None and len(caput_contextui) > 0:
            for _res in caput_contextui:
                if _res and len(_res) > 0:
                    _item = _res.split('+')[0]
                    contextus.append(_item)

        # obo:BFO_0000029 -> obo:bfo29
        # _rdf_spatia_nominalibus_prefix_simplici('aaa')
        _rdf_parts = ''

        # if rdftypisego is not None:
        if rdftypisego is not None and est_trivium is True:
            # print('fooooi')
            for _item in rdftypisego:
                _item2 = _rdf_spatia_nominalibus_prefix_simplici(_item)
                rdf_parts.append('rdf_a_{0}'.format(
                    _item2.replace(':', '_')
                ))
            if rdftrivio:
                rdf_parts.append('rdf_s_u2200_s{0}'.format(
                    rdftrivio
                ))

        if len(rdf_parts) > 0:
            rdf_parts.sort()
            _rdf_parts = '+' + '+'.join(rdf_parts)
        if est_linguae:
            return '#item+rem+{0}+{1}{2}'.format(
                _linguae[0], _script[0], _rdf_parts
            )
        else:
            if len(_rdf_parts) == 0:
                return '#meta+rem+i_qcc+is_zxxx+ix_ERROR+{0}'.format(
                    self.hashtag.replace('#', '')
                )
                # return '#meta+ERROR+{0}'.format(
                #     self.hashtag.replace('#', '')
                # )
            return '#item+rem+i_qcc+is_zxxx{0}'.format(
                _rdf_parts
            )

        return '@todo ' + self.hashtag

    def praeparatio(self):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """
        self.hashtag = None
        self.tag = None
        self.attributes = []
        self.ad_rdf = None

        _parts = self.originale.split('+')
        while len(_parts) > 0:
            if self.tag is None:
                _part = _parts.pop(0)
                if not _part.startswith('#'):
                    raise SyntaxError('# ad [{0}]? <{1}>'.format(
                        _part, self.originale))
                self.tag = _part.lower()
                continue

            _part = _parts.pop(0)
            self.attributes.append(_part.lower())

        # self.hashtag = self.tag
        if len(self.attributes) > 0:
            self.hashtag = '{0}+{1}'.format(
                self.tag, '+'.join(self.attributes))
        else:
            self.hashtag = self.tag

        if self.tag in HXL_HASHTAGS_AD_RDF:
            self.ad_rdf = HXL_HASHTAGS_AD_RDF[self.tag]
            if '__hxlattrs' in self.ad_rdf:
                referentia = HXL_ATTRIBUTES_AD_RDF[self.ad_rdf['__hxlattrs']]
                # for _attr, _option in self.ad_rdf['__hxlattrs'].items():
                for _attr, _option in referentia.items():
                    # print(_attr)
                    if self.habeo_attributa(_attr):
                        self.ad_rdf['hxlattrs'] = _option
                        break
                # pass

        return self


def hxl_hashtag_normalizatio(hashtag: str) -> str:
    """hxl_hashtag_normalizatio Canonical form of HXL and HXLTM hashtags

    Quick function to remove extra spaces and repeated attributes. If
    the hashtag starts with HXLTM pattern, we also do extra sorting
    and make sure have at least language and stript.

    Note: this function will not do further syntax checking. The
    bcp47_langtag() / hxl_hashtag_to_bcp47() could be used for this.

    Args:
        hashtag (str): HXL / HXLTM hashtag

    Returns:
        str: canonized hashtag

    >>> hxl_hashtag_normalizatio(' #date+stArT ')
    '#date+start'
    >>> hxl_hashtag_normalizatio(
    ...     '#item+rem+i_qcc+is_zxxx+ix_zza+rdf_t_xsd_datetime+ix_aaa+ix_aaa')
    '#item+rem+i_qcc+is_zxxx+ix_aaa+ix_zza+rdf_t_xsd_datetime'

    >>> hxl_hashtag_normalizatio(
    ...     '#item+rem+i_qcc+is_zxxx+ix_zza+rdf_t_xsd_int+rdf_t_xsd_int+ix_aaz')
    '#item+rem+i_qcc+is_zxxx+ix_aaz+ix_zza+rdf_t_xsd_int'

    >>> hxl_hashtag_normalizatio('#item+rem+i_qcc+is_zxxx+ix_aaz+ix_zza' +
    ... '+rdf_t_xsd_int+rdf_s_u2200_s+rdf_p_skos_preflabel' +
    ... '+rdf_y_u001d_u007c+rdf_y_u001d_u007c+rdf_a_owl_thing')
    '#item+rem+i_qcc+is_zxxx+ix_aaz+ix_zza+rdf_a_owl_thing\
+rdf_p_skos_preflabel+rdf_s_u2200_s+rdf_t_xsd_int+rdf_y_u001d_u007c'

    """

    # @TODO this function could be simplified in half even further if we change
    #       +i_lll (the lang attribute) to have an additional letter
    #       sorting before all others +is_llll and +ix_xxxxx. Yet we may keep
    #       as it is (Rocha, 2022-07-31 14:12 UTC)

    hxltm_prefix = ('#item+rem', '#meta+rem', '#status+rem')
    hxltm_hashtag = None
    _hxltm_attrs = []
    _hxltm_attrs_lang = None
    _hxltm_attrs_script = None
    _hxltm_attrs_ix = []
    _hxltm_attrs_rest = []

    if not hashtag or not hashtag.strip().startswith('#'):
        raise SyntaxError(hashtag)
    hashtag = hashtag.strip().lower()
    if not hashtag.startswith(hxltm_prefix):
        # Assime is a generic HXL hashtag. Return as it is
        # Similar to #item+conceptum+(...), which are simpler and require
        # no special sorting
        return hashtag

    for item in hxltm_prefix:
        if hashtag.startswith(item):
            hxltm_hashtag = item
            _hxltm_attrs = hashtag.replace(item + '+', '').split('+')
            break

    # Remove duplicates
    _hxltm_attrs = list(dict.fromkeys(_hxltm_attrs))

    for item in _hxltm_attrs:
        if item.startswith('i_') and len(item) == 5:
            _hxltm_attrs_lang = item
        elif item.startswith('is_') and len(item) == 7:
            _hxltm_attrs_script = item
        elif item.startswith('ix_') and len(item) >= 5:
            _hxltm_attrs_ix.append(item)
        else:
            _hxltm_attrs_rest.append(item)

    if _hxltm_attrs_lang is None:
        raise SyntaxError(f'+i_qqq? <{hashtag}>')
    if _hxltm_attrs_script is None:
        raise SyntaxError(f'+is_qqqq? <{hashtag}>')

    hxltm_hashtag += '+' + _hxltm_attrs_lang
    hxltm_hashtag += '+' + _hxltm_attrs_script
    if len(_hxltm_attrs_ix) > 0:
        _hxltm_attrs_ix.sort()
        hxltm_hashtag += '+{0}'.format('+'.join(_hxltm_attrs_ix))
    if len(_hxltm_attrs_rest) > 0:
        _hxltm_attrs_rest.sort()
        hxltm_hashtag += '+{0}'.format('+'.join(_hxltm_attrs_rest))

    return hxltm_hashtag


def hxl_hashtag_to_bcp47(
    hashtag: str,
) -> str:
    """hxl_hashtag_to_bcp47 _summary_

    _extended_summary_

    Args:
        hashtag (str): an full HXL hashtag

    Returns:
        str: an BCP47 language tag
    """
    if not hashtag.startswith('#'):
        raise ValueError('{0} not start with #'.format(hashtag))

    hashtag = hxl_hashtag_normalizatio(hashtag)

    # This this AST is similar to bcp47_langtag
    result = {
        # The input Language-Tag, _as it is_
        'Language-Tag': None,
        # The Language-Tag normalized syntax, if no errors
        'Language-Tag_normalized': None,
        'language': None,
        'script': None,
        'region': None,
        'variant': [],
        'extension': {
            # Based on AST of bcp47_langtag_callback_hxl()
            'r': {
                'rdf:Statement_raw': None,
                # 'bcp47_extension_r_normalized': None,
                'rdf:subject': [],
                'rdf:predicate': [],
                'rdf:object': [],
                'rdf:type': [],
                'rdfs:Datatype': None,
                'xsl:transform': [],
                '_unknown': [],
                '_error': [],
                # 'csvw:separator': '', # Added only if necessary
                # 'prefix': [],  # Added only if necessary
            }
        },
        'privateuse': [],  # Example: ['wadegile', 'private1']
        'grandfathered': None,
        '_callbacks': {
            'hxl_attrs': None,
            'hxl_minimal': None,
            'hxl_original': hashtag,
            'rdf_parts': None
        },
        '_unknown': [],
        '_error': [],
    }

    _bcp47_g_parts = []

    # raise ValueError(hashtag)

    # _know_hardcoded = {**BCP47_EX_HXL, **BCP47_EX_HXL_EXTRAS}
    _know_hardcoded = BCP47_EX_HXL
    if hashtag in _know_hardcoded:
        _hashtag = _know_hardcoded[hashtag]['hxl']
        result['_callbacks']['hxl_original'] = hashtag
        # raise ValueError(hashtag, _hashtag)
        hashtag = _hashtag

    if hashtag in HXL_HASH_ET_ATTRIBUTA_AD_RDF:
        _hashtag = _know_hardcoded[hashtag]['hxl']
        result['_callbacks']['hxl_original'] = hashtag
        hashtag = HXL_HASH_ET_ATTRIBUTA_AD_RDF[hashtag]['__no1hxl__']

    parts = hashtag.split('+')
    # bash_hashtag = parts.pop(0)
    # print('parts', parts)
    privateuse = []
    rdf_parts = []
    for item in parts:
        if item.startswith('i_'):
            result['language'] = item.lower().replace('i_', '')

        if item.startswith('is_'):
            result['script'] = item.replace('is_', '').capitalize()

        # Did we really use ir_ as prefix for region? Is necessary at all?
        # Anyway adding here, just in case may relevant later
        # (Rocha, 2022-05-31)
        if item.startswith('ir_'):
            region = item.replace('ir_', '')
            if region.isdigit():
                result['region'] = region
            else:
                result['region'] = region.upper()

        if item.startswith('ix_'):
            privateuse.append(item.lower().replace('ix_', ''))

        if item.startswith('rdf_'):
            rdf_parts.append(item.replace('rdf_', ''))

    if len(privateuse) > 0:
        privateuse.sort()
        result['privateuse'] = privateuse

    if len(rdf_parts) > 0:
        # rdf_parts = sorted(rdf_parts)
        rdf_parts.sort()
        result['_callbacks']['rdf_parts'] = rdf_parts
        # value_prefixes = None
        for item in rdf_parts:
            if item.startswith('s_'):
                # _subject= item.replace('s_', '').replace('_', ':')
                _subject_code, _subjec_value = item.replace(
                    's_', '').split('_')
                _subject_code = _subject_code.upper()
                _subjec_value = _subjec_value.replace('s', '')

                _subjec_value = _subjec_value.replace('s', '')
                result['extension']['r']['rdf:subject'].append(
                    '{0}||{1}:NOP'.format(
                        _subject_code, _subjec_value
                    ))

                _bcp47_g_parts.append('s{0}-s{1}-snop'.format(
                    _subject_code, _subjec_value
                ))

            elif item.startswith('a_'):
                prefix = item.replace('a_', '').replace('_', ':')
                prefix_normali = _rdf_spatia_nominalibus_prefix_simplici(
                    prefix)
                # raise ValueError(item, prefix_normali)
                type_prefix, type_item = prefix_normali.split(':')
                type_meta = prefix_normali + '||0:NOP'
                result['extension']['r']['rdf:type'].append(type_meta)
                _bcp47_g_parts.append('a{0}-a{1}-anop'.format(
                    type_prefix.upper(), type_item.lower()
                ))

            elif item.startswith('p_'):
                # print('item', item)
                # _item_parts = item.replace('p_', '').replace('_', ':')
                # _item_parts = item.lstrip('p_').replace('_', ':')
                _item_parts = item[2:].replace('_', ':')
                _item_parts = _item_parts + ':NOP'
                result['extension']['r']['rdf:predicate'].append(_item_parts)
                _index_p = result['extension']['r']['rdf:predicate'].index(
                    _item_parts
                )

                # _subject_nop = 'NOP' reserved for potential future use
                # print('_item_parts', _item_parts)
                _predicate_ns, _predicate_item, _subject, _subject_nop = \
                    _item_parts.split(':')

                _subject = ''.join(filter(str.isdigit, _subject))

                raw_predicate = f'{_predicate_ns}:{_predicate_item}'
                normalized_predicate = None

                if raw_predicate in RDF_SPATIA_NOMINALIBUS_PREFIX:
                    normalized_predicate = RDF_SPATIA_NOMINALIBUS_PREFIX[raw_predicate]
                elif raw_predicate in RDF_SPATIA_NOMINALIBUS_PREFIX_EXTRAS:
                    normalized_predicate = \
                        RDF_SPATIA_NOMINALIBUS_PREFIX_EXTRAS[raw_predicate]

                if normalized_predicate is not None:
                    if normalized_predicate.startswith('obo:'):
                        normalized_predicate = normalized_predicate.lower()
                        _predicate_ns = 'obo'
                        _predicate_item_raw = \
                            normalized_predicate.replace('obo:', '')
                        _predicate_item_raw_digits = ''.join(
                            filter(str.isdigit, _predicate_item_raw))
                        _predicate_item_raw_alpha = \
                            _predicate_item_raw.replace(
                                _predicate_item_raw_digits, '').replace(
                                    '_', '')
                        _predicate_item = '{0}{1}'.format(
                            _predicate_item_raw_alpha,
                            _predicate_item_raw_digits.lstrip('0')
                        )
                    else:
                        _predicate_ns, _predicate_item = \
                            normalized_predicate.split(':')
                        # pass

                    result['extension']['r']['rdf:predicate'][_index_p] = \
                        '{0}||{1}:{2}'.format(
                        RDF_SPATIA_NOMINALIBUS_PREFIX[raw_predicate],
                        _subject, _subject_nop)
                else:
                    result['extension']['r']['rdf:predicate'][_index_p] = \
                        '{0}||{1}:{2}'.format(
                        raw_predicate,
                        _subject, _subject_nop)

                _bcp47_g_parts.append('p{0}-p{1}-ps{2}'.format(
                    _predicate_ns.upper(), _predicate_item, _subject,
                ))
                # _bcp47_g_parts.append('p{0}-p{1}-p{2}'.format(
                #     _predicate_ns.upper(), _predicate_item,
                #     _subject, _subject_nop.lower()
                # ))

            elif item.startswith('y_'):
                # _cell_transformer = item.replace('y_', '').lower()
                _cell_transformer = item[2:]
                tverb, tval_1 = _cell_transformer.split('_')
                # if _tkey == 'csvwseparator':
                # print('oi', _tkey, _tvalue)
                if tverb.lower() == EXTRA_OPERATORS['GS']['hxl']:
                    # _cell_separator = CSVW_SEPARATORS[_tvalue]
                    decoded_separator = None
                    if tval_1 in CSVW_SEPARATORS:
                        decoded_separator = tval_1
                        # encoded_separator = CSVW_SEPARATORS[_tvalue]

                    if decoded_separator is None:
                        raise NotImplementedError(
                            '[{0}] [{1}] not implemented in <{2}>'.format(
                                tval_1, hashtag, CSVW_SEPARATORS
                            ))

                    result['extension']['r']['xsl:transform'].append(
                        '{0}||{1}:NOP'.format(tverb.upper(), tval_1.lower()))

                    _bcp47_g_parts.append('y{0}-y{1}-ynop'.format(
                        EXTRA_OPERATORS['GS']['hxl'].upper(), decoded_separator
                    ))
                elif tverb == EXTRA_OPERATORS['STX']['hxl']:
                    result['extension']['r']['xsl:transform'].append(
                        '{0}||{1}:NOP'.format(tverb.upper(), tval_1.lower()))

                    _bcp47_g_parts.append('y{0}-y{1}-ynop'.format(
                        tverb.upper(), tval_1.lower()
                    ))
                else:
                    result['_unknown'].append('rdf_parts [{0}]'.format(item))

                # result['extension']['r']['xsl:transform'].append(_item_parts)
                # _index_p = result['extension']['r']['rdf:predicate'].index(
                #     _item_parts
                # )

            elif item.startswith('t_'):
                # _cell_transformer = item.replace('y_', '').lower()
                # print(item)
                _cell_transformer = item[2:]
                # print(_cell_transformer)
                tverb, tval_1 = _cell_transformer.split('_')
                # print(tverb, tval_1)
                # raise ValueError(item)
                _bcp47_g_parts.append('t{0}-t{1}-tnop'.format(
                    tverb.upper(), tval_1.lower()
                ))
                _rdf_prefix = '{0}:{1}'.format(tverb, tval_1)
                if _rdf_prefix in RDF_SPATIA_NOMINALIBUS_PREFIX:
                    _rdf_prefix = RDF_SPATIA_NOMINALIBUS_PREFIX[_rdf_prefix]

                if result['extension']['r']['rdfs:Datatype'] is not None:
                    result['_error'].append(
                        '> 1 explicity rdfs:Datatype. Old <{0}>'.format(
                            str(result['extension']['r']['rdfs:Datatype'])
                        ))

                result['extension']['r']['rdfs:Datatype'] = \
                    '{0}||NOP:NOP'.format(_rdf_prefix)

            elif item.startswith('o_'):
                _object = item.replace('o_', '').replace('_', ':')
                result['extension']['r']['rdf:object'].append(_object)
                raise NotImplementedError(
                    'o [{0}] <{1}>'.format(item, hashtag))
            else:
                result['_unknown'].append('rdf_parts [{0}]'.format(item))
            # pass
        if len(_bcp47_g_parts) > 0:
            _bcp47_g_parts.sort()
            result['extension']['r']['rdf:Statement_raw'] = \
                'r-' + '-'.join(_bcp47_g_parts)
            # norm.append('r-' + '-'.join(_bcp47_g_parts))

    # resultatum = '-'.join(resultatum)

    # based on bcp47_langtag() without grandfathered and -r- implemented
    if len(result['_error']) == 0:
        norm = []
        if result['language']:
            norm.append(result['language'])
        if result['script']:
            norm.append(result['script'])
        if result['region']:
            norm.append(result['region'])
        if len(result['variant']) > 0:
            norm.append('-'.join(result['variant']))

        # if len(result['extension']) > 0:
        #     sorted_extension = {}
        #     for key in sorted(result['extension']):
        #         sorted_extension[key] = result['extension'][key]
        #     result['extension'] = sorted_extension

        #     # for key in result['extension']:
        #     #     # if result['extension'][key][0] is None:
        #     #     #     norm.append(key)
        #     #     # else:
        #     #     #     norm.append(key)
        #     #     #     # norm.extend(result['extension'][key])
        #     #     #     norm.append(result['extension'][key])
        #     #     norm.append(key)
        #     #     # norm.extend(result['extension'][key])
        #     #     norm.append(result['extension'][key])

        if result['extension']['r']['rdf:Statement_raw']:
            # norm.append('r-' + '-'.join(_bcp47_g_parts))
            norm.append(result['extension']['r']['rdf:Statement_raw'])

        if len(result['privateuse']) > 0:
            norm.append('x-' + '-'.join(result['privateuse']))

        result['Language-Tag_normalized'] = '-'.join(norm)

    _hxl_original = result['_callbacks']['hxl_original'].lower()

    _hxl_attrs = _hxl_original.replace('#item+rem', '').replace(
        '#item+status', '').replace('#item+conceptum', '').replace('#item', '')

    result['_callbacks']['hxl_attrs'] = _hxl_attrs
    # result['_callbacks']['hxl_minimal'] = bcp47_langtag_callback_hxl(
    #     _hxl_attrs, False)
    result['_callbacks']['hxl_minimal'] = bcp47_langtag_callback_hxl_minimal(
        result, False)

    # print(result['_callbacks']['hxl_minimal'])

    # print(result['Language-Tag_normalized'] , result['_error'])
    # print('[[{0}]]'.format(result['Language-Tag_normalized']))

    return result


def hxltm_data_referentibus(data_referentibus_index: str, columna: str):
    _caput_columna = [columna]
    _data_columna = []

    return _caput_columna, _data_columna


def hxltm__est_data_referentibus(*options):
    dr_regex = r'(DATA_REFERENTIBUS)\(\s*(?P<data_referentibus>.*)\s*;\s*(?P<c3>.*)\s*\)'
    resultatum = set()
    for options_l1 in options:
        # print(options_l1)
        if not options_l1:
            continue
        if not isinstance(options_l1, list):
            options_l1 = [options_l1]
        for options_l2 in options_l1:
            # print('options_l2', options_l2)
            _regex_result = re.search(dr_regex, options_l2)
            # print('options_l2 _regex_result', _regex_result)
            if _regex_result:
                resultatum.add(_regex_result.group('data_referentibus'))

    return resultatum


def hxltm__ex_dict(
        hxltm_dict: dict, caput: list = None, caput_aliis: dict = None):
    """hxltm__ex_dict boostrap an HXLTM caput/data from a Python dict

    Args:
        hxltm_dict (dict): _description_
        caput (list, optional): _description_. Defaults to None.
        caput_aliis (dict, optional): _description_. Defaults to None.

    Returns:
        (Tuple): caput_novo, data_novis
    """

    caput_novo = []
    data_novis = []
    if caput is None:
        _first = list(hxltm_dict.keys())[0]
        caput = hxltm_dict[_first].keys()

    # print(caput)

    if caput_aliis is not None:
        _hxltm_dict = {}
        for clavem, v in hxltm_dict.items():
            _hxltm_dict[clavem] = v
            for k_old, k_new in caput_aliis.items():
                if k_old in _hxltm_dict[clavem]:
                    _hxltm_dict[clavem][k_new] = _hxltm_dict[clavem].pop(k_old)
    else:
        _hxltm_dict = hxltm_dict

    for _index, res in _hxltm_dict.items():
        linea_novae = []
        for item in caput:
            if item in res:
                linea_novae.append(res[item])
            else:
                linea_novae.append('')
        data_novis.append(linea_novae)

    caput_novo = caput
    return caput_novo, data_novis


def hxltm__ixattr_ex_urn(res: str, urn_basi: str, gs_encoder: str = 'x29') -> str:
    """hxltm__ixattr_ex_urn Create normalized +ix_urn attribute

    For some special cases which ammount of data to strictly label with RDF
    is way to much, this function can be used to create an +ix_ attribute
    which will use only [a-z0-9] range in a way that is reversible.
    However, this strategy of encoding will generate -x- attributes which
    will be > 8, as per BCP47 definition of allowed values to -x-.

    Trivia, the gs_encoder default is a hint to ASCII "29 GS (Group separator)"

    Args:
        res (str): the argument
        urn_basi (str): the base for URN.
        gs_encoder (str, optional): character to replace non [a-z0-9].
                                    Defaults to 'x29'.

    >>> hxltm__ixattr_ex_urn('SH.STA.WASH.P5', 'worldbank')
    'ix_urnx29worldbankx29shx29stax29washx29p5'

    >>> hxltm__ixattr_ex_urn('001', 'unm49')
    'ix_urnx29unm49x29001'

    """
    # - ASCII: 29 GS (Group separator)
    #   - x29 = replace non a-z0-9 with this
    res_normali = res.strip().lower()

    if res_normali.find(gs_encoder) > - 1:
        raise NotImplementedError(f'[{res}] have gs_encoder [{gs_encoder}]')

    res_encoded = re.sub('([^a-z0-9z])', gs_encoder, res_normali)
    return f'ix_urn{gs_encoder}{urn_basi}{gs_encoder}{res_encoded}'


def hxltm__quod_data_referentibus(index_nomini: str):
    _path = '{0}/999999/0/{1}.index.json'.format(
        NUMERORDINATIO_BASIM,
        index_nomini,
    )
    if not exists(_path):
        raise FileNotFoundError(
            '{0} not ready. Please use {1}. [{2}]'.format(
                index_nomini,
                "--methodus='index_praeparationi'",
                _path))
    with open(_path, "r") as archivum:
        data = json.load(archivum)
        return data


def hxltm__data_referentibus(
    # al1: str, b2: str,
    significatus: dict,
    caput: list = None, linea: list = None, data_referentibus: dict = None
):
    # print(significatus)
    # print('TODO hxltm__data_referentibus')
    if not significatus['data_referentibus'] in data_referentibus:
        raise FileNotFoundError(
            '{0} not ready. Please use {1}. [{2}]'.format(
                significatus['data_referentibus'],
                "--methodus='index_praeparationi'"))

    # print(_caput)
    # print(significatus)
    if 'b2h_indici' not in significatus or significatus['b2h_indici'] is None:
        significatus['b2h_indici'] = caput.index(significatus['b2h'])

    res = linea[significatus['b2h_indici']]
    if res in data_referentibus[significatus['data_referentibus']]:
        return data_referentibus[significatus['data_referentibus']][res]
    else:
        # @TODO maybe allow raise error instead of return empty
        return ''


def hxltm__data_sort(fonti: str, sortkeys: list = None) -> list:
    """hxltm__data_sort sort contents of HXLTM on disk

    Load all data into memory. Return full List[List]. Header first row

    Args:
        fonti (str): path to file
        sortkeys (list, optional): List of keys to sort. Defaults to None.

    Returns:
        list: sorted data
    """

    if not sortkeys:
        sortkeys = []

    if '#item+conceptum+codicem' not in sortkeys:
        sortkeys.insert(0, '#item+conceptum+codicem')

    _data = []
    caput = []
    with open(fonti, 'r') as _fons:
        _csv_reader = csv.reader(_fons)
        # started = False
        for linea in _csv_reader:
            # print(linea)
            if len(caput) == 0:
                caput = linea
                continue
            _data.append(linea)

    _i0 = caput.index(sortkeys[0])
    if len(sortkeys) == 1:
        _data = sorted(_data, key=lambda row: int(row[_i0]))
    elif len(sortkeys) == 2:
        _i1 = caput.index(sortkeys[1])
        _data = sorted(_data, key=lambda row: (int(row[_i0]), row[_i1]))
    elif len(sortkeys) == 3:
        _i1 = caput.index(sortkeys[1])
        _i2 = caput.index(sortkeys[2])
        _data = sorted(
            _data, key=lambda row: (int(row[_i0]), row[_i1], row[_i2]))
    else:
        raise NotImplementedError('len > 3; [{}] <{}>'.format(
            len(sortkeys), sortkeys))

    resultatum = []
    resultatum.append(caput)
    resultatum.extend(_data)

    return resultatum


def hxltm__data_pivot_wide(caput: list, data: list, is_hotfix_need: bool = False
                           ) -> Tuple[list, List[list]]:
    """hxltm__data_pivot_wide

    For an HXL dataset with strict HXLTM documented attributes, convert data
    from narrow (or long) to wide.
    It works by getting attributes from an column +ix_xyadhxltrivio ("ad")
    and using to pivo from its value all other columns with
    +ix_xyexhxltrivio ("ex").

    Args:
        caput (list): header
        data (list): all data, long format
        is_hotfix_need (bool): edge case for de_hxltm_ad_hxltm_wide

    Returns:
        Tuple[list, List[list]]: caput, data
    """

    referens_columnae = '#item+rem+i_qcc+is_zxxx+ix_xyadhxltrivio'

    # No + as prefix here
    hxlatt_ex = 'ix_xyexhxltrivio'
    if referens_columnae not in caput:
        raise SyntaxError('[{}] must exist on header <{}>'.format(
            referens_columnae, caput))

    referens_per_indici = caput.index(referens_columnae)
    referens_ad_indici = []
    for item in caput:
        if hxlatt_ex in item:
            referens_ad_indici.append(caput.index(item))
    if len(referens_ad_indici) == 0:
        raise SyntaxError('{} not in header <{}>'.format(hxlatt_ex, caput))

    referens_hxlattrs = []
    for linea in data:
        if linea[referens_per_indici] not in referens_hxlattrs:
            referens_hxlattrs.append(linea[referens_per_indici])

    sorted(referens_hxlattrs)

    # raise ValueError(len(referens_hxlattrs) * len(referens_ad_indici))

    columna_novae__list = []
    columna_novae__mapping = {}

    for item_I in referens_hxlattrs:
        for item_II in referens_ad_indici:

            res = hxl_hashtag_normalizatio(
                caput[item_II].replace(hxlatt_ex, item_I))
            if res in columna_novae__list:
                continue
            columna_novae__list.append(res)
            columna_novae__mapping[item_I] = columna_novae__list.index(res)
            # columna_novae__mapping[item_I] = len(columna_novae__list) - 1
            # pass
        # columna_novae__list

    # raise ValueError(columna_novae__list)

    data_novae__dict = {}
    _codice_indici = caput.index('#item+conceptum+codicem')
    _matrix_size = len(referens_hxlattrs) * len(referens_ad_indici)
    # raise ValueError(_matrix_size, [''] * _matrix_size)

    if is_hotfix_need:
        _matrix_size = _matrix_size + 1

    _do_not_merge = []
    # These already will be on data
    _do_not_merge.extend(referens_ad_indici)
    for linea in data:
        _codicem = linea[_codice_indici]
        _referens = linea[referens_per_indici]
        __start = columna_novae__mapping[_referens]
        __diff = len(columna_novae__list)
        if _codicem not in data_novae__dict:
            data_novae__dict[_codicem] = {
                'originalis': linea,
                # 'data_novae': [''] * len(columna_novae__list),
                # 'data_novae': [''] * (len(columna_novae__list) + 1),
                'data_novae': [''] * _matrix_size,
                'data_meta': [],
            }
        else:
            # This will try to compare if we can safely add columns that would
            # be equal anyway
            for item_index, item_value in enumerate(linea):
                if item_index in _do_not_merge:
                    continue
                if item_value != \
                        data_novae__dict[_codicem]['originalis'][item_index]:
                    _do_not_merge.append(item_index)

        __loop = 0
        # __loop = -1
        # print(len(data_novae__dict[_codicem]['data_novae']))
        # raise NotImplementedError(referens_ad_indici)
        for index_originalis in referens_ad_indici:
            # index_novae = __start + index_originalis
            # print(__loop, len(data_novae__dict[_codicem]['data_novae']))
            index_novae = __start + __loop
            if not index_novae < len(data_novae__dict[_codicem]['data_novae']):
                break
            # if index_novae not in data_novae__dict[_codicem]['data_novae']:
            #     print('error', __start, index_novae, len(data_novae__dict[_codicem]['data_novae']))
            #     continue
            # print('antes', index_novae, len(data_novae__dict[_codicem]['data_novae']))
            data_novae__dict[_codicem]['data_novae'][index_novae] = \
                linea[index_originalis]
            # print('depois', index_novae)
            # data_novae__dict[_codicem]['data_novae'][index_originalis] = \
            #     linea[index_originalis]
            __loop += 1

    for codicem in data_novae__dict:
        for _old_index, _old_value in enumerate(data_novae__dict[codicem]['originalis']):
            if _old_index not in _do_not_merge:
                data_novae__dict[codicem]['data_meta'].append(_old_value)

    _gambiarra = []
    for codicem in data_novae__dict:
        data_novae__dict[codicem]['data_finalis'] = []
        data_novae__dict[codicem]['data_finalis'].extend(
            data_novae__dict[codicem]['data_meta']
        )

        if is_hotfix_need:
            # now we remove that last part we added early
            data_novae__dict[codicem]['data_finalis'].extend(
                data_novae__dict[codicem]['data_novae'][1:]
            )
        else:
            data_novae__dict[codicem]['data_finalis'].extend(
                data_novae__dict[codicem]['data_novae']
            )

        _gambiarra.append(list(data_novae__dict[codicem]['data_finalis']))

        # raise ValueError(data_novae__dict[codicem]['data_finalis'])

    # raise ValueError(_gambiarra[0])

    # for item in caput:
    #     for item_II in referens_ad_indici:

    #     pass
    # raise ValueError(_old_index)
    _caput_novo_meta = []
    for _old_index, _old_caput in enumerate(caput):
        if _old_index not in _do_not_merge:
            _caput_novo_meta.append(_old_caput)

    caput_novo = _caput_novo_meta + columna_novae__list
    # data_novo = []
    # for codicem in data_novae__dict:
    # _novus = [*data_novae__dict[codicem]['data_meta'], *data_novae__dict[_codicem]['data_novae']]

    # _novus = data_novae__dict[codicem]['data_meta']
    # _novus.extend(data_novae__dict[_codicem]['data_novae'])
    # data_novo.append(data_novae__dict[codicem]['data_meta'] \
    #     + data_novae__dict[_codicem]['data_novae'])

    # Weid way to merge 2 lists. Whatever. Eventually review
    # _novus = []
    # for item in list(data_novae__dict[codicem]['data_meta']):
    #     _novus.append(item)
    # for item in list(data_novae__dict[_codicem]['data_novae']):
    #     _novus.append(item)

    # # raise ValueError(_novus, data_novae__dict[_codicem]['data_novae'])

    # _novus = list(data_novae__dict[_codicem]['data_novae'])

    # data_novo.append(_novus)
    # raise ValueError(data_novae__dict[_codicem]['data_finalis'])
    # data_novo.append(data_novae__dict[_codicem]['data_finalis'])

    # raise NotImplementedError(data_novo[0], data_novae__dict['4'])
    # raise NotImplementedError(caput_novo, data_novo[0])
    # raise NotImplementedError(data_novo[0], data_novae__dict['4']['data_novae'])
    # raise NotImplementedError(data_novae__dict['4'])
    # raise NotImplementedError(columna_novae__list, columna_novae__mapping)

    _duplicates = []
    for item in caput_novo:
        if caput_novo.count(item) != 1 and item not in _duplicates:
            _duplicates.append(item)
    if len(_duplicates) > 0:
        raise SyntaxError('duplicates <{0}> at header'.format(_duplicates))
    # return caput_novo, data_novo

    # raise ValueError(caput_novo, _gambiarra[0])

    return caput_novo, _gambiarra


def hxltm__concat(
    # al1: str, b2: str,
    significatus: dict,
    caput: list = None, linea: list = None, data_referentibus: dict = None
):
    if significatus['a1'] in caput and significatus['b2'] in caput:
        return (linea[caput.index(significatus['a1'])] +
                linea[caput.index(significatus['b2'])])
    else:
        raise SyntaxError('{0} <{1}>? <{2}>'.format(
            'hxltm__concat', significatus, caput))


def hxltm__concat_literal(
    # al1: str, bl2: str,
    significatus: dict,
    caput: list = None, linea: list = None, data_referentibus: dict = None
):
    return significatus['al1'] + significatus['bl2']


def hxltm__concat_prefix(
    # al1: str, b2: str,
    significatus: dict,
    caput: list = None, linea: list = None, data_referentibus: dict = None
):
    # if significatus['b2'] in caput:
    if significatus['bh2'] in caput:
        bh2_indici = caput.index(significatus['bh2'])
        # print('bb2', caput.index(significatus['bh2']), linea[bh2_indici])
        # if len(linea[caput[significatus['b2']]]) > 0:
        if len(linea[bh2_indici]) > -1:
            return (significatus['al1'] +
                    linea[bh2_indici])
    else:
        raise SyntaxError('{0} <{1}>? <{2}>'.format(
            'hxltm__concat_prefix', significatus, caput))
    return ''


def hxltm__concat_suffix(
    # al1: str, b2: str,
    significatus: dict,
    caput: list = None, linea: list = None, data_referentibus: dict = None
):
    # print(significatus)
    # print(caput)
    if significatus['a1'] in caput:
        if len(linea[caput.index(significatus['a1'])]) > 0:
            return (linea[caput.index(significatus['a1'])] +
                    significatus['bl2'])
    else:
        raise SyntaxError('{0} <{1}>? <{2}>'.format(
            'hxltm__concat_suffix', significatus, caput))
    return ''


# https://docs.python.org/3/library/operator.html
# https://en.wiktionary.org/wiki/opus#Latin
HXLTM_OPERA_2 = {
    '==': lambda a1, b2: a1 == b2,
    '!=': lambda a1, b2: a1 != b2,
    '>': lambda a1, b2: a1 > b2,
    '<': lambda a1, b2: a1 < b2,
    '>=': lambda a1, b2: a1 >= b2,
    '<=': lambda a1, b2: a1 <= b2,
}
# HXLTM_OPERA_2 = {
#     '==': lambda a1, b2, _q=None, _dr=None: a1 == b2,
#     '!=': lambda a1, b2, _q=None, _dr=None: a1 != b2,
#     '>': lambda a1, b2, _q=None, _dr=None: a1 > b2,
#     '<': lambda a1, b2, _q=None, _dr=None: a1 < b2,
#     '>=': lambda a1, b2, _q=None, _dr=None: a1 >= b2,
#     '<=': lambda a1, b2, _q=None, _dr=None: a1 <= b2,
# }

HXLTM_OPERA_2_EX = {

    # both are existing data columns
    r'(DATA_REFERENTIBUS)\(\s*(?P<data_referentibus>.*)\s*;\s*(?P<b2h>.*)\s*\)': hxltm__data_referentibus,

    # Too generic
    # r'(CONCAT)\((?P<a1>.*)\s?;\s?(?P<b2>.*)\)': lambda a1: print('manual'),

    # both are existing data columns
    r'(CONCAT)\((?P<a1>#.*)\s?;\s?(?P<b2>#.*)\)': hxltm__concat,

    # Harcoded prefix
    # r'(CONCAT)\(\s*[\"\'](?P<al1>.*)[\"\']\s*;\s*(?P<b2>#.*)\s*\)': hxltm__concat_prefix,
    r'(CONCAT)\(\s*[\"\'](?P<al1>.*)[\"\']\s*;\s*(?P<bh2>#.*)\s*\)': hxltm__concat_prefix,

    # Hardcoded suffix
    r'(CONCAT)\(\s*(?P<a1>#.*)\s*;\s*[\"\'](?P<bl2>.*)[\"\']\s*\)': hxltm__concat_suffix,

    # Harcoded prefix and suffix
    r'(CONCAT)\(\s*[\"\'](?P<al1>.*)[\"\']\s*;\s*[\"\'](?P<bl2>.*)[\"\']\s*\)': hxltm__concat_literal,
}

# CONCAT(#meta+a1;#item+a2)
# CONCAT(#meta+a1;"teste after")
# CONCAT("pre";#item+a2)
# CONCAT(#meta+a1 ; #item+a2)
# CONCAT(#meta+a1 ; "teste after")
# CONCAT("pre" ; #item+a2)
# CONCAT("pre" ; "post")

HXLTM_OPERA_1 = {
    r'(CAPITALIZE)\((?P<a1>.*)\)': lambda a1: str(a1).capitalize(),
    r'(LOWER)\((?P<a1>.*)\)': lambda a1: str(a1).lower(),
    r'(UPPER)\((?P<a1>.*)\)': lambda a1: str(a1).upper(),
    r'(TITLE)\((?P<a1>.*)\)': lambda a1: str(a1).title(),
    r'(TRIM)\((?P<a1>.*)\)': lambda a1: str(a1).strip(),
    # ex (+ ablative) https://en.wiktionary.org/wiki/ex#Latin
    # referentibus, pl, m/f/n, dativus, https://en.wiktionary.org/wiki/referens
    # data, pl, n, nominativus, https://en.wiktionary.org/wiki/datum#Latin
    # https://regex101.com/r/3J42kO/1
    r'(DATA_REFERENTIBUS)\(\s*(?P<data_referentibus>.*)\s*;\s*(?P<c3>.*)\s*\)': "<complex>",
}

# CONCAT(#meta+a1;#item+a2)
# CONCAT(#meta+a1;"teste after")
# CONCAT("pre";#item+a2)
# #country+code+v_unm49=DATA_REFERENTIBUS(i1603_45_49; #country+code+v_iso3 )
# CONCAT("BR";DATA_REFERENTIBUS(i1603_45_49; #country+code+v_iso3 ))
# #adm2+code+v_pcode=DATA_REFERENTIBUS(i1603_45_49; #country+code+v_iso3 )

# HXLTM_OPERA_X = {
#     r'(?P<a1>.*)=(?P<b2>.*)': lambda a1, b3: print('manual'),
# }


def hxltm__quaestio_significatis_i(quaestio: str, caput: list = None) -> dict:
    """hxltm__quaestio_significatis_i parse a 1-statement query

    Args:
        quaestio (str): query
        caput (list, optional): HXL header. Defaults to None.

    Raises:
        SyntaxError: _description_

    Returns:
        dict: _description_
    """
    # - quaestiō, f, s,  dativus, https://en.wiktionary.org/wiki/significatus
    # - significātīs, m/f/n, s, dativus,
    #     https://en.wiktionary.org/wiki/significatus
    significātus = {
        'opus': None,
        'opus_rebus': [],
        'a1': '',
        'a1_indici': None,
        # 'a1_operi': None,  # HXLTM_OPERA_1 lambda lambda function call
        '_datetime': False,
    }

    # filtrum = None

    for regex_str, _lambda in HXLTM_OPERA_1.items():
        # print(regex_str)
        _regex_result = re.search(regex_str, quaestio)
        if _regex_result:
            # significātus['a1_operi'] = _lambda
            significātus['opus'] = _lambda
            for _nomen, _res in _regex_result.groupdict().items():
                significātus[_nomen] = _res
                significātus['opus_rebus'].append(_nomen)
            # print('done', _regex_result.groupdict())
            # print('done', significātus)
        # _regex_parsed = regex_str.match(quaestio)
        # print(_regex_result, _regex_result.group('a1'), regex_str)

    if significātus['opus'] is None:
        raise SyntaxError('{0}? quaestio [{1}] [{2}] <[{3}]>'.format(
            'hxltm__quaestio_significatis_i',
            quaestio,
            significātus,
            HXLTM_OPERA_1.keys()
        ))

    if significātus['a1'].startswith('#'):
        significātus['_datetime'] = significātus['a1'].startswith('#date')
        if caput.index(significātus['a1']) > -1:
            significātus['a1_indici'] = caput.index(significātus['a1'])
        else:
            raise SyntaxError('{0}? quaestio [{1}] [{2}] <[{3}]>'.format(
                'hxltm__quaestio_significatis_i (index #)',
                quaestio,
                significātus,
                HXLTM_OPERA_1.keys()
            ))

    # raise NotImplementedError(
    #     '@TODO hxltm__quaestio_significatis_i {0}'.format(significātus))

    return significātus


def hxltm__quaestio_significatis_ii(quaestio: str, caput: list = None) -> dict:
    """hxltm__quaestio_significatis_ii parse a 2 statement query

    Args:
        quaestio (str): query
        caput (list, optional): HXL header. Defaults to None.

    Raises:
        SyntaxError: _description_

    Returns:
        dict: _description_
    """
    # - quaestiō, f, s,  dativus, https://en.wiktionary.org/wiki/significatus
    # - significātīs, m/f/n, s, dativus,
    #     https://en.wiktionary.org/wiki/significatus
    significātus = {
        'opus': None,
        'opus_rebus': ['a1', 'b2'],
        'a1': '',
        'a1_indici': None,
        'a1_operi': None,  # HXLTM_OPERA_1 lambda lambda function call
        'b2': '',
        'b2_indici': None,
        'b2_operi': None,  # HXLTM_OPERA_1 lambda lambda function call
        '_datetime': False,
    }

    for item in HXLTM_OPERA_2.keys():
        if quaestio.find(item) > -1:
            significātus['opus'] = item
            significātus['a1'], significātus['b2'] = quaestio.split(item)
            break

    if len(significātus['a1']) == 0 or len(significātus['b2']) == 0:
        raise SyntaxError(
            '<{0}>? a1 [{1}] b2 [{2}] caput <[{3}>]'.format(
                quaestio,
                significātus['a1'],
                significātus['b2'],
                caput
            ))

    if significātus['a1'].startswith('#'):
        significātus['_datetime'] = significātus['a1'].startswith('#date')
        if caput.index(significātus['a1']) > -1:
            significātus['a1_indici'] = caput.index(significātus['a1'])
        else:
            SyntaxError('{0} <{1}> <{2}>'.format(
                significātus['a1'], quaestio, caput))

    if significātus['b2'].startswith('#'):
        significātus['_datetime'] = significātus['b2'].startswith('#date')
        if caput.index(significātus['b2']) > -1:
            significātus['b2_indici'] = caput.index(significātus['b2'])
        else:
            SyntaxError('{0} <{1}> <{2}>'.format(
                significātus['b2'], quaestio, caput))

    return significātus


def hxltm__quaestio_significatis_x(
    quaestio: str, caput: list = None, data_referentibus: dict = None
) -> dict:
    """hxltm__quaestio_significatis_i parse a assigment (like add columns)

    Args:
        quaestio (str): query
        caput (list, optional): HXL header. Defaults to None.

    Raises:
        SyntaxError: _description_

    Returns:
        dict: _description_
    """
    # - quaestiō, f, s,  dativus, https://en.wiktionary.org/wiki/significatus
    # - significātīs, m/f/n, s, dativus,
    #     https://en.wiktionary.org/wiki/significatus
    significātus = {
        'opus': None,
        # 'opus_rebus': [],
        'data_referentibus': None,
        'a1': '',
        'a1_indici': None,
        # 'a1_operi': None,
        'b2': '',
        'b2_indici': None,
        # 'b2_operi': None,
        '_datetime': False,
        '__len_before': len(caput),
        '__len_after': len(caput),
    }

    HXLTM_OPERA_X_REG = r'(?P<a1>.*)=(?P<b2>.*)'

    # HXLTM_OPERA_X = {
    #     r'(?P<a1>.*)=(?P<b2>.*)': lambda a1, b3: print('manual'),
    # }
    # for regex_str, _lambda in HXLTM_OPERA_X.items():
    #     # print(regex_str)
    #     _regex_result = re.search(regex_str, quaestio)
    #     _regex_result = re.search(regex_str, quaestio)
    #     if _regex_result:
    #         # print('foi', _regex_result)
    #         # significātus['a1_operi'] = _lambda
    #         significātus['opus'] = _lambda
    #         for _nomen, _res in _regex_result.groupdict().items():
    #             significātus[_nomen] = _res
    #             # significātus['opus_rebus'].append(_nomen)

    # print('    quaestio', quaestio)
    _reg1 = re.search(HXLTM_OPERA_X_REG, quaestio)
    # print('_reg1', _reg1, _reg1.groupdict())
    # print('', _reg1.group('a1'))
    # print('', _reg1.group('b2'))

    if _reg1:
        significātus['a1'] = _reg1.group('a1')
        significātus['b2'] = _reg1.group('b2')

    if significātus['a1'].startswith('#'):
        significātus['_datetime'] = significātus['a1'].startswith('#date')
        # print(significātus['a1'])
        # print(caput)
        # if caput.index(significātus['a1']) > -1:
        if significātus['a1'] in caput:
            significātus['a1_indici'] = caput.index(significātus['a1'])
        else:
            # If is not replacing an existing table, we add at the end
            significātus['a1_indici'] = len(caput)
            significātus['__len_after'] = len(caput) + 1
            # SyntaxError('{0} <{1}> <{2}>'.format(
            #     significātus['a1'], quaestio, caput))

    _b2_okay = False
    if significātus['b2'].startswith('#'):
        _b2_okay = True
        significātus['_datetime'] = significātus['b2'].startswith('#date')
        if caput.index(significātus['b2']) > -1:
            significātus['b2_indici'] = caput.index(significātus['b2'])
        else:
            SyntaxError('{0} <{1}> <{2}>'.format(
                significātus['b2'], quaestio, caput))
    else:
        for regex_str, _lambda in HXLTM_OPERA_2_EX.items():
            # print(regex_str)
            _regex_result = re.search(regex_str, quaestio)
            if _regex_result:
                _b2_okay = True
                for _nomen, _res in _regex_result.groupdict().items():
                    significātus[_nomen] = _res.strip()
                significātus['opus'] = _lambda
                # import inspect
                # print(inspect.getsource(_lambda))

                # print('_regex_result', _regex_result)
                # print('_regex_result items',_regex_result.groupdict().items())
                # print('significātus', significātus)

                # significātus['opus_rebus'].append(_nomen.strip())
                # if _regex_result.group('b2h'):
                #     if caput.index(significātus['b2h']) > -1:
                #         significātus['b2h_indici'] = caput.index(
                #             significātus['b2h'])
                #     else:
                #         SyntaxError('{0} <{1}> <{2}>'.format(
                #             significātus['b2h'], quaestio, caput))
                # print(caput)
                # significātus['b2h_incici'] = \
                #     caput.index(significātus['b2h'])
                # print('foi', _regex_result)
                # print('foi', _regex_result.groupdict().items())
                # significātus['a1_operi'] = _lambda

    if _b2_okay is False:
        SyntaxError('{0} <{1}>'.format(
            'hxltm__quaestio_significatis_x', [quaestio, caput, significātus]))
    # print(significātus)
    # raise NotImplementedError(significātus)

    return significātus


class HXLTMAdRDFSimplicis:
    """HXLTM ad RDF

    - ad (+ accusativus),https://en.wiktionary.org/wiki/ad#Latin
    - HXLTM, https://hxltm.etica.ai/
    - RDF, ...
    - simplicis, m/f/n, s, Gen., https://en.wiktionary.org/wiki/simplex#Latin

    """
    # fōns, m, s, nominativus, https://en.wiktionary.org/wiki/fons#Latin
    fons_configurationi: dict = {}
    # methodus_ex_tabulae: dict = {}
    objectivum_formato: str = 'application/x-turtle'
    # methodus: str = ''

    caput: list = []
    data: list = []
    praefixo: str = ''

    # _hxltm: '#meta+{caput}'

    #  '#meta+{{caput_clavi_normali}}'
    # _hxltm_hashtag_defallo: str = '#meta+{{caput_clavi_normali}}'
    # _hxl_hashtag_defallo: str = '#meta+{{caput_clavi_normali}}'

    # identitās, f, s, nom., https://en.wiktionary.org/wiki/identitas#Latin
    # ex (+ ablative), https://en.wiktionary.org/wiki/ex#Latin
    # locālī, n, s, dativus, https://en.wiktionary.org/wiki/localis#Latin
    # identitas_locali_ex_hxl_hashtag: str = '#item+conceptum+codicem'
    identitas_locali_index: int = -1
    venandum_insectum: bool = False  # noqa
    _hxltm_linguae_index: list = []
    _hxltm_linguae_info: dict = {}
    _hxltm_meta_index: list = []
    _hxltm_meta_info: dict = {}
    _hxltm_unlabeled_index: list = []
    _hxltm_unlabeled_info: dict = {}
    _hxltm_labeled = [
        '#item+conceptum+numerordinatio',
        '#item+conceptum+codicem',
        '#status+conceptum+codicem',
        '#status+conceptum+definitionem',
        '#item+rem+i_qcc+is_zxxx+ix_codexfacto',
    ]

    def __init__(
        self,
        fons_configurationi: dict,
        objectivum_formato: str,
        caput: list = None,
        data: list = None,
        venandum_insectum: bool = False  # noqa
    ):
        """__init__ _summary_

        Args:
            methodus_ex_tabulae (dict):
        """
        self.fons_configurationi = fons_configurationi
        self.objectivum_formato = objectivum_formato
        self.venandum_insectum = venandum_insectum
        self.caput = caput
        self.data = data
        # self.methodus = methodus

        self.praefixo = numerordinatio_neo_separatum(
            self.fons_configurationi['numerordinatio']['praefixo'], ':')

        self._post_init()

    def _post_init(self):
        # @TODO esse desgambiarrizar esse _post_init

        if 'identitas_locali_ex_hxl_hashtag' in \
                self.fons_configurationi['numerordinatio']:
            _test = self.fons_configurationi['numerordinatio']['identitas_locali_ex_hxl_hashtag']
            # print("{0}".format(_test))
            # print("{0}".format(self.caput))
            for item in _test:
                if item in self.caput:
                    # self.identitas_locali_ex_hxl_hashtag = item
                    self.identitas_locali_index = self.caput.index(item)
                    break
            if self.identitas_locali_index == -1:
                raise ValueError("HXLTMAdRDFSimplicis [{0}] ?? <{1}>".format(
                    _test, self.caput))

        for _index, item in enumerate(self.caput):
            attrs = item.replace('#item+rem', '')
            bcp47_simplici = qhxl_hxlhashtag_2_bcp47(attrs)
            # lingua = bcp47_langtag(bcp47_simplici, [
            #     # 'Language-Tag',
            #     'Language-Tag_normalized',
            #     'language'
            # ], strictum=False)
            # bcp47_langtag

            if item not in self._hxltm_labeled:
                # _index = self.caput.index(item)
                if item.startswith('#meta'):
                    self._hxltm_meta_index.append(_index)
                    self._hxltm_meta_info[_index] = {
                        'hxltm_hashtag': item,
                        'bcp47': bcp47_simplici
                    }
                    continue
                self._hxltm_unlabeled_index.append(_index)
                self._hxltm_unlabeled_info[_index] = {
                    'hxltm_hashtag': item,
                    'bcp47': bcp47_simplici
                }

            # Language tags only
            if not item.startswith('#item+rem'):
                continue
            # attrs = item.replace('#item+rem', '')
            # hxlattslinguae = qhxl_attr_2_bcp47(attrs)
            # lingua = bcp47_langtag(hxlattslinguae, [
            #     # 'Language-Tag',
            #     'Language-Tag_normalized',
            #     'language'
            # ], strictum=False)
            if bcp47_simplici and not bcp47_simplici.startswith(('qcc', 'zxx')):
                self._hxltm_linguae_index.append(_index)
                self._hxltm_linguae_info[_index] = {
                    'hxltm_hashtag': item,
                    'bcp47': bcp47_simplici
                }

    def resultatum_ad_csv(self):
        """resultatum ad csv text/csv

        Returns:
            (int): status code
        """
        _writer = csv.writer(sys.stdout)
        _writer.writerow(self.caput)
        _writer.writerows(self.data)
        return EXIT_OK

    # resultātum, n, s, nominativus, https://en.wiktionary.org/wiki/resultatum
    def resultatum_ad_ntriples(self):
        """resultatum ad n triples application/n-triples

        Returns:
            (int): status code
        """
        print('# TODO HXLTMAdRDFSimplicis.resultatum_ad_ntriples')
        print('# ' + str(self.fons_configurationi))

        return EXIT_OK

    # resultātum, n, s, nominativus, https://en.wiktionary.org/wiki/resultatum
    def resultatum_ad_turtle(self):
        """resultatum ad n turtle application/x-turtle

        Returns:
            (int): status code
        """
        print('# [{0}]'.format(self.praefixo))
        print('@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .')
        print('@prefix skos: <http://www.w3.org/2004/02/skos/core#> .')
        print('')
        if self.venandum_insectum:
            print('# @TODO melhorar HXLTMAdRDFSimplicis.resultatum_ad_turtle')
            print('# fons_configurationi ' + str(self.fons_configurationi))
            print('# _hxltm_unlabeled_info ' + str(self._hxltm_unlabeled_info))
            print('# _hxltm_meta_info ' + str(self._hxltm_meta_info))
            print('# _hxltm_linguae_info ' + str(self._hxltm_linguae_info))
            print('')
            print('# @TODO adicionar mais prefixos de '
                  'https://www.wikidata.org/wiki/EntitySchema:E49')
            print('')
        # print('<urn:1603:63:101> a skos:ConceptScheme ;')
        # print('  skos:prefLabel "1603:63:101"@mul-Zyyy-x-n1603 .')
        print("<urn:{0}> a skos:ConceptScheme ; \n"
              "  skos:prefLabel \"{0}\"@mul-Zyyy-x-n1603 .".format(
                  self.praefixo))
        print('  # @TODO skos:hasTopConcept')
        print('')

        # @TODO: implementar qhx-Latn (HXLStandard), refs
        #        https://github.com/EticaAI/hxltm/blob/main/docs
        #        /codex-simplex-ontologiae/ontologia.yml

        for linea in self.data:
            # print('# {0}'.format(linea))
            # print('# {0}'.format(self.identitas_locali_index))
            # _codex_locali = self.quod(
            #     linea, '#item+rem+i_qcc+is_zxxx+ix_wdatap1585')
            _codex_locali = str(int(linea[self.identitas_locali_index]))
            print('<urn:{0}:{1}> a skos:Concept ;'.format(
                self.praefixo,
                _codex_locali
            ))
            # print('  skos:prefLabel "{0}:{1}"@mul-Zyyy-x-n1603 .'.format(
            print('  skos:prefLabel "{0}:{1}"@mul-Zyyy-x-n1603 ;'.format(
                self.praefixo,
                _codex_locali
            ))
            _skos_related = []
            _skos_related_raw = []
            for _index, item in enumerate(linea):
                if item and _index in self._hxltm_unlabeled_index and \
                        self._hxltm_unlabeled_info[_index]['bcp47']:
                    _skos_related_raw.append('"{0}"@{1}'.format(
                        rdf_literal_escape(item),
                        self._hxltm_unlabeled_info[_index]['bcp47']
                    ))
                    # pass

            # @TODO add other related
            _skos_related = _skos_related_raw
            # linguae = self._quod_linguae(res)
            if len(_skos_related) > 0:
                print("  skos:related\n    {0} .".format(
                    " ,\n    ".join(_skos_related_raw)
                ))
            for _index, item in enumerate(linea):
                # if len(item) and _index in self._hxltm_unlabeled_index:
                #     print('  # {0} [{1}]'.format(
                #         self._hxltm_unlabeled_info[_index]['hxltm_hashtag'], item))
                if self.venandum_insectum:
                    if len(item) and _index in self._hxltm_meta_index:
                        print('  # verbose: {0} [{1}]'.format(
                            self._hxltm_meta_info[_index]['hxltm_hashtag'],
                            item))

            # linguae = self._quod_linguae(res)
            # if len(linguae) > 0:
            #     print("  skos:prefLabel\n    {0} .".format(
            #         " ,\n    ".join(linguae)
            #     ))

            print('')

        return EXIT_OK

    def quod(self, linea: list, caput_item: str) -> str:
        if caput_item in self.caput:
            index = self.caput.index(caput_item)
            # print('## {0}'.format(linea))
            # print('## {0}'.format(linea[index]))
            return linea[index]
        return ''


def hxltm_adde_columna(
    caput: list, data: list, quaestio: str, data_referentibus: dict = None
) -> Tuple[list, list]:
    """hxltm_cum_columna add new column (variables)
    Trivia:
      - cum (+ ablativus), https://en.wiktionary.org/wiki/cum#Latin
      - columnā, s, f, ablativus, https://en.wiktionary.org/wiki/columna#Latin
      - adde, verbum, sing, imper. https://en.wiktionary.org/wiki/addo#Latin
    Args:
        caput (list): _description_
        data (list): _description_
        quaestio (str): _description_
        data_referentibus (dict): Pre-loaded external referential data
    Returns:
        Tuple[list, list]: _description_
    """

    # https://en.wiktionary.org/wiki/columna#Latin
    significātus = hxltm__quaestio_significatis_x(
        quaestio, caput, data_referentibus)

    caput_novo = caput
    caput_novo.append(significātus['a1'])
    data_novis = []

    # caput.append(significātus['a1'])
    for _, linea in enumerate(data):
        linea_novae = []
        linea_novae.extend(linea)
        if significātus['opus']:
            # import inspect
            # print(inspect.getsource(significātus['opus']))
            res = significātus['opus'](
                significātus,
                caput_novo,
                linea_novae,
                data_referentibus
            )
        else:
            # @TODO: potential bug or simpler cases
            if 'b2' in significātus:
                res = significātus['b2']

        linea_novae.append(res)
        data_novis.append(linea_novae)
        # data[index] = data[index].append(significātus['a1'])

    # print('significātus', significātus)
    # raise NotImplementedError('{0} {1}'.format(
    #     'hxltm_cum_columna', quaestio))
    # # print(caput, columnae)
    # for item in columnae:
    #     index_columnae.append(caput.index(item))

    # raise NotImplementedError(data_novis)
    # _caput = columnae
    return caput_novo, data_novis


def hxltm_carricato(
    archivum_trivio: str = None,
    est_stdin: bool = False,
    punctum_separato=","
) -> list:
    """hxltm_carricato load entire raw CSV file to memory.

    Note: this helper is not as efficent as read line by line. But some
    operations already require such task.

    Trivia:
    - carricātō, n, s, dativus, https://en.wiktionary.org/wiki/carricatus#Latin
      - verbum: https://en.wiktionary.org/wiki/carricatus#Latin

    Args:
        archivum_trivio (str, optional): Path to file. Defaults to None.
        est_stdin (bool, optional): Is the file stdin?. Defaults to False.

    Returns:
        list: list of [caput, data], where data is array of lines
    """
    caput = []

    if est_stdin:
        _data = []
        for linea in sys.stdin:
            if len(caput) == 0:
                # caput = linea
                # _reader_caput = csv.reader(linea)
                _gambi = [linea, linea]
                _reader_caput = csv.reader(_gambi, delimiter=punctum_separato)
                caput = next(_reader_caput)
            else:
                _data.append(linea)
        _reader = csv.reader(_data)
        return caput, list(_reader)
    # else:
    #     fons = archivum_trivio
    data = []
    with open(archivum_trivio, 'r') as _fons:
        _csv_reader = csv.reader(_fons, delimiter=punctum_separato)
        for linea in _csv_reader:
            if len(caput) == 0:
                # caput = linea
                # _reader_caput = csv.reader(linea)
                # _gambi = [linea, linea]
                # _reader_caput = csv.reader(_gambi)
                # caput = next(_reader_caput)
                caput = linea
            else:
                data.append(linea)
            # print(row)

    # for line in fons:
    #     print(line)
        # json_fonti_texto += line

    # _reader = csv.reader(_data)
    # return caput, list(_reader)
    return caput, data


def hxltm_carricato_brevibus(
    archivum_trivio: str = None,
    est_stdin: bool = False,
    punctum_separato: str = ",",
    data_lineis: int = 3,
    est_hxl: bool = False
) -> list:
    """hxltm_carricato_brevibus read only header and part of the data

    Note: this helper is not as efficent as read line by line. But some
    operations already require such task.

    Trivia:
    - carricātō, n, s, dativus, https://en.wiktionary.org/wiki/carricatus#Latin
      - verbum: https://en.wiktionary.org/wiki/carricatus#Latin
    - capitī, s, n, https://en.wiktionary.org/wiki/caput#Latin
    - brevibus, pl, m/f/n, https://en.wiktionary.org/wiki/brevis#Latin

    Args:
        archivum_trivio (str, optional): Path to file. Defaults to None.
        est_stdin (bool, optional): Is the file stdin?. Defaults to False.

    Returns:
        list: list of [caput, data], where data is array of lines
    """
    caput = []

    if est_stdin:
        _data = []
        for linea in sys.stdin:
            if len(caput) == 0:
                # caput = linea
                # _reader_caput = csv.reader(linea)
                _gambi = [linea, linea]
                _reader_caput = csv.reader(_gambi, delimiter=punctum_separato)
                caput = next(_reader_caput)
            else:
                if data_lineis <= 0:
                    pass
                else:
                    data_lineis -= 1
                    _data.append(linea)
        _reader = csv.reader(_data)
        return caput, list(_reader)
        # return caput
    # else:
    #     fons = archivum_trivio
    data = []
    with open(archivum_trivio, 'r') as _fons:
        _csv_reader = csv.reader(_fons, delimiter=punctum_separato)
        for linea in _csv_reader:
            if len(caput) == 0:
                caput = linea
            else:
                if data_lineis <= 0:
                    break
                else:
                    data_lineis -= 1
                    data.append(linea)

    return caput, data


def hxltm_cum_ordinibus_ex_columnis(
    caput: list, data: list, quaestio: list, data_referentibus: dict = None
) -> Tuple[list, list]:
    """hxltm_cum_columnis_desideriis with preferred order (not enforced)

    Trivia:
      - cum (+ ablativus), https://en.wiktionary.org/wiki/cum#Latin
      - ōrdinibus, pl, m, ablativus, https://en.wiktionary.org/wiki/ordo#Latin
      - ex (+ ablativus) https://en.wiktionary.org/wiki/ex#Latin
      - columnā, s, f, ablativus, https://en.wiktionary.org/wiki/columna#Latin
      - columnīs, pl, f, ablativus,
      - columnae, pl, f, vocātīvus,
      - dēsīderiīs, pl, n, vocātīvus,

    Args:
        caput (list): _description_
        data (list): _description_
        quaestio (str): _description_
        data_referentibus (dict): Pre-loaded external referential data

    Returns:
        Tuple[list, list]: _description_
    """
    # _ordo_novo = []
    caput_novo = []
    data_novis = []
    quaestio_dict = {}
    for index, item in enumerate(caput):
        quaestio_dict[int(index)] = item

    # print('pre quaestio_dict', quaestio_dict)
    for item in quaestio:
        # print(item)
        if item.find(':') == -1:
            raise SyntaxError('{0} ":" et <{1}>?? <{2}>'.format(
                __name__, item, quaestio
            ))
        _num, _hash = item.split(':')
        if _num.lstrip('+-').isnumeric():
            _num = int(_num)
        else:
            raise SyntaxError('{0} "numero" et <{1}> <{2}>?? <{3}>'.format(
                __name__, _num, item, quaestio
            ))

        quaestio_dict[int(_num)] = _hash
    # print('post quaestio_dict', quaestio_dict)
    # ordo_novo = set()
    ordo_novo = []

    # print(sorted(quaestio_dict.items(), key=lambda item: int(item[0])))
    quaestio_dict_sorted = \
        sorted(quaestio_dict.items(), key=lambda item: int(item[0]))

    for index, item in quaestio_dict_sorted:
        # print('quaestio_dict_sorted item', item)
        if item in caput and caput.index(item) not in ordo_novo:
            # print(item, caput.index(item))
            ordo_novo.append(caput.index(item))

    # print('   ordo_novo', ordo_novo)

    # return [], [[]]
    # for index in enumerate(ordo_novo):
    for index in ordo_novo:
        caput_novo.append(caput[index])

    # print(caput)
    # print(caput_novo)
    # print(list(range(0, len(caput) -1)))
    if caput_novo == caput:
        # print('already equal. return same input')
        return caput, data

    for _, linea in enumerate(data):
        linea_novae = []
        for index in ordo_novo:
            linea_novae.append(linea[index])
        # linea_novae.extend(linea)
        data_novis.append(linea_novae)

    return caput_novo, data_novis


def hxltm_cum_aut_sine_columnis_simplicibus(
    caput: list, data: list, columnae: list,
    _data_referentibus: dict = None, cum_columnis: bool = True
) -> Tuple[list, list]:
    """hxltm_cum_aut_sine_columnis_simplicibus select/exclude only these columns

    This function will not do advanced checks.

    Trivia:
      - cum (+ ablativus), https://en.wiktionary.org/wiki/cum#Latin
      - columnīs, pl, f, ablativus, https://en.wiktionary.org/wiki/columna#Latin
      - columnae, pl, f, vocātīvus,
      - columnae, pl, f, nominativus,
      - simplicibus, pl, m/f/n, dativus https://en.wiktionary.org/wiki/simplex

    Args:
        caput (list): _description_
        data (list): _description_
        quaestio (str): _description_
        data_referentibus (dict): Pre-loaded external referential data

    Returns:
        Tuple[list, list]: _description_
    """

    # https://en.wiktionary.org/wiki/columna#Latin
    index_columnae = []
    _data = []
    caput_novo = []
    data_novis = []
    # print('    caput', caput)
    # print('    columnae', columnae)
    if cum_columnis:
        for item in columnae:
            caput_indici = caput.index(item)
            index_columnae.append(caput_indici)
            caput_novo.append(caput[caput_indici])
    else:
        for item in caput:
            if item not in columnae:
                caput_indici = caput.index(item)
                index_columnae.append(caput_indici)
                caput_novo.append(caput[caput_indici])

    # @TODO: return same data if the caput are equal

    for linea in data:
        _linea = []
        for index in index_columnae:
            _linea.append(linea[index])
        data_novis.append(_linea)

    return caput_novo, data_novis


def hxltm_cum_filtro(
        caput: list, data: list, quaestio: list, data_referentibus: dict = None
) -> Tuple[list, list]:
    """hxltm_cum_filtris Apply filters for existing columns.

    Trivia:
      - cum (+ ablativus), https://en.wiktionary.org/wiki/cum#Latin
      - filtrō, n, s, ablativus, https://en.wiktionary.org/wiki/filtrum#Latin
      - filtrīs, n, pl, ablativus, https://en.wiktionary.org/wiki/filtrum#Latin

    Args:
        caput (list): _description_
        data (list): _description_
        quaestio (list): The query
        data_referentibus (dict): Pre-loaded external referential data

    Returns:
        Tuple[list, list]: _description_
    """
    # https://en.wiktionary.org/wiki/columna#Latin

    significātus = hxltm__quaestio_significatis_i(quaestio, caput)

    # @TODO: maybe implement only appli filters if match a rule
    for _index, linea in enumerate(data):
        data[_index][significātus['a1_indici']] = significātus['opus'](
            data[_index][significātus['a1_indici']]
        )

    return caput, data


# def hxltm_ex_columnis(
def hxltm_sine_columnis(
        caput: list, data: list, columnae: list, data_referentibus: dict = None
) -> Tuple[list, list]:
    """hxltm_sine_columnis cut columns (variables)

    Trivia:
      - ex (+ ablativus), https://en.wiktionary.org/wiki/ex#Latin
      - sine (+ ablativus), https://en.wiktionary.org/wiki/sine#Latin
      - columnīs, f, pl, ablativus, https://en.wiktionary.org/wiki/columna#Latin

    Args:
        caput (list): _description_
        data (list): _description_
        columnae (list): _description_
        data_referentibus (dict): Pre-loaded external referential data

    Returns:
        Tuple[list, list]: _description_
    """
    raise DeprecationWarning('use hxltm_sine_columnis')


# def hxltm_sine_columnis(
#         caput: list, data: list, columnae: list, _data_referentibus: dict = None
# ) -> Tuple[list, list]:
#     """hxltm_sine_selectis _summary_

#     Trivia:
#       - sine (+ ablativus), https://en.wiktionary.org/wiki/sine#Latin
#       - columnīs, f, pl, ablativus, https://en.wiktionary.org/wiki/columna#Latin

#     Args:
#         caput (list): _description_
#         data (list): _description_
#         columnae (list): _description_
#         data_referentibus (dict): Pre-loaded external referential data

#     Returns:
#         Tuple[list, list]: _description_
#     """
#     pass


def hxltm_ex_selectis(
        caput: list, data: list, quaestio: str, data_referentibus: dict = None
) -> Tuple[list, list]:
    """hxltm_ex_selectis select rows (lines of data)

    Args:
        caput (list): _description_
        data (list): _description_
        quaestio (list): The query
        data_referentibus (dict): Pre-loaded external referential data

    Returns:
        Tuple[list, list]: _description_
    """
    # _op_list = ['==', '!=', '<', '>=', '>', '<=']
    _op_list = HXLTM_OPERA_2.keys()
    res_1 = ''
    res_1_isdate = False
    res_2 = ''
    res_2_isdate = False
    op = ''
    _data = []

    significātus = hxltm__quaestio_significatis_ii(quaestio, caput)

    for linea in data:
        a1 = significātus['a1']
        b2 = significātus['b2']

        if significātus['a1_indici'] is not None:
            a1 = linea[significātus['a1_indici']]
        if significātus['b2_indici'] is not None:
            b2 = linea[significātus['b2_indici']]

        if significātus['_datetime'] is True:
            a1 = date.fromisoformat(a1)
            b2 = date.fromisoformat(b2)

        op_signifo = HXLTM_OPERA_2[significātus['opus']](a1, b2)
        if op_signifo == True:
            _data.append(linea)
        # else:
        #     print('DEBUG: skip [<{0}> {1} <{2}>]'.format(
        #         _value_1, op, _value_2))
    return caput, _data


def hxltm_hashtag_ix_ad_rdf(
    hashtag: str,
    proprietates: list = None,
    rdf_trivio: int = 1603,
    limes_minimo: int = 1,
    limes_maximo: int = 1,
    proprietates_ignorato: list = None,
    strictum: bool = False
) -> str:
    """hxltm_hashtag_ix_ad_rdf selectively upgrade some +ix_ to +rdf_p_

    Args:
        hashtag (str): The original hashtag
        proprietates (list, optional): _description_. Defaults to None.
        rdf_trivio (int, optional): _description_. Defaults to 1603.
        limes_minimo (int, optional): Mininum number of upgrades. Defaults to 1.
        limes_maximo (int, optional): Maximum number of ugrades. Defaults to 1.
        proprietates_ignorato (list, optional): Regex list of ix_ hashtags
                                                to ignore. Set to False to
                                                disable defaults.
        strictum (bool, optional): Less tolerant behavior. Defaults to False.


    >>> _basi = '#item+rem+i_qcc+is_zxxx'
    >>> _ix_III = '+ix_iso8601v2020+ix_xywdatap1540+ix_xywdatap2899v65'

    ### Simpler case: if only one option is possible, assume it what user want
    >>> hxltm_hashtag_ix_ad_rdf(_basi + '+ix_xywdatap1540')
    '#item+rem+i_qcc+is_zxxx+rdf_p_wdata_p1540_s1603'

    >>> hxltm_hashtag_ix_ad_rdf(_basi + _ix_III, proprietates=[
    ...    'ix_xywdatap1540'])
    '#item+rem+i_qcc+is_zxxx+ix_iso8601v2020+ix_xywdatap2899v65\
+rdf_p_wdata_p1540_s1603'

    >>> hxltm_hashtag_ix_ad_rdf(_basi + _ix_III, proprietates=[
    ...    r"ix_xywdatap1540"])
    '#item+rem+i_qcc+is_zxxx+ix_iso8601v2020+ix_xywdatap2899v65\
+rdf_p_wdata_p1540_s1603'


    ###>>> hxltm_hashtag_ix_ad_rdf(_basi + _ix_III)

    Returns:
        str: HXL hashtag
    """
    hashtag = hxl_hashtag_normalizatio(hashtag)

    hxltm_prefix = ('#item+rem', '#meta+rem', '#status+rem')

    if proprietates_ignorato is None:
        proprietates_ignorato = [
            # Example: ix_iso8601v2020
            r"^ix_iso.*$",
            # Example: ix_xywdatap2899v65 (but not ix_xywdatap2899)
            r"^ix_xywdata(p|q)[0-9]{1,12}v.*$"
        ]

    ix_attrs = []
    rdf_p_attrs = []
    if not hashtag.startswith(hxltm_prefix):
        # Assime is a generic HXL hashtag. Return as it is
        # Similar to #item+conceptum+(...), which are simpler and require
        # no special sorting
        return hashtag

    for item in hxltm_prefix:
        if hashtag.startswith(item):
            hxltm_hashtag = item
            _hxltm_attrs = hashtag.replace(item + '+', '').split('+')
            for item_II in _hxltm_attrs:
                _ignore = False
                if proprietates_ignorato is not None:
                    for _ig_test in proprietates_ignorato:
                        if re.match(_ig_test, item_II):
                            _ignore = True
                if not _ignore and item_II.startswith('ix_'):
                    ix_attrs.append(item_II)
                if item_II.startswith('rdf_p_'):
                    rdf_p_attrs.append(item_II)
            break

    def _helper(rem: str, strictum: bool = True):
        if rem.startswith('ix_xywdata'):
            # ix_xywdatap
            # ix_xywdataq
            return 'rdf_p_wdata_{0}_s{1}'.format(
                rem.replace('ix_xywdata', ''), str(rdf_trivio)
            )
        if strictum:
            raise NotImplementedError('[{0}] [{1}]'.format(
                rem, hashtag
            ))
        else:
            return False

    if proprietates is None:
        if limes_minimo == 1 and len(rdf_p_attrs) == 1:
            # Already parsed before, let's keep at it is
            pass
        elif limes_minimo == 1 and len(ix_attrs) == 1:
            _old = ix_attrs[0]
            _new = _helper(_old)
            _hxltm_attrs[_hxltm_attrs.index(_old)] = _new
        elif limes_minimo == 0 and len(ix_attrs) == 0:
            # Allowed to ignore
            pass
        else:
            if not hxltm_hashtag.startswith('#meta'):
                # Tolerate malformated #meta's
                raise SyntaxError(
                    '[{0}] limes[{1}] limes_minimo [{2}]'.format(
                        hashtag, len(ix_attrs), limes_minimo))
    else:
        _old_and_new = {}
        _done = False
        for prop_test in proprietates:
            if isinstance(prop_test, str):
                for ix_item in ix_attrs:
                    if ix_item.startswith(prop_test):
                        _new = _helper(ix_item)
                        _old_and_new[ix_item] = _new
                # resultatum.append(_ht_novo)

                if len(_old_and_new.keys()) <= limes_maximo:
                    _done = True
                    break
        if len(_old_and_new.keys()) <= limes_maximo and \
                len(_old_and_new.keys()) >= limes_minimo:
            for _old, _new in _old_and_new.items():
                _hxltm_attrs[_hxltm_attrs.index(_old)] = _new

    # @TODO replace _hxltm_attrs final values
    return hxl_hashtag_normalizatio(
        hxltm_hashtag + '+{0}'.format('+'.join(_hxltm_attrs)))


def hxltm_index_praeparationi(
    caput: list, data: list,
    index_ad_columnam: str = None, strictum: bool = False
) -> dict:
    """hxltm_index_praeparationi add new columns (variables)

    Trivia:
      - index, s, m, nominativus, https://en.wiktionary.org/wiki/index#Latin
      - ad (+ accusativus) https://en.wiktionary.org/wiki/ad#Latin
      - columnam, s, f, acc., https://en.wiktionary.org/wiki/columna#Latin

    Args:
        caput (list): _description_
        data (list): _description_
        index_ad_columnam (str): _description_
        strictum (str): raise errors if same source keys can point to different
                        objective indexes.

    Returns:
        dict: _description_
    """

    # print('caput', caput)

    if strictum:
        raise NotImplementedError('{0} strictum'.format(
            'hxltm_index_praeparationi'))

    if not index_ad_columnam:
        index_ad_columnam = 0
    else:
        de_facto = qhxl_select(
            caput, index_ad_columnam, unicum=True, strictum=True)
        # print('de_facto', de_facto, caput)
        # index_ad_columnam = caput.index(index_ad_columnam)
        index_ad_columnam = caput.index(de_facto)

    _data_json = {}
    data_json = {}

    for linea in data:
        _clavis_ad = linea[index_ad_columnam]
        _clavis_ex = set()
        for res in linea:
            if res:
                _clavis_ex.add(res)
        for item in _clavis_ex:
            _data_json[item] = _clavis_ad

    _claves_n = []
    _claves_l = []
    _clāvēs_nl = []

    for item in _data_json.keys():
        if item.isnumeric():
            _claves_n.append(item)
        else:
            _claves_l.append(item)

    if len(_claves_n) > 0:
        _clāvēs_nl.extend(sorted(_claves_n, key=lambda x: int(x)))

    if len(_claves_l) > 0:
        _claves_l.sort(key=lambda item: (len(item), item))
        _clāvēs_nl.extend(_claves_l)

    for item in _clāvēs_nl:
        data_json[item] = _data_json[item]

    return data_json


# def hxltm_per_columnas(
#         caput: list, data: list, columnae: list) -> Tuple[list, list]:
#     """hxltm_per_columnas Apply filters to existing columns.

#     Trivia:
#       - per (+ accusative), https://en.wiktionary.org/wiki/per#Latin
#         - through, by means of
#           - Qua re per exploratores nuntiata
#       - columnās, f, pl, accusativus, https://en.wiktionary.org/wiki/columna

#     Args:
#         caput (list): _description_
#         data (list): _description_
#         columnae (list): _description_

#     Returns:
#         Tuple[list, list]: _description_
#     """
#     # https://en.wiktionary.org/wiki/columna#Latin
#     index_columnae = []
#     _data = []
#     raise NotImplementedError('@TODO implement add new column features')
#     # print(caput, columnae)
#     for item in columnae:
#         index_columnae.append(caput.index(item))

#     for linea in data:
#         _linea = []
#         for index in index_columnae:
#             _linea.append(linea[index])
#         _data.append(_linea)

#     # _caput = columnae
    return columnae, _data


def qhxl_hxlhashtag_2_bcp47(
        hxlhashtag: str, hxlstd11_compat: bool = False) -> str:
    """qhxl_hxlhashtag_2_bcp47 (no -r- RDF extension support)

    (try) to convert full HXL hashtag to BCP47

    Args:
        hxlhashtag (str):
        hxlstd11_compat (bool):

    Returns:
        str:
    """
    # needs simplification
    if not hxlhashtag:
        return None
    if hxlhashtag.find('i_') == -1:
        return None
    if hxlhashtag.find('is_') == -1 and not hxlstd11_compat:
        return None

    hxlhashtag_parts = hxlhashtag.split('+')
    # langattrs = []
    # print('hxlhashtag_parts', hxlhashtag_parts)
    _bcp_lang = ''
    _bcp_stript = ''
    _bcp_extension = []
    for item in hxlhashtag_parts:
        if item.startswith('i_'):
            _bcp_lang = item.replace('i_', '')
        if item.startswith('is_'):
            _bcp_stript = item.replace('is_', '')
        if item.startswith('ix_'):
            _bcp_extension.append(item.replace('ix_', ''))
        # if not item.startswith(('i_', 'is_', 'ix_')):
        #     continue
        # langattrs.append(item)

    if not _bcp_lang:
        return False
    if not _bcp_stript and hxlstd11_compat:
        return _bcp_lang

    bcp47_simplici = "{0}-{1}".format(
        _bcp_lang.lower(), _bcp_stript.capitalize())
    if len(_bcp_extension) > 0:
        _bcp_extension = sorted(_bcp_extension)
        bcp47_simplici = "{0}-x-{1}".format(
            bcp47_simplici,
            '-'.join(_bcp_extension)
        )

    return bcp47_simplici


def rdf_literal_escape(textum: str) -> str:
    # @see https://stackoverflow.com/questions
    #      /40828501/how-to-encode-rdf-n-triples-string-literals
    # @see https://github.com/RDFLib/rdflib/blob/master/rdflib/term.py
    # @TODO this strategy is obviously too lazy and should be rewriten.

    if not textum:
        return textum
    if not isinstance(textum, str):
        # Likely error, or maybe using int or something
        return textum

    textum = textum.strip()

    # raise ValueError(textum)
    # TODOs [#x22, #x5C, #xA, #xD]
    if textum.find("\r\n") > -1:
        textum = textum.replace("\r\n", ' ')
    if textum.find("\n") > -1:
        # textum = textum.replace(r'\n', r'\\n')
        textum = textum.replace("\n", ' ')
    if textum.find("\r") > -1:
        # textum = textum.replace(r'\r', r'\\r')
        textum = textum.replace("\r", ' ')
    if textum.find(r'"') > -1:
        textum = textum.replace(r'"', r'\"')

    # Since we're using rapper to validate RDFs we will find other errors.
    # Way too lazy for now.

    return textum


def rdf_namespaces_extras(archivum: str) -> dict:
    """rdf_namespaces_extras Populate global RDF_SPATIA_NOMINALIBUS_EXTRAS

    _extended_summary_

    Args:
        archivum (str): path or stdin for archive with extra namespace

    Returns:
        dict: result
    """
    caput, data = hxltm_carricato(archivum)
    index_prefix = caput.index('#x_rdf+prefix')
    index_iri = caput.index('#x_rdf+iri')
    if index_prefix == -1:
        raise SyntaxError('#x_rdf+prefix ? [{0}]'.format(archivum))
    if index_iri == -1:
        raise SyntaxError('#x_rdf+iri ? [{0}]'.format(index_iri))

    global RDF_SPATIA_NOMINALIBUS_EXTRAS
    for linea in data:
        RDF_SPATIA_NOMINALIBUS_EXTRAS[linea[index_prefix]] = linea[index_iri]

    # print(index_prefix, index_iri, RDF_SPATIA_NOMINALIBUS_EXTRAS)
    # pass
    return RDF_SPATIA_NOMINALIBUS_EXTRAS


class SetEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, set):
            return list(obj)
        return json.JSONEncoder.default(self, obj)


def numerordinatio_descendentibus(
        numerordinatio: str, collectio: list, ordo_maximo: int = None) -> list:
    """numerordinatio_descendentibus _summary_

    Trivia:
     - collēctiō, s, f, , nominativus
    - dēscendentibus, pl, m/f/n, , dativus,
        https://en.wiktionary.org/wiki/descendens#Latin
    - ōrdō, s, m, nominativus, https://en.wiktionary.org/wiki/ordo#Latin
    - ōrdinī, s, m, dativus,
    - maximō, s, m/n, dativus, https://en.wiktionary.org/wiki/maximus#Latin

    Args:
        numerordinatio (str): _description_
        collectio (list): _description_
        ordo_maximo (int, optional): _description_. Defaults to None.

    Returns:
        list: _description_
    """

    resultatum = []

    if ordo_maximo is not None:
        ordo_maximo = ordo_maximo + numerordinatio_ordo(numerordinatio)

    for item in collectio:
        if item.startswith(numerordinatio) and item != numerordinatio:
            if ordo_maximo is None:
                resultatum.append(item)
            else:
                item_ordini = numerordinatio_ordo(item)
                if item_ordini <= ordo_maximo:
                    resultatum.append(item)

    return resultatum

# @TODO https://stackoverflow.com/questions/55506715/nested-skos-concept-schemes


def numerordinatio_cum_antecessoribus(
        numerordinatio: Union[str, list],
        rdf_ontologia_ordinibus: list = None,
        praefixum: str = 'mdciii',
        radix: int = 2,
        est_urn: bool = True
) -> str:

    if est_urn is not True:
        # @TODO: implement this
        raise NotImplementedError

    owl_ontology_ranks = []
    if rdf_ontologia_ordinibus is not None and len(rdf_ontologia_ordinibus) > 0:
        owl_ontology_ranks = []
        for item in rdf_ontologia_ordinibus:
            owl_ontology_ranks.append(int(item))

    global NUMERODINATIO_ANTECESSORIBUS__OKAY
    global NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS

    if not isinstance(numerordinatio, list):
        _numerordinatio = numerordinatio_neo_separatum(numerordinatio, ':')
        numerordinatio = _numerordinatio.split(':')

    if len(numerordinatio) == 0 or len(numerordinatio[0].strip()) == 0:
        raise SyntaxError(numerordinatio)

    if ':'.join(numerordinatio) in NUMERODINATIO_ANTECESSORIBUS__OKAY:
        return NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS

    _parts = numerordinatio
    ordo = 0
    trivium = []
    trivium_antecessori = []
    # trivium_descendenti = []
    while len(_parts) > 0:
        ordo = ordo + 1
        _part = _parts.pop(0)
        trivium.append(_part)
        # trivium_descendenti = trivium
        # if len(_part) > 0:
        #     trivium_descendenti.append(_part[0])
        # else:
        #     trivium_descendenti = None

        if ':'.join(trivium) in NUMERODINATIO_ANTECESSORIBUS__OKAY:
            trivium_antecessori.append(_part)
            continue
        else:
            NUMERODINATIO_ANTECESSORIBUS__OKAY.append(':'.join(trivium))

        if len(owl_ontology_ranks) > 0 and ordo in owl_ontology_ranks and \
                ordo >= radix:
            NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS.append([
                '<urn:{0}:{1}(1)>'.format(praefixum, ':'.join(trivium)),
                'a',
                'owl:Ontology',
            ])

        # Exemplum: 1603
        # if ordo == 1:
        if ordo < radix:
            pass
        elif ordo == radix:
            NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS.extend([
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'a',
                    'skos:ConceptScheme',
                ],
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'rdfs:label',
                    '"({0})"'.format(':'.join(trivium)),
                ],
                # [
                #     '<urn:{0}:{1}>'.format(praefixum, ':'.join(trivium)),
                #     'skos:hasTopConcept',
                #     '<urn:{0}:{1}>'.format(
                #         praefixum, ':'.join(trivium_descendenti)),
                # ]
            ])
        elif ordo == (radix + 1):
            NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS.extend([
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'a',
                    # 'skos:ConceptScheme',
                    'skos:Collection',
                ],
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'rdfs:label',
                    '"({0})"'.format(':'.join(trivium)),
                ],
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    # 'skos:member',
                    # '<urn:{0}:{1}()>'.format(
                    #     praefixum, ':'.join(trivium_antecessori)),
                    'skos:inScheme',
                    # 'skos:member',
                    '<urn:{0}:{1}()>'.format(
                        praefixum, ':'.join(trivium_antecessori)),
                    # '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                ]
                # [
                #     # '<urn:{0}:{1}::>'.format(praefixum, ':'.join(trivium)),
                #     # 'skos:member',
                #     '<urn:{0}:{1}()>'.format(
                #         praefixum, ':'.join(trivium_antecessori)),
                #     # 'skos:inScheme',
                #     'skos:member',
                #     # '<urn:{0}:{1}::>'.format(
                #     #     praefixum, ':'.join(trivium_antecessori)),
                #     '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                # ]
            ])
        else:
            NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS.extend([
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'a',
                    # 'skos:ConceptScheme',
                    'skos:Collection',
                ],
                [
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                    'rdfs:label',
                    '"({0})"'.format(':'.join(trivium)),
                ],
                [
                    '<urn:{0}:{1}()>'.format(
                        praefixum, ':'.join(trivium_antecessori)),
                    # 'skos:inScheme',
                    'skos:member',
                    '<urn:{0}:{1}()>'.format(praefixum, ':'.join(trivium)),
                ]
            ])
            # pass

        trivium_antecessori.append(_part)

    return NUMERODINATIO_ANTECESSORIBUS__RDF_TRIPLIS


def numerordinatio_neo_separatum(
        numerordinatio: str, separatum: str = "_") -> str:
    resultatum = ''
    resultatum = numerordinatio.replace('_', separatum)
    resultatum = resultatum.replace('/', separatum)
    resultatum = resultatum.replace(':', separatum)
    # TODO: add more as need
    return resultatum


def numerordinatio_ordo(numerordinatio: str) -> int:
    normale = numerordinatio_neo_separatum(numerordinatio, '_')
    return (normale.count('_') + 1)


def numerordinatio_progenitori(
        numerordinatio: str, separatum: str = "_") -> int:
    # prōgenitōrī, s, m, dativus, https://en.wiktionary.org/wiki/progenitor
    normale = numerordinatio_neo_separatum(numerordinatio, separatum)
    _parts = normale.split(separatum)
    _parts = _parts[:-1]
    if len(_parts) == 0:
        return "0"
    return separatum.join(_parts)


class OntologiaSimplici:
    """Ontologia Simplicī


    Trivia:
    - ontologia, ---, https://en.wiktionary.org/wiki/ontologia#Latin
    - simplicī, s, m/f/b, dativus, https://en.wiktionary.org/wiki/simplex
    - ex (+ ablativus), https://en.wiktionary.org/wiki/ex#Latin
    - rādīcī, s, f, dativus, https://en.wiktionary.org/wiki/radix#Latin
    - archīvō, s, n, dativus, https://en.wiktionary.org/wiki/archivum

    @see https://www.w3.org/2001/sw/BestPractices/OEP/SimplePartWhole
    @see https://www.w3.org/TR/owl2-overview/#Overview

    """

    # No 1603 prefix
    ontologia_radici: str = None

    # dictionaria_radici: This affects how we infer "classes".
    # Without this we may make partsOf as if they're classes,
    # which may be wrong
    dictionaria_radici: str = None

    ordo_radici: int = None
    data_apothecae_ex: str = []
    caput_no1: List[str] = None
    data: List[list] = None

    PRAEFIXUM = [
        '@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .',
        '@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .',
        '@prefix xsd: <http://www.w3.org/2001/XMLSchema#> .',
        '@prefix owl: <http://www.w3.org/2002/07/owl#> .',
        '@prefix skos: <http://www.w3.org/2004/02/skos/core#> .',
        '@prefix p: <http://www.wikidata.org/prop/> .'
    ]

    PARENTES = []

    def __init__(
        self,
        ontologia_radici: str,
        ontologia_ex_archivo: str,
        dictionaria_radici: str = None
    ):

        self.ontologia_radici = numerordinatio_neo_separatum(
            ontologia_radici, ':')
        self.ontologia_ex_archivo = ontologia_ex_archivo
        if dictionaria_radici:
            self.dictionaria_radici = numerordinatio_neo_separatum(
                dictionaria_radici, ':')
        else:
            self.dictionaria_radici = self.ontologia_radici

        self.initiari()

    def initiari(self):
        """initiarī

        Trivia:
        - initiārī, https://en.wiktionary.org/wiki/initio#Latin
        """
        self.caput_no1, self.data = hxltm_carricato(self.ontologia_ex_archivo)

        if self.caput_no1[0] != '#item+conceptum+numerordinatio':
            raise SyntaxError('Non [{0}] index 0 ad [{1}]'.format(
                '#item+conceptum+numerordinatio',
                self.ontologia_ex_archivo
            ))

        if self.caput_no1[1] != '#item+conceptum+codicem':
            raise SyntaxError('Non [{0}] index 1 ad [{1}]'.format(
                '#item+conceptum+codicem',
                self.ontologia_ex_archivo
            ))

        self.ordo_radici = numerordinatio_ordo(self.ontologia_radici)

        _parents__parts = self.dictionaria_radici.split(':')
        _parents__parens = []
        # print('oi', _parents__parts)
        for item in _parents__parts:
            if len(_parents__parens) == 0:
                self.PARENTES.append(
                    '<urn:{0}> rdf:type owl:Ontology .'.format(item))
                self.PARENTES.append(
                    '<urn:{0}> rdf:type owl:Class .'.format(item))
                self.PARENTES.append('')
                _parents__parens.append(item)
                continue

            # if len(_parents__parens) > 0:
            #     # _parents__parens.append(item)
            #     # _aa =
            #     numerordinatio_nunc = ':'.join(_parents__parens)
            # else:
            #     # numerordinatio_nunc = item

            #     self.PARENTES.append(
            #         '<urn:{0}> rdf:type owl:Ontology .'.format(item))
            #     _parents__parens.append(item)
            #     continue
            _parents__parens_old = list(_parents__parens)
            _parents__parens.append(item)
            numerordinatio_nunc = ':'.join(_parents__parens)

            self.PARENTES.append(
                '<urn:{0}> rdf:type owl:Class .'.format(numerordinatio_nunc))
            self.PARENTES.append(
                '<urn:{0}> rdfs:subClassOf <urn:{1}> .'.format(
                    numerordinatio_nunc, ':'.join(_parents__parens_old)))
            # if len(_parents__parens) > 0:
            #     self.PARENTES.append(
            #         '<urn:{0}> rdfs:subClassOf <urn:{1}> .'.format(
            #             numerordinatio_nunc, ':'.join(_parents__parens)))

            self.PARENTES.append('')
        # self.PARENTES = []
        # pass

    def imprimere_ad_tabula(self, punctum_separato: str = ","):
        csv_imprimendo(
            self.caput_no1, self.data, punctum_separato=punctum_separato)
        # return self.resultatum


class OntologiaSimpliciAdOWL(OntologiaSimplici):
    def imprimere_ad_owl(self, punctum_separato: str = ","):
        # - part of (P361)
        #   - https://www.wikidata.org/wiki/Property:P361
        #   - https://www.wikidata.org/wiki/Special:EntityData/P361.ttl
        # - has part or parts (P527)
        #   - https://www.wikidata.org/wiki/Property:P527
        #   - https://www.wikidata.org/wiki/Special:EntityData/P527.ttl
        # - inverse property (P1696)
        #   - https://www.wikidata.org/wiki/Property:P1696
        #   - https://www.wikidata.org/wiki/Property_talk:P1696
        # - https://www.wikidata.org/wiki/EntitySchema:E49
        # ObjectInverseOf
        # owl:inverseOf
        #
        # @TODO parse list of wikidata properties
        # https://www.wikidata.org/wiki/Property:P31

        paginae = []
        paginae.append('# {0}'.format(self.ontologia_radici))
        paginae.extend(self.PRAEFIXUM)
        paginae.append('')
        paginae.append('wdata:P361 rdf:type owl:ObjectProperty .')
        paginae.append('wdata:P1696 rdf:type owl:ObjectProperty .')
        paginae.append('wdata:P361 owl:inverseOf wdata:P1696 .')
        paginae.append('')
        paginae.extend(self.PARENTES)
        paginae.append('')

        ordo_nunc = self.ordo_radici
        parēns = {
            ordo_nunc: self.ontologia_radici
        }
        # print(parēns)
        for linea in self.data:
            numerordinatio_nunc = numerordinatio_neo_separatum(linea[0], ':')
            if numerordinatio_nunc.startswith(
                    self.ontologia_radici + ':0:1603'):
                continue

            ordo_nunc = numerordinatio_ordo(linea[0])
            parēns[ordo_nunc] = numerordinatio_nunc
            numerordinatio_parentī = parēns[ordo_nunc - 1]

            # paginae.append('# {0} {1} {2}'.format(
            #     linea[0], linea[1], ordo_nunc))

            paginae.append('<urn:{0}> wdata:P361 <urn:{1}> .'.format(
                numerordinatio_nunc, numerordinatio_parentī,
                parēns[ordo_nunc]))

            paginae.append(
                '<urn:{0}> rdfs:Literal "{1}" .'.format(
                    numerordinatio_nunc, numerordinatio_nunc,
                    parēns[ordo_nunc]))

            paginae.append('')
            # ordo_nunc = self.ordo_radici
            # parēns = self.ontologia_radici

        ttl_imprimendo(paginae)


class OntologiaVocabulario:
    paginae: list = []
    praefixa: list = [
        '@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .',
        '@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .',
        '@prefix xsd: <http://www.w3.org/2001/XMLSchema#> .',
        '@prefix owl: <http://www.w3.org/2002/07/owl#> .',
    ]

    def __init__(
        self,
    ):
        pass

    def imprimere_ad_ttl(self):
        self.praefixa.append('')
        ttl_imprimendo(self.praefixa)
        ttl_imprimendo(self.paginae)

    @staticmethod
    def turtle_item(subject: str, assertions: list):
        pagiane = []
        if len(assertions) == 1:
            return ['{0} {1} .'.format(subject, assertions[0]), '']
        if len(assertions) > 1:
            pagiane.append('{0} {1} ;'.format(subject, assertions.pop(0)))
            while len(assertions) > 0:
                res = assertions.pop(0)
                if len(assertions) > 0:
                    pagiane.append('  {0} ;'.format(res))
                else:
                    pagiane.append('  {0} .'.format(res))

        pagiane.append('')
        return pagiane


class OntologiaVocabularioHXL(OntologiaVocabulario):
    def praeparatio(self):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """

        self.praefixa.append(
            '@prefix wdata: '
            '<http://www.wikidata.org/wiki/Special:EntityData/> .')

        self.paginae.extend([
            '# WORKING DRAFT / Early proof of concept; likely may be replaced',
            ''
        ])

        self.paginae.extend([
            '##### Universals ######',
        ])
        self.paginae.extend([
            '<urn:hxl:vocab> a owl:Ontology ;',
            '  rdfs:label "HXLStandard/HXLTM ad hoc vocabulary"@en ;',
            '  rdfs:comment "Only applicable for simpler cases such as '
            'use with HXLTM or Numerordinatio"@en .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:h()> a owl:Class ;',
            '  rdfs:label "HXL Hashtag"@en .',
            #'  rdfs:subClassOf <urn:hxl:vocab> .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:a()> a owl:Class ;',
            '  rdfs:label "HXL Attribute"@en .',
            #'  rdfs:subClassOf <urn:hxl:vocab> .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:a:v()> a owl:Class ;',
            '  rdfs:label "Controlled vocabulary attribute"@en ;',
            '  rdfs:subClassOf <urn:hxl:vocab:a()> .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:a:i()> a owl:Class ;',
            '  rdfs:label "Language attribute"@en ;',
            '  rdfs:subClassOf <urn:hxl:vocab:a()> .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:a:is()> a owl:Class ;',
            '  rdfs:label "Writing system attribute"@en ;',
            '  rdfs:subClassOf <urn:hxl:vocab:a()> .',
            ''
        ])
        self.paginae.extend([
            '<urn:hxl:vocab:a:ix()> a owl:Class ;',
            '  rdfs:label "BCP47 -x- private extension attribute"@en ;',
            '  rdfs:subClassOf <urn:hxl:vocab:a()> .',
            ''
        ])

        self.paginae.extend([
            '##### Individuals ######',
        ])

        for item in HXL_WDATA:
            if 'hxl_ix' not in item or not item['hxl_ix']:
                continue
            assertions = ['a <urn:hxl:vocab:a:ix()>']

            if 'wdata_p' in item and item['wdata_p']:
                assertions.append(
                    'owl:sameAs wdata:{0}'.format(item['wdata_p']))

            if 'hxl_v' in item and item['hxl_v']:
                assertions.append(
                    'owl:sameAs <urn:hxl:vocab:a:v:{0}>'.format(
                        item['hxl_v'].replace('v_', '')))

            if 'iri' in item and item['iri']:
                assertions.append(
                    'rdfs:seeAlso <{0}>'.format(item['iri']))

            self.paginae.extend(OntologiaVocabulario.turtle_item(
                '<urn:hxl:vocab:a:ix:{0}>'.format(
                    item['hxl_ix'].replace('ix_', '')), assertions))

        for item in HXL_WDATA:
            if 'hxl_v' not in item or not item['hxl_v']:
                continue
            # We're using hxl_ix as pivot; so ignoring it for now
            if 'hxl_ix' not in item or not item['hxl_ix']:
                continue
            assertions = [
                'a <urn:hxl:vocab:a:v()>',
                'owl:sameAs <urn:hxl:vocab:a:ix:{0}>'.format(
                    item['hxl_ix'].replace('ix_', ''))
            ]

            self.paginae.extend(OntologiaVocabulario.turtle_item(
                '<urn:hxl:vocab:a:v:{0}>'.format(
                    item['hxl_v'].replace('v_', '')), assertions))

        return self


def owl_index() -> str:
    """Just an quick test
    """
    resultatum = """
<?xml version="1.0"?>
<rdf:RDF xmlns="http://www.semanticweb.org/fititnt/ontologies/2022/5/untitled-ontology-123456#"
     xml:base="http://www.semanticweb.org/fititnt/ontologies/2022/5/untitled-ontology-123456"
     xmlns:dc="http://purl.org/dc/elements/1.1/"
     xmlns:owl="http://www.w3.org/2002/07/owl#"
     xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
     xmlns:xml="http://www.w3.org/XML/1998/namespace"
     xmlns:xsd="http://www.w3.org/2001/XMLSchema#"
     xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#"
     xmlns:skos="http://www.w3.org/2004/02/skos/core#">
    <owl:Ontology rdf:about="http://www.semanticweb.org/fititnt/ontologies/2022/5/untitled-ontology-123456">
        <owl:imports rdf:resource="https://github.com/EticaAI/lexicographi-sine-finibus/blob/main/officina/MDCIII.owl"/>
        <owl:imports rdf:resource="file:/workspace/git/EticaAI/lexicographi-sine-finibus/officina/1603/45/16/24/0/1603_45_16_24_0.no1.owl.ttl"/>
        <owl:imports rdf:resource="file:/workspace/git/EticaAI/lexicographi-sine-finibus/officina/1603/45/16/24/1/1603_45_16_24_1.no1.owl.ttl"/>
        <owl:imports rdf:resource="file:/workspace/git/EticaAI/lexicographi-sine-finibus/officina/1603/45/16/24/2/1603_45_16_24_2.no1.owl.ttl"/>
        <owl:imports rdf:resource="file:/workspace/git/EticaAI/lexicographi-sine-finibus/officina/1603/45/16/24/3/1603_45_16_24_3.no1.owl.ttl"/>
    </owl:Ontology>
</rdf:RDF>
    """
    return resultatum


def qhxl(rem: dict, query: Union[str, list]):
    if isinstance(query, str):
        query = [query]
    for clavem, rem_item in rem.items():
        # print(clavem, rem_item, clavem.find(query))
        for query_item in query:
            # if clavem.find(query_item) > -1:
            if clavem.find(query_item) > -1 and query_item.endswith(query_item):
                return rem_item
    return None


def qhxl_attr_2_bcp47(hxlatt: str) -> str:
    """qhxl_attr_2_bcp47

    Convert HXL attribute part to BCP47

    Args:
        hxlatt (str):

    Returns:
        str:
    """
    resultatum = ''
    tempus1 = hxlatt.replace('+i_', '')
    tempus1 = tempus1.split('+is_')
    resultatum = tempus1[0] + '-' + tempus1[1].capitalize()
    # @TODO: test better cases with +ix_
    resultatum = resultatum.replace('+ix_', '-x-')
    # raise ValueError

    return resultatum


def qhxl_bcp47_2_hxlattr(bcp47: str) -> str:
    """qhxl_bcp47_2_hxlattr

    Convert BCP47 to HXL attribute part

    Args:
        hxlatt (str):

    Returns:
        str:

    >>> qhxl_bcp47_2_hxlattr('lat-Latn-x-private1-private2')
    '+i_lat+is_latn+ix_private1+ix_private2'
    >>> qhxl_bcp47_2_hxlattr('qcc-Zxxx-x-wikip')
    '+i_qcc+is_zxxx+ix_wikip'
    """
    resultatum = ''
    bcp47_parsed = bcp47_langtag(bcp47)

    resultatum = '+i_' + bcp47_parsed['language']
    resultatum += '+is_' + bcp47_parsed['script'].lower()
    if len(bcp47_parsed['privateuse']) > 0:
        for item in bcp47_parsed['privateuse']:
            resultatum += '+ix_' + item

    return resultatum


def qhxl_select(caput: list, query: str, unicum: bool = False, strictum: bool = False):
    """qhxl_select

    Select HXL hashtags from list of header of hashtags without
    need be exact

    Args:
        caput (list): HXL hashtags header
        query (str): query to search for parts
        unicum (bool): Return just one item instead of always list
        strictum (bool): Force have at least one value

    Returns:
        list: 0 or more results
    """
    result_filtered = []

    if query in caput:
        # Exact match
        return query if unicum else [query]

    query_parts = query.replace('#', '').strip('+').split('+')
    # query_parts = filter(None, query_parts)
    for res in caput:
        _failed = False
        _res_parts = res.replace('#', '').strip('+').split('+')
        # _res_parts = filter(None, _res_parts)
        for _q in query_parts:
            if _q not in _res_parts:
                _failed = True
                continue
        if not _failed:
            result_filtered.append(res)

    if strictum is True and len(result_filtered) == 0:
        raise ValueError('<{0}> not found in <{1}>'.format(query, caput))

    if strictum is True and unicum is True and len(result_filtered) > 1:
        raise ValueError('Non unicum <{0}>: <{1}>; caput <{2}>'.format(
            query, result_filtered, caput))

    if unicum:
        return result_filtered[0]

    return result_filtered


def res_interlingualibus_formata(rem: dict, query) -> str:
    # pylint: disable=too-many-return-statements

    if not rem[query]:
        return ''

    if query.find('#status+conceptum+definitionem') > -1:
        return "{0} +++<sup><em>(1-100)</em></sup>+++".format(
            rem[query])
    if query.find('#status+conceptum+codicem') > -1:
        return "{0} +++<sup><em>(1-100)</em></sup>+++".format(
            rem[query])

    if query.find('+ix_wikiq') > -1 and query.endswith('+ix_wikiq'):
        return "https://www.wikidata.org/wiki/{0}[{0}]".format(
            rem[query])

    if query.find('+ix_wdatap3916') > -1 and query.endswith('+ix_wdatap3916'):
        # No https?
        return "http://vocabularies.unesco.org/thesaurus/{0}[{0}]".format(
            rem[query])

    if query.find('+ix_wdatap') > -1 and query.endswith('+ix_wdatap'):
        return "https://www.wikidata.org/wiki/Property:{0}[{0}]".format(
            rem[query])

    if query.find('+ix_ta98') > -1 and query.endswith('+ix_ta98'):
        term = rem[query].replace('A', '')
        resultatum = (
            'link:++https://ifaa.unifr.ch/Public/EntryPage/'
            'TA98%20Tree/Entity%20TA98%20EN/{0}%20Entity%20TA98%20EN.htm++['
            '{1}]').format(term, rem[query])
        return resultatum

    return rem[query]


class TabulaAdHXLTM:
    """Tabula ad HXLTM

    - tabula, f, s, nominativus, https://en.wiktionary.org/wiki/tabula
    - ad (+ accusativus),https://en.wiktionary.org/wiki/ad#Latin
    - ex (+ ablativus)
    - HXLTM, https://hxltm.etica.ai/

    """
    methodus_ex_tabulae: dict = {}
    # methodus_ex_tabulae: dict = {}
    objectivum_formato: str = 'hxltm_csv'
    methodus: str = ''

    # _hxltm: '#meta+{caput}'

    #  '#meta+{{caput_clavi_normali}}'
    _hxltm_hashtag_defallo: str = '#meta+{{caput_clavi_normali}}'
    _hxl_hashtag_defallo: str = '#meta+{{caput_clavi_normali}}'

    def __init__(
        self,
        methodus_ex_tabulae: dict,
        objectivum_formato: str,
        methodus: str,
    ):
        """__init__ _summary_

        Args:
            methodus_ex_tabulae (dict):
        """
        self.methodus_ex_tabulae = methodus_ex_tabulae
        self.objectivum_formato = objectivum_formato
        self.methodus = methodus

    def caput_translationi(self, caput: list) -> list:
        """Caput trānslātiōnī

        - trānslātiōnī, f, s, dativus, https://en.wiktionary.org/wiki/translatio
        - caput, n, s, nominativus, https://en.wiktionary.org/wiki/caput#Latin

        Args:
            caput (list): _description_

        Returns:
            list: _description_
        """

        # if self.objectivum_formato.find('hxltm') > -1:
        #     # neo_caput = map(self.clavis_ad_hxl, caput, 'hxltm')
        #     # neo_caput = map(self.clavis_ad_hxl, caput, 'hxltm')
        #     neo_caput = map(self.clavis_ad_hxl, caput)
        #     # neo_caput = map(self.clavis_ad_hxl, caput)
        # if self.objectivum_formato.find('hxl') > -1:
        #     # neo_caput = map(self.clavis_ad_hxl, caput, 'hxl')
        #     neo_caput = map(self.clavis_ad_hxl, caput)
        if self.objectivum_formato.find('hxl') > -1:
            neo_caput = map(self.clavis_ad_hxl, caput)
        else:
            neo_caput = map(self.clavis_normationi, caput)
        return neo_caput

    def clavis_normationi(self, clavis: str) -> str:
        """clāvis nōrmātiōnī

        - clāvis, f, s, normativus, https://en.wiktionary.org/wiki/clavis#Latin
        - nōrmātiōnī, f, s, dativus, https://en.wiktionary.org/wiki/normatio

        Args:
            clavis (str):

        Returns:
            str:
        """
        if not clavis or len(clavis) == 0:
            return ''
        clavis_normali = clavis.strip().lower()\
            .replace(' ', '_').replace('-', '_')

        return clavis_normali

    # def clavis_ad_hxl(self, clavis: str, classis: str = 'hxltm') -> str:
    def clavis_ad_hxl(self, clavis: str) -> str:
        """clavis_ad_hxltm

        - clāvis, f, s, normativus, https://en.wiktionary.org/wiki/clavis#Latin
        - nōrmātiōnī, f, s, dativus, https://en.wiktionary.org/wiki/normatio

        Args:
            clavis (str):

        Returns:
            str:
        """
        clavis_normationi = self.clavis_normationi(clavis)

        if not clavis or len(clavis) == 0:
            return ''
        # print(classis)
        # neo_caput = 'hxl_hashtag'
        # forma = self._hxl_hashtag_defallo
        # if classis == 'hxltm' or not classis:
        if self.objectivum_formato.find('hxltm') > -1:
            neo_caput = 'hxltm_hashtag'
            forma = self._hxltm_hashtag_defallo
        # elif classis == 'hxl':
        elif self.objectivum_formato.find('hxl') > -1:
            neo_caput = 'hxl_hashtag'
            forma = self._hxl_hashtag_defallo

        if clavis_normationi in self.methodus_ex_tabulae['caput'].keys():
            if self.methodus_ex_tabulae['caput'][clavis_normationi] and \
                neo_caput in self.methodus_ex_tabulae['caput'][clavis_normationi] and \
                    self.methodus_ex_tabulae['caput'][clavis_normationi][neo_caput]:
                forma = self.methodus_ex_tabulae['caput'][clavis_normationi][neo_caput]

        hxl_hashtag = forma.replace(
            '{{caput_clavi_normali}}', clavis_normationi)

        return hxl_hashtag

    # def clavis_ad_hxl(self, clavis: str) -> str:
    #     pass


class TabulaSimplici:
    """Tabula simplicī /Simple Table/@eng-Latn

    Trivia:
    - tabula, s, f, nominativus, https://en.wiktionary.org/wiki/tabula#Latin
    - simplicī, s, m/f/n, Dativus, https://en.wiktionary.org/wiki/simplex#Latin
    """

    archivum_trivio: str = ''
    nomen: str = ''
    statum: bool = None
    caput: list = []
    concepta: dict = None
    res_totali: int = 0
    ex_radice: bool = False
    archivum_trivio_ex_radice: str = ''
    archivum_nomini: str = ''
    est_bcp47: bool = False
    urn: str = ''
    # codex_opus: list = []
    # opus: list = []
    # in_limitem: int = 0
    # in_ordinem: str = None
    # quaero_numerordinatio: list = []
    _primaryKey: str = None
    _propertyUrl: str = None
    _numerordinatio_columnae: str = None

    def __init__(
        self,
        archivum_trivio: str,
        nomen: str,
        ex_radice: bool = False
    ):
        self.archivum_trivio = archivum_trivio
        self.nomen = nomen
        self.ex_radice = ex_radice
        self.urn = 'urn:mdciii:{0}'.format(numerordinatio_neo_separatum(
            nomen, ':'
        ))
        # self.initiari()

    def _initiari(self):
        """initiarī

        Trivia:
        - initiārī, https://en.wiktionary.org/wiki/initio#Latin
        """
        if not os.path.exists(self.archivum_trivio):
            # print('self.archivum_trivio', self.archivum_trivio)
            self.statum = False
            return self.statum

        if self.archivum_trivio.find('bcp47') > -1:
            self.est_bcp47 = True

        self.archivum_trivio_ex_radice = \
            self.archivum_trivio.replace(NUMERORDINATIO_BASIM, '')

        self.archivum_nomini = Path(self.archivum_trivio_ex_radice).name

        with open(self.archivum_trivio) as csvfile:
            reader = csv.reader(csvfile)
            for lineam in reader:
                if len(self.caput) == 0:
                    self.caput = lineam
                    for _, _columna in \
                            BCP47_EX_HXL['#item+conceptum+codicem'].items():
                        if _columna in self.caput:
                            self._primaryKey = _columna
                            break
                    for _, _columna in \
                            BCP47_EX_HXL[
                                '#item+conceptum+numerordinatio'].items():
                        if _columna in self.caput:
                            self._numerordinatio_columnae = _columna
                            break
                    continue
                # TODO: what about empty lines?
                self.res_totali += 1

        if self._numerordinatio_columnae is not None:
            self._propertyUrl = '{{{0}}}'.format(
                self._numerordinatio_columnae)
        elif self._primaryKey is not None:
            self._propertyUrl = '{0}:{{{1}}}'.format(
                self.urn, self._primaryKey)

        self.statum = True
        return self.statum

    def _initiari_v2(self):
        """initiarī

        Trivia:
        - initiārī, https://en.wiktionary.org/wiki/initio#Latin
        """
        if not os.path.exists(self.archivum_trivio):
            self.statum = False
            return self.statum

        self.archivum_trivio_ex_radice = \
            self.archivum_trivio.replace(NUMERORDINATIO_BASIM, '')

        self.archivum_nomini = Path(self.archivum_trivio_ex_radice).name

        self.concepta = {}
        with open(self.archivum_trivio) as csvfile:
            reader = csv.DictReader(csvfile)
            for lineam in reader:
                # de_codex = lineam['#item+conceptum+numerordinatio']

                if self.est_bcp47:
                    de_codex = lineam['qcc-Zxxx-r-aMDCIII-alatcodicem-anop']
                else:
                    de_codex = lineam['#item+conceptum+codicem']
                self.concepta[de_codex] = lineam
                # if len(self.caput) == 0:
                #     self.caput = lineam
                #     continue
                # # TODO: what about empty lines?
                # self.res_totali += 1

        self.statum = True
        return self.statum

    def _quod_linguae(self, res: dict) -> list:
        resultatum = []
        # resultatum.append('"todo"@en')
        for clavem, item in res.items():
            if len(item) == 0 or \
                    (not self.est_bcp47 and not clavem.startswith('#item+rem')):
                continue
            if item.find('"') > -1:
                # item = item.replace('"', 'zzzz')
                item = item.replace('"', '\\"')
            attrs = clavem.replace('#item+rem', '')

            if self.est_bcp47:
                hxlattslinguae = clavem
            else:
                hxlattslinguae = qhxl_attr_2_bcp47(attrs)

            # lingua = bcp47_langtag(clavem, [
            lingua = bcp47_langtag(hxlattslinguae, [
                # 'Language-Tag',
                'Language-Tag_normalized',
                'language'
            ], strictum=False)
            if lingua['language'] not in ['qcc', 'zxx']:
                resultatum.append('"{0}"@{1}'.format(
                    rdf_literal_escape(item),
                    lingua['Language-Tag_normalized']))
        return resultatum

    def _quod_descendentia(
            self, dictionaria_codici: list, item_codici: str) -> list:
        # dēscendentia, n, pl, Nominativus,
        #     https://en.wiktionary.org/wiki/descendens#Latin
        resultatum = []

        # de_codex_n = numerordinatio_progenitori(de_codex, ':')

        # for clavem, _item in dictionaria_radici.items():
        for clavem in dictionaria_codici:
            # clavem_n = numerordinatio_progenitori(clavem, ':')
            # print('clavem', de_codex_n, clavem_n)
            progenitor = numerordinatio_progenitori(clavem, ':')
            # print('clavem', item_codici, progenitor)
            if progenitor == item_codici:
                resultatum.append(clavem)
            # pass
        return resultatum

    def praeparatio(self):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """
        self._initiari()
        return self.statum

    def quod_csvw(self) -> dict:
        if self.ex_radice is True:
            _path = self.archivum_trivio_ex_radice
        else:
            _path = self.archivum_nomini

        resultatum = {
            # 'name': self.nomen,
            # 'path': self.nomen,
            'url': _path,
            # 'profile': 'tabular-data-resource',
            'tableSchema': {
                'columns': [],
                'primaryKey': self._primaryKey,
                'propertyUrl': self._propertyUrl
                # 'foreignKeys': []
            },
            # 'stats': {
            #     'fields': len(self.caput),
            #     'rows': self.res_totali,
            # }
        }

        for caput_rei in self.caput:
            item = {
                'name': caput_rei,
                # TODO: actually get rigth type from reference dictionaries
                'datatype': 'string',
            }
            resultatum['tableSchema']['columns'].append(item)

        return resultatum

    def quod_datapackage(self) -> dict:
        if self.ex_radice is True:
            _path = self.archivum_trivio_ex_radice
        else:
            _path = self.archivum_nomini

        resultatum = {
            # 'format': 'csv',
            # 'mediatype': 'text/csv',
            'name': self.nomen,
            # 'path': self.nomen,
            'path': _path,
            'profile': 'tabular-data-resource',
            'schema': {
                'fields': [],
                'primaryKey': self._primaryKey
            },
            'stats': {
                'fields': len(self.caput),
                'rows': self.res_totali,
            }
        }
        _caput_codicem = BCP47_EX_HXL['#item+conceptum+codicem'].values()
        _caput_numerordinatio = \
            BCP47_EX_HXL['#item+conceptum+numerordinatio'].values()

        for caput_rei in self.caput:
            item = {
                'name': caput_rei,
                # TODO: actually get rigth type from reference dictionaries
                'type': 'string',
            }
            if caput_rei in _caput_codicem:
                item['constraints'] = {
                    'unique': True
                }
            if caput_rei in _caput_numerordinatio:
                item['constraints'] = {
                    'unique': True
                }

            resultatum['schema']['fields'].append(item)

        return resultatum

    def quod_rdf_skos_ttl_concepta(self) -> list:
        self._initiari_v2()

        paginae = []

        nomen_radici = numerordinatio_neo_separatum(self.nomen, ':')

        # https://www.w3.org/2015/03/ShExValidata/ (near ok)
        # https://skos-play.sparna.fr/skos-testing-tool (needs more work)
        # paginae.append("<urn:{0}> a skos:Concept ;".format(nomen_radici))
        paginae.append("<urn:{0}> a skos:ConceptScheme ;".format(nomen_radici))
        # paginae.append("  skos:prefLabel\n    {0} .".format(
        #     ",\n    ".join(linguae)
        # ))
        paginae.append("  skos:prefLabel \"{0}\"@{1} .".format(
            rdf_literal_escape(nomen_radici),
            'mul-Zyyy-x-n1603'
        ))
        paginae.append('')

        dictionaria_codici = []

        # for codex_de, res in enumerate(self.concepta):
        for codex_de, _res in self.concepta.items():
            codex_de_n = numerordinatio_neo_separatum(codex_de, ':')
            numerodinatio = nomen_radici + ':' + str(codex_de_n)
            dictionaria_codici.append(numerodinatio)

        for codex_de, res in self.concepta.items():

            codex_de_n = numerordinatio_neo_separatum(codex_de, ':')
            if codex_de_n.startswith('0:1603'):
                continue
            # This is deprecated use; used on 1603_25_1
            if codex_de_n.startswith('0:999'):
                continue
            numerodinatio = nomen_radici + ':' + str(codex_de_n)
            # paginae.append(':{0} a skos:Concept ;'.format(codex_de))
            paginae.append("<urn:{0}> a skos:Concept ;".format(numerodinatio))

            progenitor = numerordinatio_progenitori(numerodinatio, ':')

            if nomen_radici == progenitor:
                paginae.append('  skos:topConceptOf\n    <urn:{0}> ;'.format(
                    progenitor))

            descendentia = self._quod_descendentia(
                dictionaria_codici, numerodinatio)
            if len(descendentia) > 0:
                # print('descendentia', descendentia)
                # paginae.append('  skos:broader\n    {0} ;'.format(
                paginae.append('  skos:narrower\n    {0} ;'.format(
                    ' ,\n    '.join(
                        map(lambda x: '<urn:' + x + '>', descendentia))
                ))

            # paginae.append('  rdfs:subClassOf <urn:{0}> ;'.format(
            #     progenitor
            # ))
            # paginae.append('  skos:narrowerTransitive <urn:{0}> ;'.format(
            # paginae.append('  skos:narrower\n    <urn:{0}> ;'.format(

            # AVOID: tchbc - Top Concepts Having Broader Concepts
            if nomen_radici != progenitor:
                paginae.append('  skos:broader\n    <urn:{0}> ;'.format(
                    progenitor
                ))
            # @TODO: implement inverse, skos:broader

            linguae = self._quod_linguae(res)
            if len(linguae) > 0:
                paginae.append("  skos:prefLabel\n    {0} .".format(
                    " ,\n    ".join(linguae)
                ))
            else:
                # The "." also need to be on last statement
                raise NotImplementedError('{0} / {0} needs be fixed'.format(
                    nomen_radici, numerodinatio))

            # paginae.append('  rdfs:subClassOf <urn:{0}> . '.format(
            # paginae.append('  skos:topConceptOf <urn:{0}> . '.format(
            #     progenitor
            # ))
            # paginae.append('  skos:topConceptOf <http://vocabularies.unesco.org/thesaurus> .')
            # paginae.append('  skos:topConceptOf <http://vocabularies.unesco.org/thesaurus> ;')
            paginae.append('')

        # @TODO: ...

        # @TODO: edge case, such as 1603:25:1
        #   <urn:1603:25:1:4> a skos:Concept ;
        #   skos:prefLabel
        #       "extremitates"@lat-Latn-x-wikip3982 .

        return paginae


def ttl_imprimendo(
        paginae: Iterator[str],
        archivum_trivio: str = None):
    if archivum_trivio:
        raise NotImplementedError('{0}'.format(archivum_trivio))
    # imprimendō, v, s, dativus, https://en.wiktionary.org/wiki/impressus#Latin

    for linea in paginae:
        print(linea)


class XLSXSimplici:
    """Read-only wrapper for XLSX files

    - XLSX http://officeopenxml.com/anatomyofOOXML-xlsx.php
    - simplicī, m/f/n, s, dativus, https://en.wiktionary.org/wiki/simplex#Latin
    """

    archivum_trivio: str = ''
    workbook: None
    active: str = None
    active_col_start: int = 0
    active_col_end: int = 1
    active_col_ignore: list = []  # Example: headers without text
    active_row_start: int = 0
    active_row_end: int = 1
    _formatum: str = 'csv'

    def __init__(self, archivum_trivio: str) -> None:
        """__init__

        Args:
            archivum_trivio (str):
        """
        # from openpyxl import load_workbook
        # print(archivum_trivio)
        self.archivum_trivio = archivum_trivio
        self.workbook = load_workbook(
            archivum_trivio, data_only=True, read_only=True)
        # self.workbook = load_workbook(
        #     archivum_trivio, data_only=True)

        self.workbook.iso_dates = True
        self.sheet_active = None
        # pass

    def de(self, worksheet_reference: str = None) -> bool:
        if isinstance(worksheet_reference, str) and \
                len(worksheet_reference) == 1:
            meta = self.meta()
            if worksheet_reference in meta['sheet_cod_ab']:
                self.active = meta['sheet_cod_ab'][worksheet_reference]
                return True
        if worksheet_reference in self.workbook:
            self.active = worksheet_reference
            return True
        return False

    def meta(self) -> dict:
        """meta

        Returns:
            dict: _description_
        """
        resultatum = {
            # '_': self.workbook,
            'sheetnames': self.workbook.sheetnames,
            'sheet': {},
            'sheet_active': {},
            'sheet_cod_ab': {},
            'cod_ab_level': [],
        }
        for item in self.workbook.sheetnames:
            resultatum['sheet'][item] = {
                # '__': self.workbook[item],
                # 'max_col': self.workbook[item].max_col
                'max_column': self.workbook[item].max_column,
                'max_row': self.workbook[item].max_row
            }
            _item_num = re.sub('[^0-9]', '', item)
            if len(_item_num) == 1:
                #_likely_ab = _item_num
                resultatum['sheet_cod_ab'][_item_num] = item
                resultatum['cod_ab_level'].append(int(_item_num))
            if len(resultatum['cod_ab_level']) > 0:
                resultatum['cod_ab_level'] = sorted(resultatum['cod_ab_level'])
        # resultatum = wb2
        # print(wb2.keys())
        # workbook.close()
        return resultatum

    def finis(self) -> None:
        """finis Close XLSX file immediately to save memory
        """
        # https://en.wiktionary.org/wiki/finis#Latin
        self.workbook.close()

    def imprimere(self, formatum: str = None) -> list:
        """imprimere /print/@eng-Latn

        Trivia:
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - fōrmātum, s, n, nominativus, https://en.wiktionary.org/wiki/formatus

        Args:
            formatum (str, optional): output format. Defaults to 'csv'.

        Returns:
            [list]: data, caput
        """
        # pylint: disable=chained-comparison
        caput = []
        data = []
        if formatum:
            self._formatum = formatum

        fons = self.workbook[self.active]

        for row_index, row in enumerate(fons.rows):
            # print(row_index, row, self.active_row_end)
            if row_index >= self.active_row_start and \
                    row_index <= self.active_row_end:
                linea = []
                for col_index in range(
                        self.active_col_start, self.active_col_end):
                    textum = row[col_index].value

                    if textum:
                        # if isinstance(textum, datetime.datetime):
                        if isinstance(textum, datetime):
                            textum = str(textum).replace(' 00:00:00', '')
                        linea.append(textum)
                    else:
                        linea.append('')

                if len(caput) == 0:
                    caput = linea
                else:
                    data.append(linea)
            # a_dict[row[0].value+','+row[1].value].append(str(row[2].value))

        # return data, caput
        return caput, data

    def praeparatio(self):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """

        if self.workbook[self.active].max_column is None or \
                self.workbook[self.active].max_row is None:
            # print('Excel was saved without metas; opening without read-only')
            self.workbook = load_workbook(
                self.archivum_trivio, data_only=True, read_only=False)

            self.workbook.iso_dates = True
            self.sheet_active = None

        self.active_col_end = self.workbook[self.active].max_column
        self.active_row_end = self.workbook[self.active].max_row

        # fons = self.workbook[self.active]

        # for row_index, row in enumerate(fons.rows):
        #     # TODO: discover where it starts
        #     pass

        # for index_cols in fons.rows

        # @TODO: calculate if HXL, if have more columns than ideal, etc

        return True
